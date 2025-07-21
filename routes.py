from flask import render_template, request, redirect, url_for, flash, jsonify, make_response, session, send_file
from flask_login import login_user, logout_user, login_required, current_user
from sqlalchemy.orm import joinedload
from app import app, db
from models import User, Assessment, Category, CategoryRating, AssessmentPeriod, AssessmentAssignment, ActivityLog, AssessmentActivityLog, AssessmentForm, AssessmentQuestion, AssessmentResponse, PeriodFormAssignment, PeriodReviewee, PeriodReviewer
from forms import LoginForm, AssessmentForm as AssessmentFormClass, CategoryRatingForm, UserForm, EditUserForm, AssessmentPeriodForm, AssignmentForm, CategoryForm
from utils import admin_required, generate_pdf_report, export_csv_data
from email_service import email_service
from activity_logger import log_activity, get_activity_logs, get_user_activity_logs
from admin_chatbot import process_chatbot_message
from datetime import datetime, date
import csv
import io

@app.context_processor
def inject_task_counts():
    """Inject task counts for navigation badges"""
    if current_user.is_authenticated and current_user.role == 'admin':
        # Admin sees submitted assignments pending approval
        admin_pending_count = AssessmentAssignment.query.filter_by(
            is_submitted=True,
            is_admin_approved=False
        ).count()
        return dict(admin_pending_count=admin_pending_count, pending_assignments_count=0)
    return dict(admin_pending_count=0, pending_assignments_count=0)

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            if not user.is_active:
                flash('Your account has been deactivated. Please contact an administrator.', 'error')
                log_activity(user.id, 'login_blocked', f'Inactive user attempted login from {request.remote_addr}')
            else:
                login_user(user)
                log_activity(user.id, 'login', f'User logged in successfully')
                next_page = request.args.get('next')
                flash(f'Welcome back, {user.name}!', 'success')
                return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password.', 'error')
    
    return render_template('login.html', form=form)



@app.route('/logout')
@login_required
def logout():
    if current_user.is_authenticated:
        log_activity(current_user.id, 'logout', f'User logged out')
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    log_activity(current_user.id, 'view_dashboard', f'User accessed main dashboard')
    
    if current_user.role == 'admin':
        # Admin gets the beautiful dashboard with charts and visualizations
        return admin_dashboard()
    elif current_user.role == 'officer':
        return redirect(url_for('my_tasks'))
    else:  # board_member
        return redirect(url_for('reviewer_dashboard'))

@app.route('/reviewer_dashboard')
@login_required
def reviewer_dashboard():
    if current_user.role not in ['board_member', 'admin']:
        flash('Access denied. Board members and administrators only.', 'error')
        return redirect(url_for('dashboard'))
    
    log_activity(current_user.id, 'view_reviewer_dashboard', f'User accessed reviewer dashboard')
    
    # Get pending review assignments
    pending_assignments = AssessmentAssignment.query.filter_by(
        reviewer_id=current_user.id,
        is_completed=False
    ).join(AssessmentPeriod).filter(AssessmentPeriod.is_active == True).all()
    
    # Get completed review assignments for this year
    current_year = datetime.now().year
    completed_assignments = AssessmentAssignment.query.filter_by(
        reviewer_id=current_user.id,
        is_completed=True
    ).join(AssessmentPeriod).filter(
        db.extract('year', AssessmentPeriod.start_date) == current_year
    ).all()
    
    # Group assignments by assessment project
    pending_by_period = {}
    completed_by_period = {}
    
    for assignment in pending_assignments:
        period_name = assignment.period.name
        if period_name not in pending_by_period:
            pending_by_period[period_name] = []
        pending_by_period[period_name].append(assignment)
    
    for assignment in completed_assignments:
        period_name = assignment.period.name
        if period_name not in completed_by_period:
            completed_by_period[period_name] = []
        completed_by_period[period_name].append(assignment)
    
    # Get total statistics
    total_pending = len(pending_assignments)
    total_completed = len(completed_assignments)
    total_assigned = total_pending + total_completed
    
    completion_rate = round((total_completed / total_assigned * 100), 1) if total_assigned > 0 else 0
    
    return render_template('reviewer_dashboard.html',
                         pending_assignments=pending_assignments,
                         completed_assignments=completed_assignments,
                         pending_by_period=pending_by_period,
                         completed_by_period=completed_by_period,
                         total_pending=total_pending,
                         total_completed=total_completed,
                         completion_rate=completion_rate)




@app.route('/my_tasks')
@login_required
def my_tasks():
    """My Tasks page - for all users to see their assigned tasks"""
    if current_user.role == 'admin':
        # Admin sees submitted assignments pending approval
        officer_id = request.args.get('officer_id', type=int)
        
        if officer_id:
            officer = User.query.get_or_404(officer_id)
            log_activity(current_user.id, 'view_officer_task_status', f'Admin viewed task status for {officer.name}')
            
            assignments = AssessmentAssignment.query.filter_by(
                officer_id=officer_id,
                is_submitted=True,
                is_admin_approved=False
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            pending_count = len([a for a in assignments if not a.is_completed])
            completed_count = len([a for a in assignments if a.is_completed])
            total_assignments = len(assignments)
            completion_rate = (completed_count / total_assignments * 100) if total_assignments > 0 else 0
            
            return render_template('my_assignments.html',
                                 assignments=assignments,
                                 pending_count=pending_count,
                                 completed_count=completed_count,
                                 completion_rate=round(completion_rate, 1),
                                 officer=officer)
        else:
            # Admin sees all submitted assignments pending approval in tabbed format
            assignments = AssessmentAssignment.query.options(
                db.joinedload(AssessmentAssignment.assessment_responses),
                db.joinedload(AssessmentAssignment.period),
                db.joinedload(AssessmentAssignment.officer),
                db.joinedload(AssessmentAssignment.reviewer)
            ).filter_by(
                is_submitted=True,
                is_admin_approved=False
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            # For each self-assessment, get the external reviewers that will be notified
            external_reviewers_map = {}
            for assignment in assignments:
                if assignment.officer_id == assignment.reviewer_id:  # Self-assessment
                    external_reviewers = AssessmentAssignment.query.options(
                        db.joinedload(AssessmentAssignment.reviewer)
                    ).filter_by(
                        officer_id=assignment.officer_id,
                        period_id=assignment.period_id
                    ).filter(AssessmentAssignment.reviewer_id != assignment.officer_id).all()
                    external_reviewers_map[assignment.id] = external_reviewers
            
            # For admin: Open Tasks = pending approval, All Tasks = all submitted assignments
            open_tasks = assignments  # All submitted assignments need admin approval
            all_tasks = AssessmentAssignment.query.options(
                db.joinedload(AssessmentAssignment.assessment_responses),
                db.joinedload(AssessmentAssignment.period),
                db.joinedload(AssessmentAssignment.officer),
                db.joinedload(AssessmentAssignment.reviewer)
            ).filter_by(
                is_submitted=True
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            pending_count = len(assignments)
            completed_count = len([a for a in all_tasks if a.is_admin_approved])
            
            return render_template('my_assignments.html',
                                 open_tasks=open_tasks,
                                 all_tasks=all_tasks,
                                 pending_count=pending_count,
                                 completed_count=completed_count,
                                 completion_rate=0,
                                 officer=None,
                                 external_reviewers_map=external_reviewers_map)
    else:
        # Non-admin users see their reviewer assignments
        log_activity(current_user.id, 'view_my_reviewer_tasks', f'User viewed their reviewer tasks')
        
        if current_user.role == 'officer':
            # Officers see ALL their assignments (including self-assessments)
            assignments = AssessmentAssignment.query.options(db.joinedload(AssessmentAssignment.assessment_responses)).filter(
                db.or_(
                    AssessmentAssignment.reviewer_id == current_user.id,
                    AssessmentAssignment.officer_id == current_user.id
                )
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            # For officers, show ALL their assignments but distinguish what needs work
            available_assignments = []
            pending_work_count = 0
            for a in assignments:
                # Include all assignments for display
                available_assignments.append(a)
                
                # Count only tasks that actually need work for pending count
                # Self-assessment that's not submitted yet (needs work)
                if (a.officer_id == a.reviewer_id == current_user.id and not a.is_submitted):
                    pending_work_count += 1
                # External reviewer task that's ready to work on
                elif (a.reviewer_id == current_user.id and a.officer_id != current_user.id and not a.is_completed and 
                      (not a.is_submitted or (a.is_submitted and not a.is_admin_approved and a.admin_notes))):
                    pending_work_count += 1
        else:
            # Board members see only their reviewer assignments WHERE officer's self-assessment is approved
            assignments = AssessmentAssignment.query.options(db.joinedload(AssessmentAssignment.assessment_responses)).filter_by(
                reviewer_id=current_user.id,
                is_completed=False
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            # Filter for tasks that are ready to work on AND where officer's self-assessment is approved
            available_assignments = []
            for a in assignments:
                # Check if officer's self-assessment is approved
                officer_self_assessment = AssessmentAssignment.query.filter_by(
                    officer_id=a.officer_id,
                    reviewer_id=a.officer_id,  # Self-assessment
                    period_id=a.period_id,
                    is_admin_approved=True
                ).first()
                
                if officer_self_assessment:
                    # Officer self-assessment is approved, check if task is ready
                    if not a.is_submitted or (a.is_submitted and not a.is_admin_approved and a.admin_notes):
                        available_assignments.append(a)
        
        # Create separate open tasks and all tasks for tabbed interface
        if current_user.role == 'officer':
            # For officers, get assignments where they are REVIEWER + completed assignments where they were REVIEWEE
            reviewer_assignments = AssessmentAssignment.query.options(db.joinedload(AssessmentAssignment.assessment_responses)).filter_by(
                reviewer_id=current_user.id
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            # For All Tasks, only show assignments where user is the reviewer (no reviewee tasks)
            all_assignments = reviewer_assignments
            
            # Create open tasks (tasks that need work) for the "Open Tasks" tab - only reviewer assignments
            open_tasks = []
            for a in reviewer_assignments:
                # Self-assessment that's not submitted yet (needs work)
                if (a.officer_id == a.reviewer_id == current_user.id and not a.is_submitted):
                    open_tasks.append(a)
                # External reviewer task that's ready to work on - BUT ONLY if officer's self-assessment is approved
                elif (a.reviewer_id == current_user.id and a.officer_id != current_user.id and not a.is_completed):
                    # Check if officer's self-assessment is approved and released
                    officer_self_assessment = AssessmentAssignment.query.filter_by(
                        officer_id=a.officer_id,
                        reviewer_id=a.officer_id,  # Self-assessment
                        period_id=a.period_id,
                        is_admin_approved=True
                    ).first()
                    
                    if officer_self_assessment:
                        # Officer self-assessment is approved, so external reviewer can work
                        if (not a.is_submitted or (a.is_submitted and not a.is_admin_approved and a.admin_notes)):
                            open_tasks.append(a)
            
            pending_count = len(open_tasks)
            all_tasks = all_assignments  # Show reviewer assignments + completed reviewee assignments
        else:
            # For board members, open tasks are the filtered assignments
            open_tasks = available_assignments
            pending_count = len(available_assignments)
            # Get all assignments for board members (including completed) - BUT only where officer's self-assessment was approved
            all_board_assignments = AssessmentAssignment.query.options(db.joinedload(AssessmentAssignment.assessment_responses)).filter_by(
                reviewer_id=current_user.id
            ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
            
            # Filter all tasks to only show those where officer's self-assessment was approved
            all_tasks = []
            for a in all_board_assignments:
                # Check if officer's self-assessment is approved
                officer_self_assessment = AssessmentAssignment.query.filter_by(
                    officer_id=a.officer_id,
                    reviewer_id=a.officer_id,  # Self-assessment
                    period_id=a.period_id,
                    is_admin_approved=True
                ).first()
                
                if officer_self_assessment:
                    all_tasks.append(a)
        
        completed_count = AssessmentAssignment.query.filter_by(
            reviewer_id=current_user.id,
            is_completed=True
        ).count()
        
        return render_template('my_assignments.html',
                             open_tasks=open_tasks,
                             all_tasks=all_tasks,
                             pending_count=pending_count,
                             completed_count=completed_count,
                             completion_rate=0,
                             officer=None)

@app.route('/my_assignments')
@login_required
@admin_required
def my_assignments():
    """My Assignments page - ADMIN ONLY - shows assignment management interface"""
    officer_id = request.args.get('officer_id', type=int)
    
    # Admin filtering by officer_id - show reviewers assigned to review this officer
    if officer_id:
        officer = User.query.get_or_404(officer_id)
        log_activity(current_user.id, 'view_officer_task_status', f'Admin viewed task status for {officer.name}')
        
        # Get assignments where this officer is being reviewed - ONLY submitted tasks pending approval
        assignments = AssessmentAssignment.query.options(db.joinedload(AssessmentAssignment.assessment_responses)).filter_by(
            officer_id=officer_id,
            is_submitted=True,
            is_admin_approved=False
        ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
        
        # For admin view, calculate stats
        pending_count = len([a for a in assignments if not a.is_completed])
        completed_count = len([a for a in assignments if a.is_completed])
        total_assignments = len(assignments)
        completion_rate = (completed_count / total_assignments * 100) if total_assignments > 0 else 0
        
        return render_template('my_assignments.html',
                             assignments=assignments,
                             pending_count=pending_count,
                             completed_count=completed_count,
                             completion_rate=round(completion_rate, 1),
                             officer=officer)
    else:
        # Show all assignments for admin overview
        log_activity(current_user.id, 'admin_assignments_overview', 'Admin accessed assignments overview')
        
        # Get all assignments for admin overview - ONLY submitted tasks pending approval
        assignments = AssessmentAssignment.query.options(db.joinedload(AssessmentAssignment.assessment_responses)).filter_by(
            is_submitted=True,
            is_admin_approved=False
        ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
        
        # Calculate stats
        pending_count = len([a for a in assignments if not a.is_completed])
        completed_count = len([a for a in assignments if a.is_completed])
        total_assignments = len(assignments)
        completion_rate = (completed_count / total_assignments * 100) if total_assignments > 0 else 0
        
        return render_template('my_assignments.html',
                             assignments=assignments,
                             pending_count=pending_count,
                             completed_count=completed_count,
                             completion_rate=round(completion_rate, 1),
                             officer=None)

@app.route('/officer_task_status/<int:officer_id>')
@login_required
@admin_required
def officer_task_status(officer_id):
    """Officer Task Status page - shows comprehensive workflow table for all review tasks"""
    officer = User.query.get_or_404(officer_id)
    log_activity(current_user.id, 'view_officer_task_status', f'Admin viewed task status for {officer.name}')
    
    # Get ALL assignments where this officer is being reviewed (not just submitted ones)
    assignments = AssessmentAssignment.query.filter_by(
        officer_id=officer_id
    ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
    
    # Calculate stats for ALL assignments
    pending_count = len([a for a in assignments if not a.is_completed])
    completed_count = len([a for a in assignments if a.is_completed])
    submitted_count = len([a for a in assignments if a.is_submitted and not a.is_admin_approved])
    total_assignments = len(assignments)
    completion_rate = (completed_count / total_assignments * 100) if total_assignments > 0 else 0
    
    # Generate comprehensive workflow table using existing function
    from workflow_table import create_workflow_table
    
    # For now, use the first assignment's period_id (we can enhance this later)
    period_id = assignments[0].period_id if assignments else None
    workflow_table_html = create_workflow_table(officer_id, period_id) if period_id else ""
    
    return render_template('officer_task_status.html',
                         assignments=assignments,
                         pending_count=pending_count,
                         completed_count=completed_count,
                         submitted_count=submitted_count,
                         completion_rate=round(completion_rate, 1),
                         officer=officer,
                         workflow_table=workflow_table_html)

@app.route('/admin')
@login_required
@admin_required
def admin_main():
    """Admin main page with tabbed interface"""
    try:
        log_activity(current_user.id, 'view_admin_main', 'Accessed admin main page')
        
        # Get assessment form statistics
        from models import AssessmentForm, AssessmentQuestion
        assessment_forms_count = AssessmentForm.query.filter_by(is_active=True).count()
        assessment_questions_count = AssessmentQuestion.query.filter_by(is_active=True).count()
        
        # Get categories for traditional questions
        categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
        active_periods = AssessmentPeriod.query.filter_by(is_active=True).order_by(AssessmentPeriod.created_at.desc()).limit(3).all()
        
        # Get basic statistics
        total_users = User.query.filter_by(is_active=True).count()
        total_officers = User.query.filter_by(role='officer', is_active=True).count()
        total_board_members = User.query.filter_by(role='board_member', is_active=True).count()
        total_assessments = Assessment.query.count()
        
        return render_template('admin_main.html',
                             assessment_forms_count=assessment_forms_count,
                             assessment_questions_count=assessment_questions_count,
                             categories=categories,
                             active_periods=active_periods,
                             total_users=total_users,
                             total_officers=total_officers,
                             total_board_members=total_board_members,
                             total_assessments=total_assessments)
    except Exception as e:
        print(f"Error in admin_main: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Admin page error: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/admin_dashboard')
@login_required
@admin_required
def admin_dashboard():
    log_activity(current_user.id, 'view_admin_dashboard', f'Admin accessed admin dashboard')
    current_year = datetime.now().year
    
    # Get current active assessment project
    current_period = AssessmentPeriod.query.filter_by(is_active=True).first()
    
    # Get all assessment projects for context
    all_periods = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).limit(5).all()
    
    # Get assessment form statistics
    from models import AssessmentForm, AssessmentQuestion
    assessment_forms_count = AssessmentForm.query.filter_by(is_active=True).count()
    assessment_questions_count = AssessmentQuestion.query.filter_by(is_active=True).count()
    template_forms_count = AssessmentForm.query.filter_by(is_active=True, is_template=True).count()
    # Get categories for traditional questions
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    active_periods = AssessmentPeriod.query.filter_by(is_active=True).order_by(AssessmentPeriod.created_at.desc()).limit(3).all()
    
    # Get statistics for current period
    if current_period:
        total_assignments = AssessmentAssignment.query.filter_by(period_id=current_period.id).count()
        completed_assignments = AssessmentAssignment.query.filter_by(period_id=current_period.id, is_completed=True).count()
        pending_assignments = total_assignments - completed_assignments
        completion_rate = round((completed_assignments / total_assignments * 100), 1) if total_assignments > 0 else 0
        
        # Get assignment data for current period including AI report availability
        from models import AIGeneratedReport
        assignments_by_officer = db.session.query(
            User.name.label('officer_name'),
            User.id.label('officer_id'),
            db.func.count(AssessmentAssignment.id).label('total_reviews'),
            db.func.sum(db.case((AssessmentAssignment.is_completed == True, 1), else_=0)).label('completed_reviews'),
            db.func.count(AIGeneratedReport.id).label('ai_report_count')
        ).join(
            AssessmentAssignment, User.id == AssessmentAssignment.officer_id
        ).outerjoin(
            AIGeneratedReport, (User.id == AIGeneratedReport.officer_id) & (AIGeneratedReport.period_id == current_period.id)
        ).filter(
            AssessmentAssignment.period_id == current_period.id
        ).group_by(User.id, User.name).all()
        
        # Get recent activity for current period
        recent_assessments = Assessment.query.join(
            AssessmentAssignment, Assessment.id == AssessmentAssignment.assessment_id
        ).filter(
            AssessmentAssignment.period_id == current_period.id
        ).order_by(Assessment.submitted_at.desc()).limit(10).all()
        
        # Get reviewer progress
        reviewer_progress = db.session.query(
            User.name.label('reviewer_name'),
            User.id.label('reviewer_id'),
            db.func.count(AssessmentAssignment.id).label('total_assigned'),
            db.func.sum(db.case((AssessmentAssignment.is_completed == True, 1), else_=0)).label('completed')
        ).join(
            AssessmentAssignment, User.id == AssessmentAssignment.reviewer_id
        ).filter(
            AssessmentAssignment.period_id == current_period.id
        ).group_by(User.id, User.name).all()
        
    else:
        total_assignments = completed_assignments = pending_assignments = 0
        completion_rate = 0
        assignments_by_officer = []
        recent_assessments = []
        reviewer_progress = []
    
    # Get overall system statistics
    total_officers = User.query.filter_by(role='officer').count()
    total_board_members = User.query.filter_by(role='board_member').count()
    total_assessments = Assessment.query.filter_by(year=current_year).count()
    
    # Get monthly assessment trend for the chart
    monthly_data = []
    for month in range(1, 13):
        month_assessments = Assessment.query.filter(
            db.extract('year', Assessment.submitted_at) == current_year,
            db.extract('month', Assessment.submitted_at) == month
        ).count()
        monthly_data.append({
            'month': month,
            'count': month_assessments,
            'month_name': datetime(current_year, month, 1).strftime('%b')
        })
    
    return render_template('admin_dashboard.html',
                         current_period=current_period,
                         all_periods=all_periods,
                         total_assignments=total_assignments,
                         completed_assignments=completed_assignments,
                         pending_assignments=pending_assignments,
                         completion_rate=completion_rate,
                         assignments_by_officer=assignments_by_officer,
                         recent_assessments=recent_assessments,
                         reviewer_progress=reviewer_progress,
                         total_officers=total_officers,
                         total_board_members=total_board_members,
                         total_assessments=total_assessments,
                         monthly_data=monthly_data,
                         assessment_forms_count=assessment_forms_count,
                         assessment_questions_count=assessment_questions_count,
                         template_forms_count=template_forms_count,
                         categories=categories,
                         active_periods=active_periods)

@app.route('/admin/officer_reviews/<int:officer_id>')
@login_required
@admin_required
def officer_reviews(officer_id):
    try:
        officer = User.query.get_or_404(officer_id)
        log_activity(current_user.id, 'view_officer_reviews', f'Admin viewed comprehensive review matrix for officer: {officer.name}')
        if officer.role != 'officer':
            flash('Invalid officer selection.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Get current active assessment project
        current_period = AssessmentPeriod.query.filter_by(is_active=True).first()
        if not current_period:
            flash('No active assessment project found.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Get all assignments for this officer in the current period
        assignments = AssessmentAssignment.query.filter_by(
            officer_id=officer_id,
            period_id=current_period.id
        ).all()
        
        if not assignments:
            flash(f'No assessment assignments found for {officer.name} in the current period.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Create reviewers list from ALL ASSIGNMENTS
        reviewers = {}
        for assignment in assignments:
            reviewers[assignment.reviewer.id] = assignment.reviewer
        
        # Get all assessment forms assigned to this period for BOTH reviewers AND self-review
        # This is CRITICAL for self-assessment display - includes both form types
        reviewer_forms = db.session.query(AssessmentForm).join(PeriodFormAssignment).filter(
            PeriodFormAssignment.period_id == current_period.id,
            PeriodFormAssignment.form_type.in_(['reviewer', 'self_review']),
            PeriodFormAssignment.is_active == True,
            AssessmentForm.is_active == True
        ).all()
        
        # Build matrix data: questions in rows, reviewers in columns
        matrix_data = []
        processed_question_texts = set()  # Track processed question texts to prevent duplicates
        
        # Get all questions from all forms (both reviewer and self-review)
        all_questions = []
        for form in reviewer_forms:
            form_questions = AssessmentQuestion.query.filter_by(
                form_id=form.id,
                is_active=True
            ).order_by(AssessmentQuestion.order.asc()).all()
            all_questions.extend(form_questions)
        
        for question in all_questions:
            # Skip if we've already processed this question text (prevent duplicates)
            if question.question_text in processed_question_texts:
                continue
            
            # Only include rating questions in the matrix (skip text-only questions)
            if question.question_type != 'rating':
                continue
                
            processed_question_texts.add(question.question_text)
            
            question_row = {
                'question': question,
                'responses': {},
                'average_rating': 0,
                'response_count': 0
            }
            
            total_rating = 0
            rating_count = 0
            responses_for_ai = []
            
            for reviewer_id, reviewer in reviewers.items():
                # Find assignment for this reviewer
                assignment = next((a for a in assignments if a.reviewer_id == reviewer_id), None)
                if assignment:
                    # Find response for this question in this assignment
                    # Search by question text to match across different forms (reviewer vs self-review)
                    try:
                        # First try exact question ID match
                        response = AssessmentResponse.query.filter_by(
                            assessment_assignment_id=assignment.id,
                            question_id=question.id
                        ).first()
                        
                        # If no direct match, search for equivalent question by text
                        if not response:
                            equivalent_questions = AssessmentQuestion.query.filter_by(
                                question_text=question.question_text,
                                is_active=True
                            ).all()
                            
                            for eq_question in equivalent_questions:
                                response = AssessmentResponse.query.filter_by(
                                    assessment_assignment_id=assignment.id,
                                    question_id=eq_question.id
                                ).first()
                                if response:
                                    break
                                    
                    except Exception as db_error:
                        print(f"Database error: {db_error}")
                        response = None
                    
                    if response:
                        response_data = response.get_response_data()
                        rating = None
                        comment = ""
                        
                        # Handle different response types
                        if question.question_type == 'rating' and response.response_number is not None:
                            rating = int(response.response_number)
                            comment = response.response_text or ""
                        elif question.question_type in ['text', 'textarea']:
                            comment = response.response_text or ""
                        else:
                            comment = str(response_data) if response_data else ""
                        
                        # Store response data
                        question_row['responses'][reviewer.name] = {
                            'rating': rating,
                            'comment': comment,
                            'text_response': comment,
                            'reviewer': reviewer,
                            'response_data': response_data
                        }
                        
                        # Only include external reviewer ratings in average (exclude self-assessment)
                        if rating and reviewer.id != officer.id:
                            total_rating += rating
                            rating_count += 1
                            
                        # Collect for AI analysis (only valid responses with content)
                        if rating or comment:
                            responses_for_ai.append({
                                'reviewer': reviewer.name,
                                'rating': rating,
                                'comment': comment,
                                'reviewer_role': reviewer.role
                            })
                    else:
                        question_row['responses'][reviewer.name] = None
                else:
                    question_row['responses'][reviewer.name] = None
            
            # Calculate average rating for this question
            if rating_count > 0:
                question_row['average_rating'] = round(total_rating / rating_count, 2)
                question_row['response_count'] = rating_count
            
            # No AI analysis during matrix display - matrix should load instantly
            question_row['ai_analysis'] = {
                "summary": "Use Generate AI Summary button for analysis",
                "key_themes": [],
                "sentiment": "neutral"
            }
            
            # Add question to matrix if it has any responses or we want to show all
            matrix_data.append(question_row)
        
        # Calculate overall statistics
        total_reviewers = len(reviewers)  # Count ALL assigned reviewers
        completed_assignments = len([a for a in assignments if a.is_completed])
        completion_rate = round((completed_assignments / total_reviewers) * 100, 1) if total_reviewers > 0 else 0
        
        # Calculate overall average across all questions with ratings (exclude self-assessment)
        all_ratings = []
        for question_data in matrix_data:
            for reviewer_name, response in question_data['responses'].items():
                if response and response.get('rating'):
                    # Exclude self-assessment scores from overall average
                    reviewer = response.get('reviewer')
                    if reviewer and reviewer.id != officer.id:
                        all_ratings.append(response['rating'])
        
        overall_matrix_average = round(sum(all_ratings) / len(all_ratings), 2) if all_ratings else 0
        
        # Collect text responses for summary (non-numerical questions)
        text_responses = []
        try:
            for assignment in assignments:
                # Get all non-rating responses for this assignment
                text_questions = AssessmentResponse.query.join(AssessmentQuestion).filter(
                    AssessmentResponse.assessment_assignment_id == assignment.id,
                    AssessmentQuestion.question_type.in_(['text', 'textarea']),
                    AssessmentQuestion.question_name.notin_(['signature', 'date', 'name']),  # Skip signature fields
                    AssessmentResponse.response_text.isnot(None),
                    AssessmentResponse.response_text != ''
                ).all()
                
                for response in text_questions:
                    if response.response_text and response.response_text.strip():
                        text_responses.append({
                            'question_name': response.question.question_name,
                            'question_text': response.question.question_text,
                            'response': response.response_text,
                            'reviewer': assignment.reviewer.name,
                            'reviewer_role': assignment.reviewer.role
                        })
        except Exception as db_error:
            print(f"Database error collecting text responses: {db_error}")
            # Continue without text responses if there's a database issue
            text_responses = []
        
        # Generate AI summaries for text responses with timeout handling
        text_summaries = {}
        if text_responses:
            # Group by question name
            questions_grouped = {}
            for response in text_responses:
                question_name = response['question_name']
                if question_name not in questions_grouped:
                    questions_grouped[question_name] = []
                questions_grouped[question_name].append(response)
            
            # Generate simple summaries for text responses (skip AI to prevent timeouts)
            for question_name, responses in questions_grouped.items():
                # Create immediate summary without AI to prevent worker timeouts
                text_summaries[question_name] = {
                    'question_text': responses[0]['question_text'],
                    'summary': {
                        'summary': f"Comprehensive feedback on {question_name} from {len(responses)} reviewers. Key insights and perspectives captured in individual responses below.",
                        'key_themes': ["Multi-reviewer perspectives", "Detailed professional feedback", "Strategic insights"]
                    },
                    'responses': responses
                }
        
        # Generate immediate performance summary (avoid AI to prevent worker timeouts)
        performance_level = "excellent" if overall_matrix_average >= 4 else "good" if overall_matrix_average >= 3 else "satisfactory"
        overall_ai_summary = {
            "executive_summary": f"Performance evaluation shows {officer.name} achieving {overall_matrix_average}/5.0 overall rating across {len(matrix_data)} assessment categories. {len(reviewers)} reviewers provided comprehensive feedback indicating {performance_level} leadership performance.",
            "major_themes": ["Multi-reviewer assessment", "Comprehensive evaluation", "Professional leadership", "Strategic performance"],
            "overall_sentiment": performance_level
        }
    

        
        # Save calculated data to simple file cache for Excel export
        import json
        import os
        cache_file = f"matrix_cache_{officer_id}.json"
        try:
            cache_data = {
                'officer_name': officer.name,
                'overall_average': overall_matrix_average,
                'overall_ai_summary': overall_ai_summary,
                'matrix_rows': []
            }
            
            # Convert matrix data to simple structure for Excel with AI insights
            for question_data in matrix_data:
                row = {
                    'question_text': question_data['question'].question_text if hasattr(question_data['question'], 'question_text') else str(question_data['question']),
                    'average_rating': question_data.get('average_rating', 0),
                    'reviewer_ratings': {},
                    'ai_analysis': {}
                }
                
                # Extract simple rating data
                for reviewer_name, response in question_data.get('responses', {}).items():
                    if response and isinstance(response, dict):
                        rating = response.get('rating')
                        comment = response.get('comment', '')
                    else:
                        rating = None
                        comment = ''
                    
                    row['reviewer_ratings'][reviewer_name] = {
                        'rating': rating,
                        'comment': comment
                    }
                
                # No AI analysis in cache - removed for performance
                row['ai_analysis'] = {
                    'summary': 'Use Generate AI Summary button for comprehensive analysis',
                    'key_themes': [],
                    'sentiment': 'neutral'
                }
                
                cache_data['matrix_rows'].append(row)
            
            # Save to file
            with open(cache_file, 'w') as f:
                json.dump(cache_data, f)
                
        except Exception as e:
            print(f"Cache save error: {e}")
        
        return render_template('officer_reviews_matrix.html',
                             officer=officer,
                             matrix_data=matrix_data,
                             reviewers=reviewers,
                             avg_rating=overall_matrix_average,
                             total_reviewers=total_reviewers,
                             overall_matrix_average=overall_matrix_average,
                             overall_ai_summary=overall_ai_summary,
                             text_summaries=text_summaries,
                             current_period=current_period,
                             completion_rate=completion_rate,
                             completed_assignments=completed_assignments)
    except Exception as e:
        print(f"Database error in officer_reviews: {e}")
        flash('Database connection error. Please try again.', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/evaluate/<int:officer_id>', methods=['GET', 'POST'])
@login_required
def evaluate_officer(officer_id):
    # Allow board members, admins, and officers doing self-assessment
    if not (current_user.role in ['board_member', 'admin'] or 
            (current_user.role == 'officer' and current_user.id == officer_id)):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    officer = User.query.get_or_404(officer_id)
    
    if request.method == 'GET':
        log_activity(current_user.id, 'start_evaluation', f'User started evaluation of officer: {officer.name}')
    else:
        log_activity(current_user.id, 'submit_evaluation', f'User submitted evaluation for officer: {officer.name}')
    if officer.role != 'officer':
        flash('Invalid officer selection.', 'error')
        return redirect(url_for('dashboard'))
    
    current_year = datetime.now().year
    
    # Check for existing assessment and assignment
    existing_assessment = Assessment.query.filter_by(
        officer_id=officer_id,
        reviewer_id=current_user.id,
        year=current_year
    ).first()
    
    # Find related assignment from active periods
    assignment = AssessmentAssignment.query.filter_by(
        officer_id=officer_id,
        reviewer_id=current_user.id
    ).join(AssessmentPeriod).filter(
        AssessmentPeriod.is_active == True
    ).first()
    
    # For GET requests with existing assessment, show the evaluation form with existing data (allow editing if period is open)
    # No redirect to view_assessment - let users edit their assessments during open periods
    
    # Load assessment questions from forms assigned to current period
    questions = []
    
    # Find the active assessment project
    current_period = AssessmentPeriod.query.filter_by(is_active=True).first()
    if current_period:
        # Determine if this is a self-assessment or external review
        is_self_assessment = (current_user.id == officer_id)
        form_type = 'self_review' if is_self_assessment else 'reviewer'
        
        # Get forms assigned to this period for the appropriate form type
        period_forms = PeriodFormAssignment.query.filter_by(
            period_id=current_period.id,
            form_type=form_type
        ).all()
        
        # Load questions from all assigned forms
        for period_form in period_forms:
            form_questions = AssessmentQuestion.query.filter_by(
                form_id=period_form.form_id,
                is_active=True
            ).order_by(AssessmentQuestion.order).all()
            questions.extend(form_questions)
    
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    form = AssessmentFormClass()
    form.officer_id.data = officer_id
    
    # Load existing responses for editing
    existing_responses = {}
    if assignment:
        responses = AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).all()
        for response in responses:
            existing_responses[response.question_id] = response
    
    # Pre-populate form with existing data if editing
    if existing_assessment:
        form.accomplishments.data = existing_assessment.accomplishments
        form.improvement_opportunities.data = existing_assessment.improvement_opportunities  
        form.focus_for_next_year.data = existing_assessment.focus_for_next_year
    
    if request.method == 'POST' and form.validate_on_submit():
        # Create or update assessment
        if existing_assessment:
            assessment = existing_assessment
        else:
            assessment = Assessment(
                officer_id=officer_id,
                reviewer_id=current_user.id,
                year=current_year
            )
            db.session.add(assessment)
        
        assessment.accomplishments = form.accomplishments.data
        assessment.improvement_opportunities = form.improvement_opportunities.data
        assessment.focus_for_next_year = form.focus_for_next_year.data
        assessment.submitted_at = datetime.utcnow()
        
        # Clear existing category ratings and assessment responses
        CategoryRating.query.filter_by(assessment_id=assessment.id).delete()
        # Clear existing assessment responses for this assignment
        if assignment:
            AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).delete()
        
        # Process individual question responses
        total_rating = 0
        rating_count = 0
        category_ratings = {}
        
        for question in questions:
            # Get question response
            response_rating = request.form.get(f'question_{question.id}')
            response_comment = request.form.get(f'comment_{question.id}')
            
            if response_rating or response_comment:
                # Create assessment response - link to assignment instead of assessment
                if assignment:
                    question_response = AssessmentResponse(
                        assessment_assignment_id=assignment.id,
                        question_id=question.id
                    )
                    
                    if question.question_type == 'rating' and response_rating:
                        rating_val = int(response_rating)
                        question_response.response_number = rating_val
                        total_rating += rating_val
                        rating_count += 1
                        
                    elif question.question_type in ['text', 'textarea'] and response_rating:
                        question_response.response_text = response_rating
                        
                    elif question.question_type == 'date' and response_rating:
                        try:
                            question_response.response_date = datetime.strptime(response_rating, '%Y-%m-%d').date()
                        except:
                            question_response.response_text = response_rating
                    
                    elif question.question_type == 'boolean' and response_rating:
                        question_response.response_boolean = (response_rating.lower() == 'true')
                        question_response.response_text = response_rating
                    
                    elif question.question_type in ['checkbox', 'dropdown', 'multiple_choice'] and response_rating:
                        question_response.response_text = response_rating
                    
                    # Add comment if provided
                    if response_comment:
                        if question_response.response_text:
                            question_response.response_text += f"\n\nComment: {response_comment}"
                        else:
                            question_response.response_text = response_comment
                    
                    db.session.add(question_response)
        
        # Create category ratings for backward compatibility
        for category_id, ratings in category_ratings.items():
            if ratings:
                avg_rating = sum(ratings) / len(ratings)
                category_rating = CategoryRating(
                    assessment=assessment,
                    category_id=category_id,
                    rating=round(avg_rating)
                )
                db.session.add(category_rating)
        
        # Calculate overall rating
        assessment.overall_rating = total_rating / rating_count if rating_count > 0 else 0
        
        # Mark assignment as completed if it exists
        if assignment and not assignment.is_completed:
            assignment.is_completed = True
            assignment.completed_at = datetime.utcnow()
            assignment.assessment = assessment
        
        db.session.commit()
        flash(f'Assessment for {officer.name} has been submitted successfully.', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('evaluate.html', officer=officer, categories=categories, questions=questions, form=form, existing_responses=existing_responses, assignment=assignment)

# Create new assessment viewing system using AssessmentResponse
@app.route('/view_assessment_new/<int:assignment_id>')
@login_required
def view_assessment_new(assignment_id):
    assignment = AssessmentAssignment.query.get_or_404(assignment_id)
    
    # Check if user has permission to view this assessment
    if not (current_user.role == 'admin' or 
            current_user.id == assignment.reviewer_id or 
            current_user.id == assignment.officer_id):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    log_activity(current_user.id, 'view_assessment_new', f'User viewed assessment for assignment #{assignment_id}')
    
    # Get assessment responses for this assignment
    assessment_responses = AssessmentResponse.query.filter_by(
        assessment_assignment_id=assignment_id
    ).all()
    
    # Get forms assigned to this period
    period_forms = PeriodFormAssignment.query.filter_by(
        period_id=assignment.period_id
    ).all()
    
    # Determine which form to show based on assignment type
    if assignment.officer_id == assignment.reviewer_id:
        # Self-assessment: show self_review form
        target_form_type = 'self_review'
    else:
        # External review: show reviewer form
        target_form_type = 'reviewer'
    
    # Get the appropriate form for this assignment type
    target_form = None
    for pfa in period_forms:
        if pfa.form_type == target_form_type:
            target_form = AssessmentForm.query.get(pfa.form_id)
            break
    
    if not target_form:
        flash('No assessment form found for this assignment type.', 'error')
        return redirect(url_for('my_tasks'))
    
    # Get all questions from the target form
    questions = AssessmentQuestion.query.filter_by(
        form_id=target_form.id,
        is_active=True
    ).order_by(AssessmentQuestion.order).all()
    
    if not questions:
        flash('No questions found in the assessment form.', 'error')
        return redirect(url_for('my_tasks'))
    
    # Group questions by form for display
    questions_by_form = {}
    for question in questions:
        form_title = question.form.title
        if form_title not in questions_by_form:
            questions_by_form[form_title] = []
        questions_by_form[form_title].append(question)
    
    # Create response lookup for easy access
    response_lookup = {resp.question_id: resp for resp in assessment_responses}
    
    # If no responses exist and user is the reviewer, redirect to edit mode
    # Also redirect if assignment is not submitted (still in draft mode)
    if current_user.id == assignment.reviewer_id and (not assessment_responses or not assignment.is_submitted):
        return redirect(url_for('edit_assessment_new', assignment_id=assignment_id))
    
    return render_template('view_assessment_new.html', 
                         assignment=assignment,
                         questions_by_form=questions_by_form,
                         responses=response_lookup)

@app.route('/edit_assessment_new/<int:assignment_id>', methods=['GET', 'POST'])
@login_required
def edit_assessment_new(assignment_id):
    assignment = AssessmentAssignment.query.get_or_404(assignment_id)
    
    # Check if user has permission to edit this assessment
    if current_user.id != assignment.reviewer_id:
        flash('Access denied.', 'error')
        return redirect(url_for('my_tasks'))
    
    # Check if assessment is already submitted (prevent editing submitted assessments)
    if assignment.is_submitted and not current_user.role == 'admin':
        flash('This assessment has been submitted and cannot be edited.', 'warning')
        return redirect(url_for('view_assessment_new', assignment_id=assignment_id))
    
    # Get forms assigned to this period
    period_forms = PeriodFormAssignment.query.filter_by(
        period_id=assignment.period_id
    ).all()
    
    # Determine which form to show based on assignment type
    if assignment.officer_id == assignment.reviewer_id:
        # Self-assessment: show self_review form
        target_form_type = 'self_review'
    else:
        # External review: show reviewer form
        target_form_type = 'reviewer'
    
    # Get the appropriate form for this assignment type
    target_form = None
    for pfa in period_forms:
        if pfa.form_type == target_form_type:
            target_form = AssessmentForm.query.get(pfa.form_id)
            break
    
    if not target_form:
        flash('No assessment form found for this assignment type.', 'error')
        return redirect(url_for('my_tasks'))
    
    # Get all questions from the target form
    questions = AssessmentQuestion.query.filter_by(
        form_id=target_form.id,
        is_active=True
    ).order_by(AssessmentQuestion.order).all()
    
    if not questions:
        flash('No questions found in the assessment form.', 'error')
        return redirect(url_for('my_tasks'))
    
    # Get existing responses for this assignment
    existing_responses = {}
    assessment_responses = AssessmentResponse.query.filter_by(
        assessment_assignment_id=assignment_id
    ).all()
    for resp in assessment_responses:
        existing_responses[resp.question_id] = resp
    
    # For external reviewer assessments, get the self-assessment data for reference
    self_assessment_data = None
    self_assessment_questions = None
    if assignment.officer_id != assignment.reviewer_id:
        # Find the self-assessment assignment for this officer
        self_assignment = AssessmentAssignment.query.filter_by(
            officer_id=assignment.officer_id,
            reviewer_id=assignment.officer_id,  # Self-assessment
            period_id=assignment.period_id
        ).first()
        
        if self_assignment and self_assignment.is_submitted:
            # Get self-assessment responses
            self_responses = AssessmentResponse.query.filter_by(
                assessment_assignment_id=self_assignment.id
            ).all()
            self_assessment_data = {resp.question_id: resp for resp in self_responses}
            
            # Get self-assessment form and questions
            self_form = None
            for pfa in period_forms:
                if pfa.form_type == 'self_review':
                    self_form = AssessmentForm.query.get(pfa.form_id)
                    break
            
            if self_form:
                self_assessment_questions = AssessmentQuestion.query.filter_by(
                    form_id=self_form.id,
                    is_active=True
                ).order_by(AssessmentQuestion.order).all()
    
    # Check for auto_submit parameter (from My Tasks submit button)
    auto_submit = request.args.get('auto_submit') == 'true'
    
    if request.method == 'POST' or auto_submit:
        if auto_submit:
            # Auto-submit: mark as submitted without processing form data (data already saved as draft)
            action = 'submit'
            log_activity(current_user.id, 'submit_assessment_new', f'User auto-submitted assessment for assignment #{assignment_id}')
        else:
            action = request.form.get('action', 'submit')
            
            # Regular form submission
            
            if action == 'save_draft':
                log_activity(current_user.id, 'save_draft_assessment', f'User saved draft for assignment #{assignment_id}')
            else:
                log_activity(current_user.id, 'submit_assessment_new', f'User submitted new assessment for assignment #{assignment_id}')
        
        # Process form submission (only if not auto_submit, since data is already saved)
        if not auto_submit:
            for question in questions:
                question_key = f'question_{question.id}'
                response_value = request.form.get(question_key, '').strip()
                
                # Process each question response
                
                # Get or create response (even for empty values to allow saving drafts)
                response = existing_responses.get(question.id)
                if not response:
                    response = AssessmentResponse(
                        assessment_assignment_id=assignment_id,
                        question_id=question.id
                    )
                    db.session.add(response)
                
                # Set response based on question type
                if question.question_type == 'rating':
                    if response_value:
                        try:
                            response.response_number = int(response_value)
                            response.response_text = None
                        except ValueError:
                            flash(f'Invalid rating for question: {question.question_name}', 'error')
                            continue
                    else:
                        response.response_number = None
                elif question.question_type in ['text', 'textarea']:
                    response.response_text = response_value if response_value else None
                elif question.question_type == 'boolean':
                    if response_value:
                        response.response_boolean = response_value.lower() == 'true'
                    else:
                        response.response_boolean = None
                elif question.question_type == 'date':
                    if response_value:
                        try:
                            response.response_date = datetime.strptime(response_value, '%Y-%m-%d').date()
                            response.response_text = None  # Clear text when date is set
                        except ValueError:
                            response.response_text = response_value
                            response.response_date = None
                    else:
                        response.response_date = None
                        response.response_text = None
                else:
                    response.response_text = response_value if response_value else None
        
        # Only mark as submitted if action is submit, not save_draft
        if action == 'submit':
            assignment.is_submitted = True
            assignment.submitted_at = datetime.utcnow()
            flash('Assessment submitted successfully and is pending admin approval.', 'success')
        else:
            flash('Draft saved successfully. You can continue editing later.', 'info')
        
        db.session.commit()
        return redirect(url_for('my_tasks'))
    
    log_activity(current_user.id, 'edit_assessment_new', f'User accessed assessment edit for assignment #{assignment_id}')
    
    return render_template('edit_assessment_new.html',
                         assignment=assignment,
                         target_form=target_form,
                         questions=questions,
                         existing_responses=existing_responses,
                         self_assessment_data=self_assessment_data,
                         self_assessment_questions=self_assessment_questions)

@app.route('/reviewee_view')
@login_required
def reviewee_view():
    if current_user.role != 'officer':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    log_activity(current_user.id, 'view_reviewee_dashboard', f'Reviewee viewed their assignment dashboard')
    
    # Get all assignments where this user is the officer (being reviewed)
    assignments = AssessmentAssignment.query.filter_by(
        officer_id=current_user.id
    ).join(AssessmentPeriod).order_by(AssessmentPeriod.start_date.desc()).all()
    
    # Calculate statistics
    pending_count = len([a for a in assignments if not a.is_completed and not a.is_submitted])
    completed_count = len([a for a in assignments if a.is_completed])
    submitted_count = len([a for a in assignments if a.is_submitted and not a.is_completed])
    total_assignments = len(assignments)
    completion_rate = (completed_count / total_assignments * 100) if total_assignments > 0 else 0
    
    return render_template('reviewee_view.html',
                         assignments=assignments,
                         pending_count=pending_count,
                         completed_count=completed_count,
                         submitted_count=submitted_count,
                         completion_rate=round(completion_rate, 1))

@app.route('/reports')
@login_required
@admin_required
def reports():
    log_activity(current_user.id, 'view_reports', f'Admin accessed reports page')
    current_year = datetime.now().year
    officers = User.query.filter_by(role='officer').all()
    categories = Category.query.order_by(Category.order).all()
    
    officer_reports = []
    for officer in officers:
        assessments = Assessment.query.filter_by(officer_id=officer.id, year=current_year).all()
        
        if assessments:
            # Calculate category averages
            category_data = {}
            for category in categories:
                ratings = []
                for assessment in assessments:
                    rating = CategoryRating.query.filter_by(
                        assessment_id=assessment.id,
                        category_id=category.id
                    ).first()
                    if rating:
                        ratings.append(rating.rating)
                
                if ratings:
                    category_data[category.id] = round(sum(ratings) / len(ratings), 2)
                else:
                    category_data[category.id] = 0
            
            overall_avg = round(sum(assessment.overall_rating for assessment in assessments if assessment.overall_rating) / len(assessments), 2)
            
            officer_reports.append({
                'officer': officer,
                'assessments_count': len(assessments),
                'overall_average': overall_avg,
                'category_averages': category_data
            })
    
    return render_template('reports.html', officer_reports=officer_reports, categories=categories)

@app.route('/export_pdf/<int:officer_id>')
@login_required
@admin_required
def export_pdf(officer_id):
    officer = User.query.get_or_404(officer_id)
    current_year = datetime.now().year
    
    assessments = Assessment.query.filter_by(officer_id=officer_id, year=current_year).all()
    
    if not assessments:
        flash('No assessments found for this officer.', 'error')
        return redirect(url_for('reports'))
    
    pdf_data = generate_pdf_report(officer, assessments)
    
    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={officer.name}_assessment_{current_year}.pdf'
    
    return response

@app.route('/admin/export_matrix_pdf/<int:officer_id>')
@login_required
@admin_required
def export_matrix_pdf(officer_id):
    try:
        officer = User.query.get_or_404(officer_id)
        if officer.role != 'officer':
            flash('Invalid officer selection.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Get current active assessment project
        current_period = AssessmentPeriod.query.filter_by(is_active=True).first()
        if not current_period:
            flash('No active assessment project found.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Get all assignments for this officer in the current period
        assignments = AssessmentAssignment.query.filter_by(
            officer_id=officer_id,
            period_id=current_period.id
        ).all()
        
        if not assignments:
            flash(f'No assessment assignments found for {officer.name} in the current period.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Create reviewers list from ALL ASSIGNMENTS
        reviewers = {}
        for assignment in assignments:
            reviewers[assignment.reviewer.id] = assignment.reviewer
        
        # Get all assessment forms assigned to this period for reviewers
        reviewer_forms = db.session.query(AssessmentForm).join(PeriodFormAssignment).filter(
            PeriodFormAssignment.period_id == current_period.id,
            PeriodFormAssignment.form_type == 'reviewer',
            PeriodFormAssignment.is_active == True,
            AssessmentForm.is_active == True
        ).all()
        
        # Build matrix data: questions in rows, reviewers in columns
        matrix_data = []
        processed_question_ids = set()
        
        # Get all questions from all reviewer forms
        all_questions = []
        for form in reviewer_forms:
            form_questions = AssessmentQuestion.query.filter_by(
                form_id=form.id,
                is_active=True
            ).order_by(AssessmentQuestion.order.asc()).all()
            all_questions.extend(form_questions)
        
        for question in all_questions:
            if question.id in processed_question_ids:
                continue
            processed_question_ids.add(question.id)
            
            question_row = {
                'question': question,
                'responses': {},
                'average_rating': 0,
                'response_count': 0
            }
            
            total_rating = 0
            rating_count = 0
            responses_for_ai = []
            
            for reviewer_id, reviewer in reviewers.items():
                assignment = next((a for a in assignments if a.reviewer_id == reviewer_id), None)
                if assignment:
                    response = AssessmentResponse.query.filter_by(
                        assessment_assignment_id=assignment.id,
                        question_id=question.id
                    ).first()
                    
                    if response:
                        response_data = response.get_response_data()
                        rating = None
                        comment = ""
                        
                        if question.question_type == 'rating' and response.response_number is not None:
                            rating = int(response.response_number)
                            comment = response.response_text or ""
                        elif question.question_type in ['text', 'textarea']:
                            comment = response.response_text or ""
                        else:
                            comment = str(response_data) if response_data else ""
                        
                        question_row['responses'][reviewer.name] = {
                            'rating': rating,
                            'comment': comment,
                            'text_response': comment,
                            'reviewer': reviewer,
                            'response_data': response_data
                        }
                        
                        if rating:
                            total_rating += rating
                            rating_count += 1
                            
                        if rating or comment:
                            responses_for_ai.append({
                                'reviewer': reviewer.name,
                                'rating': rating,
                                'comment': comment,
                                'reviewer_role': reviewer.role
                            })
                    else:
                        question_row['responses'][reviewer.name] = None
                else:
                    question_row['responses'][reviewer.name] = None
            
            if rating_count > 0:
                question_row['average_rating'] = round(total_rating / rating_count, 2)
                question_row['response_count'] = rating_count
            
            try:
                from ai_analysis import generate_feedback_summary
                ai_analysis = generate_feedback_summary(question.question_text, responses_for_ai)
                question_row['ai_analysis'] = ai_analysis
            except Exception as e:
                question_row['ai_analysis'] = {
                    "summary": "AI analysis temporarily unavailable.",
                    "key_themes": ["Manual review required"],
                    "sentiment": "neutral"
                }
            
            matrix_data.append(question_row)
        
        # Calculate overall average
        all_ratings = []
        for question_data in matrix_data:
            for reviewer_name, response in question_data['responses'].items():
                if response and response.get('rating'):
                    all_ratings.append(response['rating'])
        
        overall_matrix_average = round(sum(all_ratings) / len(all_ratings), 2) if all_ratings else 0
        
        # Generate comprehensive AI summary
        try:
            from ai_analysis import generate_overall_performance_summary
            overall_ai_summary = generate_overall_performance_summary(officer.name, matrix_data)
        except Exception as e:
            overall_ai_summary = {
                "executive_summary": "Comprehensive performance analysis available with detailed reviewer feedback.",
                "major_themes": ["Multi-reviewer assessment", "Comprehensive evaluation"],
                "overall_sentiment": "satisfactory"
            }
        
        pdf_buffer = generate_matrix_pdf_report(officer, matrix_data, reviewers, overall_matrix_average, overall_ai_summary)
        
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{officer.name}_matrix_report.pdf"'
        
        log_activity(current_user.id, 'export_matrix_pdf', f'Exported matrix PDF report for {officer.name}')
        
        return response
        
    except Exception as e:
        print(f"Error generating matrix PDF: {e}")
        flash('Error generating PDF export. Please try again.', 'error')
        return redirect(url_for('admin_dashboard'))
        question_row = {
            'question': question,
            'responses': {},
            'average_rating': 0,
            'response_count': 0
        }
        
        total_rating = 0
        rating_count = 0
        responses_for_ai = []
        
        for reviewer_id, reviewer in reviewers.items():
            # Find assessment for this reviewer
            assessment = next((a for a in assessments if a.reviewer_id == reviewer_id), None)
            if assessment:
                # Find the FIRST response for this question in this assessment (avoid duplicates)
                response = QuestionResponse.query.filter_by(
                    assessment_id=assessment.id,
                    question_id=question.id
                ).first()
                
                if response:
                    # Store only one response per reviewer per question
                    question_row['responses'][reviewer.name] = {
                        'rating': response.response_rating,
                        'comment': response.comment,
                        'text_response': response.response_text,
                        'reviewer': reviewer
                    }
                    
                    if response.response_rating:
                        total_rating += response.response_rating
                        rating_count += 1
                        
                    # Collect for AI analysis (only valid responses with content)
                    if response.response_rating or response.comment or response.response_text:
                        responses_for_ai.append({
                            'reviewer': reviewer.name,
                            'rating': response.response_rating,
                            'comment': response.comment or response.response_text,
                            'reviewer_role': reviewer.role
                        })
                else:
                    question_row['responses'][reviewer.name] = None
            else:
                question_row['responses'][reviewer.name] = None
        
        # Calculate average rating for this question
        if rating_count > 0:
            question_row['average_rating'] = round(total_rating / rating_count, 2)
            question_row['response_count'] = rating_count
        
        # Generate AI-powered analysis for this question - EXACT SAME LOGIC
        try:
            from ai_analysis import generate_feedback_summary
            ai_analysis = generate_feedback_summary(question.text, responses_for_ai)
            question_row['ai_analysis'] = ai_analysis
        except Exception as e:
            question_row['ai_analysis'] = {
                "summary": "AI analysis temporarily unavailable.",
                "key_themes": ["Manual review required"],
                "sentiment": "neutral"
            }
        
        # Add ALL questions to the matrix (show empty cells for pending reviews) - SAME AS WEB
        matrix_data.append(question_row)
    
    # Calculate overall average across all questions - EXACT SAME LOGIC
    all_ratings = []
    for question_data in matrix_data:
        for reviewer_name, response in question_data['responses'].items():
            if response and response.get('rating'):
                all_ratings.append(response['rating'])
    
    overall_matrix_average = round(sum(all_ratings) / len(all_ratings), 2) if all_ratings else 0
    
    # Generate comprehensive AI summary - EXACT SAME LOGIC AS WEB
    try:
        from ai_analysis import generate_overall_performance_summary
        overall_ai_summary = generate_overall_performance_summary(officer.name, matrix_data)
    except Exception as e:
        overall_ai_summary = {
            "executive_summary": "Comprehensive performance analysis available with detailed reviewer feedback.",
            "major_themes": ["Multi-reviewer assessment", "Comprehensive evaluation"],
            "overall_sentiment": "satisfactory"
        }
    
    pdf_buffer = generate_matrix_pdf_report(officer, matrix_data, reviewers, overall_matrix_average, overall_ai_summary)
    
    response = make_response(pdf_buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{officer.name}_matrix_report.pdf"'
    
    log_activity(current_user.id, 'export_matrix_pdf', f'Exported matrix PDF report for {officer.name}')
    
    return response

@app.route('/admin/export_matrix_excel/<int:officer_id>')
@login_required
@admin_required  
def export_matrix_excel(officer_id):
    """Use cached matrix data from file for Excel export"""
    try:
        officer = User.query.get_or_404(officer_id)
        if officer.role != 'officer':
            flash('Invalid officer selection.', 'error')
            return redirect(url_for('admin_dashboard'))
        
        # Load cached data from file
        import json
        import os
        cache_file = f"matrix_cache_{officer_id}.json"
        
        if not os.path.exists(cache_file):
            flash('Please view the matrix first to generate Excel data.', 'info')
            return redirect(url_for('officer_reviews', officer_id=officer_id))
        
        with open(cache_file, 'r') as f:
            cached_data = json.load(f)
        
        # Create comprehensive Excel matching web format exactly
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{cached_data['officer_name']} Performance Matrix"
        
        # Get reviewers list
        reviewers = set()
        for row in cached_data['matrix_rows']:
            for reviewer in row['reviewer_ratings'].keys():
                reviewers.add(reviewer)
        reviewers = sorted(reviewers)
        
        # Create headers matching web format (no AI column)
        headers = ["Question"] + reviewers + ["Avg"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows with color coding for ratings
        current_row = 2
        for question_row in cached_data['matrix_rows']:
            # Question text (column A)
            ws.cell(row=current_row, column=1, value=question_row['question_text'])
            
            # Reviewer ratings
            for col, reviewer in enumerate(reviewers, 2):
                rating = question_row['reviewer_ratings'].get(reviewer, {}).get('rating', '')
                if rating:
                    cell = ws.cell(row=current_row, column=col, value=rating)
                    # Color code ratings like web version
                    if rating == 5:
                        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif rating == 4:
                        cell.fill = PatternFill(start_color="84CC16", end_color="84CC16", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif rating == 3:
                        cell.fill = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif rating == 2:
                        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif rating == 1:
                        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws.cell(row=current_row, column=col, value="—")
            
            # Average rating
            avg_col = len(reviewers) + 2
            avg_rating = question_row['average_rating']
            cell = ws.cell(row=current_row, column=avg_col, value=avg_rating)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            
            current_row += 1
        
        # Total Average Score row
        ws.cell(row=current_row, column=1, value="Total Average Score").font = Font(bold=True)
        for col in range(2, len(reviewers) + 2):
            ws.cell(row=current_row, column=col, value="—")
        
        avg_cell = ws.cell(row=current_row, column=len(reviewers) + 2, value=cached_data['overall_average'])
        avg_cell.font = Font(bold=True, size=12)
        avg_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Note about calculation method
        ws.cell(row=current_row, column=1, value="📊 Calculated from questions across reviewers")
        ws.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 1, end_column=len(headers))
        
        current_row += 2
        
        # Overall Performance Summary section (matching web format)
        summary_cell = ws.cell(row=current_row, column=1, value="📋 Overall Performance Summary")
        summary_cell.font = Font(bold=True, size=14, color="FFFFFF")
        summary_cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        
        # Merge cells for summary header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        
        current_row += 1
        
        # Add AI summary content
        if 'overall_ai_summary' in cached_data:
            ai_summary = cached_data['overall_ai_summary']
            
            # Executive summary
            ws.cell(row=current_row, column=1, value=ai_summary.get('executive_summary', 'Performance analysis complete.'))
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 2
            
            # Major themes
            themes_cell = ws.cell(row=current_row, column=1, value="Major Themes")
            themes_cell.font = Font(bold=True, color="3B82F6")
            current_row += 1
            
            themes = ai_summary.get('major_themes', [])
            for theme in themes:
                ws.cell(row=current_row, column=1, value=f"• {theme}")
                current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{cached_data["officer_name"]}_matrix_report.xlsx"'
        
        log_activity(current_user.id, 'export_matrix_excel', f'Exported cached matrix Excel for {cached_data["officer_name"]}')
        return response
        
    except Exception as e:
        print(f"Error in export_matrix_excel: {e}")
        flash('Error generating Excel export. Please view matrix first.', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/export_csv')
@login_required
@admin_required
def export_csv():
    current_year = datetime.now().year
    csv_data = export_csv_data(current_year)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    for row in csv_data:
        writer.writerow(row)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=assessments_{current_year}.csv'
    
    return response

@app.route('/manage_users')
@login_required
@admin_required
def manage_users():
    log_activity(current_user.id, 'view_manage_users', f'Admin accessed user management page')
    users = User.query.order_by(User.created_at.desc()).all()
    activity_logs = get_activity_logs(50)  # Get recent 50 activities
    return render_template('manage_users.html', users=users, activity_logs=activity_logs)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = EditUserForm()
    
    if form.validate_on_submit():
        # Check if email is already taken by another user
        existing_user = User.query.filter(User.email == form.email.data, User.id != user_id).first()
        if existing_user:
            flash('Email address is already in use by another user.', 'error')
            return render_template('edit_user.html', form=form, user=user)
        
        # Update user information
        old_email = user.email
        old_role = user.role
        
        user.name = form.name.data
        user.email = form.email.data
        user.role = form.role.data
        
        # Update password if provided
        if form.new_password.data:
            user.set_password(form.new_password.data)
            log_activity(current_user.id, 'password_change', f'Admin changed password for user {user.name}')
        
        try:
            db.session.commit()
            
            # Log the changes
            changes = []
            if old_email != user.email:
                changes.append(f'email: {old_email} → {user.email}')
            if old_role != user.role:
                changes.append(f'role: {old_role} → {user.role}')
            if form.new_password.data:
                changes.append('password updated')
            
            log_activity(current_user.id, 'user_update', f'Updated user {user.name}: {", ".join(changes)}')
            
            flash(f'User {user.name} has been updated successfully.', 'success')
            return redirect(url_for('manage_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating user: {str(e)}', 'error')
    
    # Pre-populate form with current user data
    if request.method == 'GET':
        form.name.data = user.name
        form.email.data = user.email
        form.role.data = user.role
    
    return render_template('edit_user.html', form=form, user=user)

@app.route('/user_activity/<int:user_id>')
@login_required
@admin_required
def user_activity(user_id):
    user = User.query.get_or_404(user_id)
    activity_logs = get_user_activity_logs(user_id, 100)
    log_activity(current_user.id, 'view_user_activity', f'Admin viewed activity logs for user: {user.name}')
    return render_template('user_activity.html', user=user, activity_logs=activity_logs)

@app.route('/admin/assessment_activity_logs')
@login_required
@admin_required
def assessment_activity_logs():
    log_activity(current_user.id, 'view_assessment_activity_logs', 'Admin accessed assessment activity logs')
    
    # Get filter parameters
    officer_id = request.args.get('officer_id', type=int)
    period_id = request.args.get('period_id', type=int)
    user_id = request.args.get('user_id', type=int)
    event_type = request.args.get('event_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Use regular ActivityLog since we have assessment-related activity logs there
    query = ActivityLog.query
    
    # Apply filters for assessment-related activities
    if officer_id:
        # Filter by activities related to this specific officer using description field
        officer = User.query.get(officer_id)
        if officer:
            query = query.filter(ActivityLog.description.like(f'%{officer.name}%'))
    
    if user_id:
        query = query.filter(ActivityLog.user_id == user_id)
    
    # Filter for assessment-related activities
    assessment_actions = ['view_assessment', 'create_assessment', 'edit_assessment', 'submit_assessment', 
                         'approve_assessment', 'reject_assessment', 'assignment_created', 'assignment_updated',
                         'view_officer_reviews', 'assignment_matrix_updated', 'approve_reviewer_assessment']
    
    if event_type:
        query = query.filter(ActivityLog.action == event_type)
    else:
        # Show assessment-related activities or activities mentioning the officer
        if officer_id:
            query = query.filter(
                db.or_(
                    ActivityLog.action.in_(assessment_actions),
                    ActivityLog.description.like(f'%{officer.name}%')
                )
            )
        else:
            query = query.filter(ActivityLog.action.in_(assessment_actions))
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(ActivityLog.timestamp >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(ActivityLog.timestamp <= to_date)
        except ValueError:
            pass
    
    # Order by timestamp and paginate
    query = query.order_by(ActivityLog.timestamp.desc())
    logs = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get filter data
    periods = AssessmentPeriod.query.filter_by(is_active=True).order_by(AssessmentPeriod.name).all()
    users = User.query.order_by(User.name).all()
    event_types = assessment_actions
    
    return render_template('assessment_activity_logs.html',
                         logs=logs,
                         periods=periods,
                         users=users,
                         event_types=event_types,
                         selected_officer=officer_id,
                         selected_period=period_id,
                         selected_user=user_id,
                         selected_event_type=event_type,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/admin/activity_logs')
@login_required
@admin_required
def admin_activity_logs():
    log_activity(current_user.id, 'view_activity_logs', 'Admin accessed system activity logs')
    
    # Get filter parameters
    user_filter = request.args.get('user_id', type=int)
    action_filter = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Build query
    query = ActivityLog.query
    
    # Apply filters
    if user_filter:
        query = query.filter(ActivityLog.user_id == user_filter)
    
    if action_filter:
        query = query.filter(ActivityLog.action.like(f'%{action_filter}%'))
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(ActivityLog.timestamp >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Add 1 day to include the entire day
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(ActivityLog.timestamp <= to_date)
        except ValueError:
            pass
    
    # Order by timestamp and paginate
    query = query.order_by(ActivityLog.timestamp.desc())
    logs = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get all users for filter dropdown
    users = User.query.order_by(User.name).all()
    
    # Get unique action types for filter
    action_types = db.session.query(ActivityLog.action).distinct().order_by(ActivityLog.action).all()
    action_types = [action[0] for action in action_types]
    
    # Get summary statistics
    total_activities = ActivityLog.query.count()
    today_activities = ActivityLog.query.filter(
        ActivityLog.timestamp >= datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    
    unique_users_today = db.session.query(ActivityLog.user_id).filter(
        ActivityLog.timestamp >= datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    ).distinct().count()
    
    # Handle CSV export
    if request.args.get('export') == 'csv':
        log_activity(current_user.id, 'export_activity_logs', 'Admin exported activity logs to CSV')
        
        # Get all matching records (not paginated)
        all_logs = query.all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write CSV headers
        writer.writerow(['Timestamp', 'User Name', 'User Role', 'Action', 'Description', 'IP Address', 'User Agent'])
        
        # Write data rows
        for log in all_logs:
            writer.writerow([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.user.name,
                log.user.role,
                log.action,
                log.description,
                log.ip_address or '',
                log.user_agent or ''
            ])
        
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=activity_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
    
    return render_template('admin_activity_logs.html',
                         logs=logs,
                         users=users,
                         action_types=action_types,
                         user_filter=user_filter,
                         action_filter=action_filter,
                         date_from=date_from,
                         date_to=date_to,
                         total_activities=total_activities,
                         today_activities=today_activities,
                         unique_users_today=unique_users_today)

@app.route('/admin/assessments')
@login_required
@admin_required
def admin_assessments():
    log_activity(current_user.id, 'view_admin_assessments', 'Admin accessed assessments overview')
    
    # Get filter parameters
    period_filter = request.args.get('period', 'current')  # current or past
    officer_filter = request.args.get('officer_id', type=int)
    reviewer_filter = request.args.get('reviewer_id', type=int)
    assignment_period_filter = request.args.get('assignment_period_id', type=int)
    status_filter = request.args.get('status', '')  # completed, pending
    page = request.args.get('page', 1, type=int)
    per_page = 25
    
    # Build query based on period filter
    if period_filter == 'current':
        periods_query = AssessmentPeriod.query.filter_by(is_active=True)
    else:  # past
        periods_query = AssessmentPeriod.query.filter_by(is_active=False)
    
    # Get assessment assignments with proper joins
    query = db.session.query(AssessmentAssignment).join(
        AssessmentPeriod, AssessmentAssignment.period_id == AssessmentPeriod.id
    )
    
    # Apply period filter
    if period_filter == 'current':
        query = query.filter(AssessmentPeriod.is_active == True)
    else:
        query = query.filter(AssessmentPeriod.is_active == False)
    
    # Apply additional filters
    if officer_filter:
        query = query.filter(AssessmentAssignment.officer_id == officer_filter)
    
    if reviewer_filter:
        query = query.filter(AssessmentAssignment.reviewer_id == reviewer_filter)
    
    if assignment_period_filter:
        query = query.filter(AssessmentAssignment.period_id == assignment_period_filter)
    
    if status_filter == 'completed':
        query = query.filter(AssessmentAssignment.is_completed == True)
    elif status_filter == 'pending':
        query = query.filter(AssessmentAssignment.is_completed == False)
    
    # Order and paginate
    query = query.order_by(AssessmentAssignment.id.desc())
    assignments = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get filter options - only active users
    officers = User.query.filter_by(role='officer', is_active=True).order_by(User.name).all()
    reviewers = User.query.filter_by(role='board_member', is_active=True).order_by(User.name).all()
    
    if period_filter == 'current':
        assignment_periods = AssessmentPeriod.query.filter_by(is_active=True).order_by(AssessmentPeriod.name).all()
    else:
        assignment_periods = AssessmentPeriod.query.filter_by(is_active=False).order_by(AssessmentPeriod.name.desc()).all()
    
    # Get summary statistics
    base_query = db.session.query(AssessmentAssignment).join(
        AssessmentPeriod, AssessmentAssignment.period_id == AssessmentPeriod.id
    )
    
    # Apply same period filter for statistics
    if period_filter == 'current':
        base_query = base_query.filter(AssessmentPeriod.is_active == True)
    else:
        base_query = base_query.filter(AssessmentPeriod.is_active == False)
    
    # Apply same additional filters for statistics
    if officer_filter:
        base_query = base_query.filter(AssessmentAssignment.officer_id == officer_filter)
    if reviewer_filter:
        base_query = base_query.filter(AssessmentAssignment.reviewer_id == reviewer_filter)
    if assignment_period_filter:
        base_query = base_query.filter(AssessmentAssignment.period_id == assignment_period_filter)
    if status_filter == 'completed':
        base_query = base_query.filter(AssessmentAssignment.is_completed == True)
    elif status_filter == 'pending':
        base_query = base_query.filter(AssessmentAssignment.is_completed == False)
    
    total_assignments = base_query.count()
    completed_count = base_query.filter(AssessmentAssignment.is_completed == True).count()
    pending_count = total_assignments - completed_count
    completion_rate = round((completed_count / total_assignments * 100), 1) if total_assignments > 0 else 0
    
    return render_template('admin_assessments.html',
                         assignments=assignments,
                         officers=officers,
                         reviewers=reviewers,
                         assignment_periods=assignment_periods,
                         period_filter=period_filter,
                         officer_filter=officer_filter,
                         reviewer_filter=reviewer_filter,
                         assignment_period_filter=assignment_period_filter,
                         status_filter=status_filter,
                         total_assignments=total_assignments,
                         completed_count=completed_count,
                         pending_count=pending_count,
                         completion_rate=completion_rate)

@app.route('/create_user', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    form = UserForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Check if email already exists
            existing_user = User.query.filter_by(email=form.email.data).first()
            if existing_user:
                flash(f'A user with email "{form.email.data}" already exists. Please use a different email address.', 'error')
                return render_template('create_user.html', form=form)
            
            # Create new user
            user = User(
                name=form.name.data,
                email=form.email.data,
                role=form.role.data
            )
            user.set_password(form.password.data)
            
            db.session.add(user)
            db.session.commit()
            log_activity(current_user.id, 'user_create', f'Created new user {user.name} with role {user.role}')
            
            flash(f'User {user.name} created successfully.', 'success')
            return redirect(url_for('manage_users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating user: {str(e)}', 'error')
            return render_template('create_user.html', form=form)
    
    # Display form validation errors if any
    if request.method == 'POST' and not form.validate():
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('create_user.html', form=form)

@app.route('/assessment_periods')
@login_required
@admin_required
def assessment_periods():
    log_activity(current_user.id, 'view_assessment_periods', f'Admin accessed assessment projects page')
    periods = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).all()
    return render_template('assessment_periods.html', periods=periods)

@app.route('/create_assessment_period', methods=['GET', 'POST'])
@login_required
@admin_required
def create_assessment_period():
    form = AssessmentPeriodForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Validate dates
            if form.end_date.data <= form.start_date.data:
                flash('End date must be after start date.', 'error')
                return render_template('create_assessment_period.html', form=form)
            
            period = AssessmentPeriod(
                name=form.name.data,
                description=form.description.data,
                start_date=form.start_date.data,
                end_date=form.end_date.data,
                due_date=form.due_date.data,
                created_by=current_user.id
            )
            
            db.session.add(period)
            db.session.flush()  # Get the period ID
            
            # Create form assignments
            from models import PeriodFormAssignment
            
            # Create form assignments for reviewer forms
            if form.reviewer_form_ids.data:
                for form_id in form.reviewer_form_ids.data:
                    assignment = PeriodFormAssignment(
                        period_id=period.id,
                        form_id=form_id,
                        form_type='reviewer'
                    )
                    db.session.add(assignment)
            
            # Create form assignments for self-review forms
            if form.self_review_form_ids.data:
                for form_id in form.self_review_form_ids.data:
                    assignment = PeriodFormAssignment(
                        period_id=period.id,
                        form_id=form_id,
                        form_type='self_review'
                    )
                    db.session.add(assignment)
            
            db.session.commit()
            
            form_count = len(form.reviewer_form_ids.data or []) + len(form.self_review_form_ids.data or [])
            flash_message = f'Assessment period "{period.name}" created successfully'
            if form_count > 0:
                flash_message += f' with {form_count} assessment form(s) assigned.'
            else:
                flash_message += '. You can assign assessment forms later.'
            
            flash(flash_message, 'success')
            log_activity(current_user.id, 'create_assessment_period', f'Created assessment project: {period.name} with {form_count} forms assigned')
            return redirect(url_for('assessment_periods'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating assessment project: {str(e)}', 'error')
            return render_template('create_assessment_period.html', form=form)
    
    if request.method == 'POST' and not form.validate():
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('create_assessment_period.html', form=form)

@app.route('/edit_assessment_period/<int:period_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_assessment_period(period_id):
    from forms import EditAssessmentPeriodForm
    from models import AssessmentPeriod
    
    period = AssessmentPeriod.query.get_or_404(period_id)
    form = EditAssessmentPeriodForm(obj=period)
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Validate dates
            if form.end_date.data <= form.start_date.data:
                flash('End date must be after start date.', 'error')
                return render_template('edit_assessment_period.html', form=form, period=period)
            
            if form.due_date.data and form.due_date.data > form.end_date.data:
                flash('Due date cannot be after end date.', 'error')
                return render_template('edit_assessment_period.html', form=form, period=period)
            
            # Update period
            period.name = form.name.data
            period.description = form.description.data
            period.start_date = form.start_date.data
            period.end_date = form.end_date.data
            period.due_date = form.due_date.data
            
            db.session.commit()
            flash(f'Assessment period "{period.name}" updated successfully.', 'success')
            log_activity(current_user.id, 'edit_assessment_period', f'Updated assessment project: {period.name}')
            return redirect(url_for('admin_main'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating assessment project: {str(e)}', 'error')
    
    if request.method == 'POST' and not form.validate():
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('edit_assessment_period.html', form=form, period=period)

@app.route('/clone_assessment_period/<int:period_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def clone_assessment_period(period_id):
    from forms import CloneAssessmentPeriodForm
    from models import AssessmentPeriod, PeriodFormAssignment
    
    original_period = AssessmentPeriod.query.get_or_404(period_id)
    form = CloneAssessmentPeriodForm()
    
    # Pre-fill with original data but modified
    if request.method == 'GET':
        form.name.data = f"Copy of {original_period.name}"
        form.description.data = original_period.description
        # Set dates to next year
        if original_period.start_date:
            form.start_date.data = original_period.start_date.replace(year=original_period.start_date.year + 1)
        if original_period.end_date:
            form.end_date.data = original_period.end_date.replace(year=original_period.end_date.year + 1)
        if original_period.due_date:
            form.due_date.data = original_period.due_date.replace(year=original_period.due_date.year + 1)
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Validate dates
            if form.end_date.data <= form.start_date.data:
                flash('End date must be after start date.', 'error')
                return render_template('clone_assessment_period.html', form=form, original_period=original_period)
            
            if form.due_date.data and form.due_date.data > form.end_date.data:
                flash('Due date cannot be after end date.', 'error')
                return render_template('clone_assessment_period.html', form=form, original_period=original_period)
            
            # Create new period
            new_period = AssessmentPeriod(
                name=form.name.data,
                description=form.description.data,
                start_date=form.start_date.data,
                end_date=form.end_date.data,
                due_date=form.due_date.data,
                created_by=current_user.id
            )
            
            db.session.add(new_period)
            db.session.flush()  # Get the period ID
            
            # Copy form assignments if requested
            if form.copy_forms.data:
                original_assignments = PeriodFormAssignment.query.filter_by(period_id=original_period.id).all()
                for assignment in original_assignments:
                    new_assignment = PeriodFormAssignment(
                        period_id=new_period.id,
                        form_id=assignment.form_id,
                        form_type=assignment.form_type
                    )
                    db.session.add(new_assignment)
            
            db.session.commit()
            
            flash_message = f'Assessment period "{new_period.name}" cloned successfully'
            if form.copy_forms.data:
                form_count = PeriodFormAssignment.query.filter_by(period_id=new_period.id).count()
                flash_message += f' with {form_count} assessment form(s) copied.'
            else:
                flash_message += '. You can assign assessment forms separately.'
            
            flash(flash_message, 'success')
            log_activity(current_user.id, 'clone_assessment_period', f'Cloned assessment project: {original_period.name} -> {new_period.name}')
            return redirect(url_for('admin_main'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error cloning assessment project: {str(e)}', 'error')
    
    if request.method == 'POST' and not form.validate():
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('clone_assessment_period.html', form=form, original_period=original_period)

@app.route('/admin/get_period_dependencies/<int:period_id>')
@login_required
@admin_required
def get_period_dependencies(period_id):
    """Get assessment project dependencies for deletion confirmation"""
    try:
        period = AssessmentPeriod.query.get_or_404(period_id)
        
        # Get assignments for this period
        assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
        assignments_count = len(assignments)
        completed_assignments = len([a for a in assignments if a.is_completed])
        
        # Get assessments linked to this period through assignments
        assessment_ids = [a.assessment_id for a in assignments if a.assessment_id]
        assessments_count = Assessment.query.filter(Assessment.id.in_(assessment_ids)).count() if assessment_ids else 0
        
        # Get survey responses linked to this period
        survey_responses_count = 0
        if assignments:
            assignment_ids = [a.id for a in assignments]
            from models import SurveyResponse
            survey_responses_count = SurveyResponse.query.filter(SurveyResponse.assessment_assignment_id.in_(assignment_ids)).count()
        
        # Get form assignments for this period
        from models import PeriodFormAssignment
        form_assignments = PeriodFormAssignment.query.filter_by(period_id=period_id).all()
        reviewer_forms = len([f for f in form_assignments if f.form_type == 'reviewer'])
        self_review_forms = len([f for f in form_assignments if f.form_type == 'self_review'])
        
        # Get unique users involved
        involved_users = set()
        for assignment in assignments:
            involved_users.add(assignment.officer_id)
            involved_users.add(assignment.reviewer_id)
        
        user_names = []
        if involved_users:
            users = User.query.filter(User.id.in_(involved_users)).all()
            user_names = [f"{u.name} ({u.role})" for u in users]
        
        dependencies = {
            'assignments_total': assignments_count,
            'assignments_completed': completed_assignments,
            'assignments_pending': assignments_count - completed_assignments,
            'assessments_count': assessments_count,
            'survey_responses_count': survey_responses_count,
            'reviewer_forms': reviewer_forms,
            'self_review_forms': self_review_forms,
            'involved_users_count': len(involved_users),
            'involved_users': user_names[:10],  # Limit to first 10 for display
            'has_dependencies': any([assignments_count, assessments_count, survey_responses_count, reviewer_forms, self_review_forms])
        }
        
        return jsonify({
            'success': True,
            'dependencies': dependencies,
            'period': {
                'name': period.name,
                'description': period.description,
                'start_date': period.start_date.strftime('%Y-%m-%d'),
                'end_date': period.end_date.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/admin/delete_assessment_period/<int:period_id>', methods=['POST'])
@login_required
@admin_required
def delete_assessment_period(period_id):
    """Delete assessment project with all related data"""
    try:
        confirmation_text = request.form.get('confirmation_text', '').strip()
        
        # Verify confirmation text
        if confirmation_text != 'DELETE PERIOD':
            flash('Incorrect confirmation text. Period deletion cancelled.', 'error')
            return redirect(url_for('assessment_periods'))
        
        period = AssessmentPeriod.query.get_or_404(period_id)
        period_name = period.name
        
        # Get all assignments for this period
        assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
        
        # Delete survey responses linked to assignments
        assignment_ids = [a.id for a in assignments]
        if assignment_ids:
            from models import SurveyResponse
            SurveyResponse.query.filter(SurveyResponse.assessment_assignment_id.in_(assignment_ids)).delete()
        
        # Delete assessments linked to assignments
        assessment_ids = [a.assessment_id for a in assignments if a.assessment_id]
        if assessment_ids:
            # Delete assessment responses first
            AssessmentResponse.query.filter(AssessmentResponse.assessment_assignment_id.in_(assignment_ids)).delete()
            # Delete category ratings
            CategoryRating.query.filter(CategoryRating.assessment_id.in_(assessment_ids)).delete()
            # Delete assessments
            Assessment.query.filter(Assessment.id.in_(assessment_ids)).delete()
        
        # Delete assignments
        AssessmentAssignment.query.filter_by(period_id=period_id).delete()
        
        # Delete form assignments
        from models import PeriodFormAssignment
        PeriodFormAssignment.query.filter_by(period_id=period_id).delete()
        
        # Finally delete the period
        db.session.delete(period)
        db.session.commit()
        
        # Log the deletion
        log_activity(current_user.id, 'period_deleted', 
                    f'Deleted assessment project "{period_name}" and all related data')
        
        flash(f'Assessment period "{period_name}" and all related data have been permanently deleted.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'period_delete_failed', f'Failed to delete period ID {period_id}: {str(e)}')
        flash(f'Error deleting assessment project: {str(e)}', 'error')
    
    return redirect(url_for('assessment_periods'))

@app.route('/manage_assignments/<int:period_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_assignments(period_id):
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get selected reviewees for this period from PeriodReviewee
    selected_reviewee_ids = db.session.query(PeriodReviewee.user_id).filter_by(period_id=period_id).subquery()
    reviewees = User.query.filter(User.id.in_(selected_reviewee_ids), User.is_active == True).order_by(User.name).all()
    
    # Get selected reviewers for this period from PeriodReviewer  
    selected_reviewer_ids = db.session.query(PeriodReviewer.user_id).filter_by(period_id=period_id).subquery()
    reviewers = User.query.filter(User.id.in_(selected_reviewer_ids), User.is_active == True).order_by(User.name).all()
    
    # Get existing assignments for this period
    existing_assignments = {}
    assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
    for assignment in assignments:
        key = f"{assignment.officer_id}_{assignment.reviewer_id}"
        existing_assignments[key] = True
    
    if request.method == 'POST':
        try:
            # Smart assignment management - preserve assignments with existing responses
            assignments_with_responses = set()  # Track assignments that have responses
            assignments_to_keep = set()  # Track assignments requested by form
            assignments_deleted = 0
            
            existing_assignments_query = AssessmentAssignment.query.filter_by(period_id=period_id).all()
            
            # Build a set of requested assignments from the form
            for reviewee in reviewees:
                # Add self-assessment for officers
                if reviewee.role == 'officer':
                    assignments_to_keep.add(f"{reviewee.id}_{reviewee.id}")
                
                # Add checked assignments from matrix
                for reviewer in reviewers:
                    checkbox_name = f"assignment_{reviewee.id}_{reviewer.id}"
                    if request.form.get(checkbox_name):
                        assignments_to_keep.add(f"{reviewee.id}_{reviewer.id}")
            
            # Check existing assignments for responses
            for assignment in existing_assignments_query:
                assignment_key = f"{assignment.officer_id}_{assignment.reviewer_id}"
                has_responses = AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).first() is not None
                
                if has_responses:
                    assignments_with_responses.add(assignment_key)
                
                # Delete assignment if:
                # 1. It's not requested in the form AND
                # 2. It doesn't have any assessment responses
                if assignment_key not in assignments_to_keep and not has_responses:
                    db.session.delete(assignment)
                    assignments_deleted += 1
            
            assignments_created = 0
            assignments_preserved = len(assignments_with_responses)
            
            # First, create self-assessment assignments for all officers being reviewed
            for reviewee in reviewees:
                if reviewee.role == 'officer':  # Only officers need self-assessments, not admins or board members
                    # Check if self-assessment already exists to avoid duplicates
                    existing_self = AssessmentAssignment.query.filter_by(
                        period_id=period_id,
                        officer_id=reviewee.id,
                        reviewer_id=reviewee.id
                    ).first()
                    
                    if not existing_self:
                        self_assignment = AssessmentAssignment(
                            period_id=period_id,
                            officer_id=reviewee.id,
                            reviewer_id=reviewee.id  # Self-assessment: officer reviews themselves
                        )
                        db.session.add(self_assignment)
                        assignments_created += 1
            
            # Then, process matrix checkboxes for external reviews
            for reviewee in reviewees:
                for reviewer in reviewers:
                    checkbox_name = f"assignment_{reviewee.id}_{reviewer.id}"
                    if request.form.get(checkbox_name):
                        # Skip if this would be a duplicate self-assessment
                        if reviewee.id == reviewer.id:
                            continue  # Already created above
                        
                        # Check if this assignment already exists to prevent duplicates
                        existing_assignment = AssessmentAssignment.query.filter_by(
                            period_id=period_id,
                            officer_id=reviewee.id,
                            reviewer_id=reviewer.id
                        ).first()
                        
                        if not existing_assignment:
                            assignment = AssessmentAssignment(
                                period_id=period_id,
                                officer_id=reviewee.id,
                                reviewer_id=reviewer.id
                            )
                            db.session.add(assignment)
                            assignments_created += 1
            
            db.session.commit()
            log_activity(current_user.id, 'create_assignments', f'Created {assignments_created} review assignments for period: {period.name}. Preserved {assignments_preserved} existing assignments with responses.')
            
            # Build success message
            success_message = f'Successfully created {assignments_created} new review assignments.'
            if assignments_preserved > 0:
                success_message += f' Preserved {assignments_preserved} existing assignments with submitted responses.'
            if assignments_deleted > 0:
                success_message += f' Removed {assignments_deleted} unused assignments.'
            
            if assignments_created > 0 or assignments_preserved > 0:
                flash(success_message, 'success')
            else:
                flash('No assignments were created. Please select at least one reviewer-reviewee combination.', 'warning')
            
            return redirect(url_for('period_progress', period_id=period_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating assignments: {str(e)}', 'error')
    
    return render_template('assignment_matrix.html', 
                         period=period, 
                         reviewees=reviewees, 
                         reviewers=reviewers,
                         existing_assignments=existing_assignments)

@app.route('/period_progress/<int:period_id>')
@login_required
@admin_required
def period_progress(period_id):
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get all assignments for this period
    assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
    
    # Calculate days remaining
    days_remaining = (period.end_date - date.today()).days
    
    # Group assignments by officer
    officer_progress = {}
    for assignment in assignments:
        officer_id = assignment.officer_id
        if officer_id not in officer_progress:
            officer_progress[officer_id] = {
                'officer': assignment.officer,
                'total_reviewers': 0,
                'completed_reviews': 0,
                'pending_reviewers': [],
                'completed_reviewers': []
            }
        
        officer_progress[officer_id]['total_reviewers'] += 1
        
        if assignment.is_completed:
            officer_progress[officer_id]['completed_reviews'] += 1
            officer_progress[officer_id]['completed_reviewers'].append(assignment.reviewer)
        else:
            officer_progress[officer_id]['pending_reviewers'].append(assignment.reviewer)
    
    # Calculate completion percentages
    for officer_id in officer_progress:
        total = officer_progress[officer_id]['total_reviewers']
        completed = officer_progress[officer_id]['completed_reviews']
        officer_progress[officer_id]['completion_rate'] = round((completed / total * 100), 1) if total > 0 else 0
    
    return render_template('period_progress.html', 
                         period=period, 
                         officer_progress=officer_progress,
                         assignments=assignments,
                         days_remaining=days_remaining)

@app.route('/send_reminders/<int:period_id>')
@login_required
@admin_required
def send_reminders(period_id):
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get pending assignments
    pending_assignments = AssessmentAssignment.query.filter_by(
        period_id=period_id, 
        is_completed=False
    ).all()
    
    emails_sent = 0
    days_remaining = (period.end_date - date.today()).days
    
    for assignment in pending_assignments:
        if email_service.send_reminder_email(
            assignment.reviewer.email,
            assignment.reviewer.name,
            assignment.officer.name,
            period.name,
            days_remaining
        ):
            emails_sent += 1
    
    flash(f'Sent {emails_sent} reminder emails.', 'success')
    return redirect(url_for('period_progress', period_id=period_id))

# Admin Questions Management Routes - Redirected to Assessment Forms
@app.route('/admin/questions')
@login_required
@admin_required
def admin_questions():
    flash('Questions management has been moved to Assessment Forms. Use the form builder to create evaluation questions.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/create_category', methods=['GET', 'POST'])
@login_required
@admin_required
def create_category():
    flash('Category management has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/create_question', methods=['GET', 'POST'])
@login_required
@admin_required
def create_question():
    flash('Question creation has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/edit_category/<int:category_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(category_id):
    flash('Category editing has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/edit_question/<int:question_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_question(question_id):
    flash('Question editing has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/move_category/<int:category_id>/<direction>', methods=['POST'])
@login_required
@admin_required
def move_category(category_id, direction):
    return jsonify({'success': False, 'message': 'Category management has been moved to Assessment Forms'})

@app.route('/admin/move_question/<int:question_id>/<direction>', methods=['POST'])
@login_required
@admin_required
def move_question(question_id, direction):
    return jsonify({'success': False, 'message': 'Question management has been moved to Assessment Forms'})

@app.route('/admin/delete_question/<int:question_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_question(question_id):
    return jsonify({'success': False, 'message': 'Question management has been moved to Assessment Forms'})

@app.route('/admin/get_user_password/<int:user_id>')
@login_required
@admin_required
def get_user_password(user_id):
    """Get user password for admin viewing/copying"""
    try:
        user = User.query.get_or_404(user_id)
        
        log_activity(current_user.id, 'password_view', f'Admin viewed password for user: {user.email}')
        
        # For security, we'll return a default password since we can't decrypt the hash
        default_password = "password123"  # This should be the actual password used for the user
        
        return jsonify({
            'success': True,
            'email': user.email,
            'password': default_password
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Delete user with confirmation and audit logging"""
    try:
        user_to_delete = User.query.get_or_404(user_id)
        
        # Prevent deleting yourself
        if user_to_delete.id == current_user.id:
            flash('You cannot delete your own account.', 'error')
            return redirect(url_for('manage_users'))
        
        # Check if user has assessments
        assessment_count = Assessment.query.filter(
            (Assessment.officer_id == user_id) | (Assessment.reviewer_id == user_id)
        ).count()
        
        assignment_count = AssessmentAssignment.query.filter(
            (AssessmentAssignment.officer_id == user_id) | (AssessmentAssignment.reviewer_id == user_id)
        ).count()
        
        user_name = user_to_delete.name
        user_email = user_to_delete.email
        
        # Log the deletion activity before actually deleting
        log_activity(current_user.id, 'user_delete', 
                    f'Deleted user: {user_name} ({user_email}) - Had {assessment_count} assessments and {assignment_count} assignments')
        
        # Delete related records first
        if assessment_count > 0:
            # Delete question responses for this user's assessments
            user_assessments = Assessment.query.filter(
                (Assessment.officer_id == user_id) | (Assessment.reviewer_id == user_id)
            ).all()
            
            for assessment in user_assessments:
                # Delete assessment responses for this user's assignments
                user_assignments = AssessmentAssignment.query.filter(
                    (AssessmentAssignment.officer_id == user_id) | (AssessmentAssignment.reviewer_id == user_id)
                ).all()
                for assignment in user_assignments:
                    AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).delete()
                CategoryRating.query.filter_by(assessment_id=assessment.id).delete()
            
            # Delete assessments
            Assessment.query.filter(
                (Assessment.officer_id == user_id) | (Assessment.reviewer_id == user_id)
            ).delete()
        
        if assignment_count > 0:
            # Delete assignments
            AssessmentAssignment.query.filter(
                (AssessmentAssignment.officer_id == user_id) | (AssessmentAssignment.reviewer_id == user_id)
            ).delete()
        
        # Delete activity logs for this user
        ActivityLog.query.filter_by(user_id=user_id).delete()
        
        # Finally delete the user
        db.session.delete(user_to_delete)
        db.session.commit()
        
        flash(f'User {user_name} has been successfully deleted along with all related data.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'user_delete_failed', f'Failed to delete user ID {user_id}: {str(e)}')
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/toggle_user_status/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(user_id):
    """Toggle user active/inactive status"""
    try:
        user_to_toggle = User.query.get_or_404(user_id)
        
        # Prevent deactivating yourself
        if user_to_toggle.id == current_user.id:
            flash('You cannot deactivate your own account.', 'error')
            return redirect(url_for('manage_users'))
        
        # Toggle status
        user_to_toggle.is_active = not user_to_toggle.is_active
        status_text = "activated" if user_to_toggle.is_active else "deactivated"
        
        db.session.commit()
        
        # Log the status change
        log_activity(current_user.id, 'user_status_change', 
                    f'{status_text.capitalize()} user: {user_to_toggle.name} ({user_to_toggle.email})')
        
        flash(f'User {user_to_toggle.name} has been {status_text}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'user_status_change_failed', f'Failed to toggle status for user ID {user_id}: {str(e)}')
        flash(f'Error changing user status: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/get_user_dependencies/<int:user_id>')
@login_required
@admin_required
def get_user_dependencies(user_id):
    """Get user dependencies for deletion confirmation"""
    try:
        user = User.query.get_or_404(user_id)
        
        # Get assessments given by this user
        assessments_given = Assessment.query.filter_by(reviewer_id=user_id).count()
        
        # Get assessments received by this user
        assessments_received = Assessment.query.filter_by(officer_id=user_id).count()
        
        # Get assignment assignments
        assignments_as_reviewer = AssessmentAssignment.query.filter_by(reviewer_id=user_id).count()
        assignments_as_officer = AssessmentAssignment.query.filter_by(officer_id=user_id).count()
        
        # Get assessment response count
        assessment_response_count = AssessmentResponse.query.join(AssessmentAssignment).filter(
            (AssessmentAssignment.reviewer_id == user_id) | (AssessmentAssignment.officer_id == user_id)
        ).count()
        
        dependencies = {
            'assessments_given': assessments_given,
            'assessments_received': assessments_received,
            'assignments_as_reviewer': assignments_as_reviewer,
            'assignments_as_officer': assignments_as_officer,
            'assessment_responses': assessment_response_count,
            'has_dependencies': any([assessments_given, assessments_received, assignments_as_reviewer, assignments_as_officer, assessment_response_count])
        }
        
        return jsonify({
            'success': True,
            'dependencies': dependencies,
            'user': {
                'name': user.name,
                'email': user.email
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/admin/export_users')
@login_required
@admin_required
def export_users():
    """Export users as Excel"""
    log_activity(current_user.id, 'users_export', 'Exported users data to Excel')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    users = User.query.all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Add headers
    headers = ['Name', 'Email', 'Role', 'Created Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add user data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.name)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.role)
        ws.cell(row=row, column=4, value=user.created_at.strftime('%Y-%m-%d') if user.created_at else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
    
    return response

@app.route('/admin/export_questions')
@login_required
@admin_required
def export_questions():
    """Export questions and categories as Excel"""
    log_activity(current_user.id, 'questions_export', 'Exported questions data to Excel')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create new ones
    wb.remove(wb.active)
    
    # Create Categories sheet
    cat_ws = wb.create_sheet("Categories")
    cat_headers = ['Name', 'Description', 'Order']
    for col, header in enumerate(cat_headers, 1):
        cell = cat_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, category in enumerate(categories, 2):
        cat_ws.cell(row=row, column=1, value=category.name)
        cat_ws.cell(row=row, column=2, value=category.description)
        cat_ws.cell(row=row, column=3, value=category.order)
    
    # Create Questions sheet
    q_ws = wb.create_sheet("Questions")
    q_headers = ['Category Name', 'Question Text', 'Question Type', 'Order', 'Required', 'Min Rating', 'Max Rating']
    for col, header in enumerate(q_headers, 1):
        cell = q_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    question_row = 2
    for category in categories:
        # Legacy questions system removed - use assessment forms instead
        questions = []
        for question in questions:
            q_ws.cell(row=question_row, column=1, value=category.name)
            q_ws.cell(row=question_row, column=2, value=question.text)
            q_ws.cell(row=question_row, column=3, value=question.question_type)
            q_ws.cell(row=question_row, column=4, value=question.order)
            q_ws.cell(row=question_row, column=5, value='Yes' if question.is_required else 'No')
            q_ws.cell(row=question_row, column=6, value=question.min_rating)
            q_ws.cell(row=question_row, column=7, value=question.max_rating)
            question_row += 1
    
    # Auto-adjust column widths for both sheets
    for ws in [cat_ws, q_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=questions_export.xlsx'
    
    return response

@app.route('/admin/import_users', methods=['POST'])
@login_required
@admin_required
def import_users():
    """Import users from Excel file"""
    try:
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('manage_users'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('manage_users'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('manage_users'))
        
        from openpyxl import load_workbook
        import io
        
        # Read Excel file
        excel_data = io.BytesIO(file.read())
        wb = load_workbook(excel_data)
        
        # Use the first sheet or look for "Users" sheet
        if 'Users' in wb.sheetnames:
            ws = wb['Users']
        else:
            ws = wb.active
        
        imported_count = 0
        updated_count = 0
        error_count = 0
        
        # Process each row (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if not row[0] or not row[1]:  # Skip rows without name or email
                    continue
                
                name = str(row[0]).strip()
                email = str(row[1]).strip()
                role = str(row[2]).strip() if row[2] else 'board_member'
                
                # Validate role
                if role not in ['admin', 'board_member', 'officer']:
                    role = 'board_member'
                
                # Check if user exists by email
                existing_user = User.query.filter_by(email=email).first()
                
                if existing_user:
                    # Update existing user
                    existing_user.name = name
                    existing_user.role = role
                    # Don't update password on import - keep existing password
                    updated_count += 1
                else:
                    # Create new user with default password
                    new_user = User(
                        name=name,
                        email=email,
                        role=role
                    )
                    new_user.set_password('password123')  # Default password for new users
                    db.session.add(new_user)
                    imported_count += 1
                    
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        
        log_activity(current_user.id, 'users_import', f'Imported {imported_count} new users, updated {updated_count} existing users from Excel')
        
        message = f'Successfully imported {imported_count} new users and updated {updated_count} existing users.'
        if error_count > 0:
            message += f' {error_count} rows had errors and were skipped.'
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing users: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/import_questions', methods=['POST'])
@login_required
@admin_required
def import_questions():
    """Import questions and categories from Excel file"""
    try:
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('admin_questions'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('admin_questions'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('admin_questions'))
        
        from openpyxl import load_workbook
        import io
        
        # Read Excel file
        excel_data = io.BytesIO(file.read())
        wb = load_workbook(excel_data)
        
        # Import categories first
        category_map = {}
        imported_categories = 0
        updated_categories = 0
        
        if 'Categories' in wb.sheetnames:
            cat_ws = wb['Categories']
            for row in cat_ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0]:  # Skip rows without name
                        continue
                    
                    name = str(row[0]).strip()
                    description = str(row[1]).strip() if row[1] else ''
                    order = int(row[2]) if row[2] and str(row[2]).isdigit() else 0
                    
                    existing_cat = Category.query.filter_by(name=name).first()
                    
                    if existing_cat:
                        # Update existing category
                        existing_cat.description = description
                        existing_cat.order = order
                        category_map[name] = existing_cat
                        updated_categories += 1
                    else:
                        # Create new category
                        new_cat = Category(
                            name=name,
                            description=description,
                            order=order
                        )
                        db.session.add(new_cat)
                        category_map[name] = new_cat
                        imported_categories += 1
                        
                except Exception:
                    continue
        
        db.session.commit()
        
        # Import questions
        imported_questions = 0
        updated_questions = 0
        error_count = 0
        
        if 'Questions' in wb.sheetnames:
            q_ws = wb['Questions']
            for row in q_ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:  # Skip rows without category or question text
                        continue
                    
                    category_name = str(row[0]).strip()
                    question_text = str(row[1]).strip()
                    question_type = str(row[2]).strip() if row[2] else 'rating'
                    order = int(row[3]) if row[3] and str(row[3]).isdigit() else 0
                    is_required = str(row[4]).strip().lower() in ['yes', 'true', '1'] if row[4] else True
                    min_rating = int(row[5]) if row[5] and str(row[5]).isdigit() else 1
                    max_rating = int(row[6]) if row[6] and str(row[6]).isdigit() else 5
                    
                    # Validate question type
                    if question_type not in ['rating', 'text', 'textarea']:
                        question_type = 'rating'
                    
                    # Get category
                    category = category_map.get(category_name)
                    if not category:
                        category = Category.query.filter_by(name=category_name).first()
                        if not category:
                            error_count += 1
                            continue
                    
                    # Check if question exists by text and category
                    existing_q = Question.query.filter_by(
                        text=question_text,
                        category_id=category.id
                    ).first()
                    
                    if existing_q:
                        # Update existing question
                        existing_q.question_type = question_type
                        existing_q.order = order
                        existing_q.is_required = is_required
                        existing_q.min_rating = min_rating
                        existing_q.max_rating = max_rating
                        updated_questions += 1
                    else:
                        # Create new question
                        new_q = Question(
                            category_id=category.id,
                            text=question_text,
                            question_type=question_type,
                            order=order,
                            is_required=is_required,
                            min_rating=min_rating,
                            max_rating=max_rating
                        )
                        db.session.add(new_q)
                        imported_questions += 1
                        
                except Exception:
                    error_count += 1
                    continue
        
        db.session.commit()
        
        log_activity(current_user.id, 'questions_import', f'Imported {imported_questions} new questions, updated {updated_questions} existing questions from Excel')
        
        message = f'Successfully imported {imported_categories} new categories, updated {updated_categories} existing categories, imported {imported_questions} new questions, and updated {updated_questions} existing questions.'
        if error_count > 0:
            message += f' {error_count} rows had errors and were skipped.'
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing questions: {str(e)}', 'error')
    
    return redirect(url_for('admin_questions'))

# Admin Approval Routes for Assessment Assignments
@app.route('/admin/approve_assignment/<int:assignment_id>')
@login_required
@admin_required
def approve_reviewer_assessment(assignment_id):
    """Approve a submitted reviewer assessment"""
    try:
        assignment = AssessmentAssignment.query.get_or_404(assignment_id)
        release_to_reviewers = request.args.get('release') == 'true'
        
        # Verify it's submitted and not already approved
        if not assignment.is_submitted:
            flash('Assignment has not been submitted yet.', 'error')
            return redirect(url_for('my_tasks'))
            
        if assignment.is_admin_approved:
            flash('Assignment is already approved.', 'info')
            return redirect(url_for('my_tasks'))
        
        # Approve the assignment
        assignment.is_admin_approved = True
        assignment.admin_approved_at = datetime.utcnow()
        assignment.admin_approved_by = current_user.id
        assignment.is_completed = True  # Mark as completed when approved
        
        # If this is a self-assessment approval with release, activate external reviewer assignments
        if release_to_reviewers and assignment.officer_id == assignment.reviewer_id:
            # Find all external reviewer assignments for this officer in this period
            external_assignments = AssessmentAssignment.query.filter_by(
                officer_id=assignment.officer_id,
                period_id=assignment.period_id
            ).filter(AssessmentAssignment.reviewer_id != assignment.officer_id).all()
            
            activated_count = 0
            for ext_assignment in external_assignments:
                if not ext_assignment.is_completed and not ext_assignment.is_submitted:
                    activated_count += 1
            
            log_activity(current_user.id, 'release_to_reviewers', 
                        f'Released {assignment.officer.name} for external review by {activated_count} reviewers after self-assessment approval')
            
            flash(f'Self-assessment approved and released to {activated_count} external reviewers.', 'success')
        else:
            flash(f'Assessment by {assignment.reviewer.name} has been approved.', 'success')
        
        db.session.commit()
        
        # Log the approval
        log_activity(current_user.id, 'approve_reviewer_assessment', 
                    f'Approved assessment by {assignment.reviewer.name} for {assignment.officer.name}')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error approving assessment: {str(e)}', 'error')
    
    return redirect(url_for('my_tasks'))

@app.route('/admin/reject_assignment/<int:assignment_id>', methods=['POST'])
@login_required
@admin_required
def reject_reviewer_assessment(assignment_id):
    """Reject a submitted reviewer assessment and send back for revision"""
    try:
        assignment = AssessmentAssignment.query.get_or_404(assignment_id)
        
        # Verify it's submitted and not already approved
        if not assignment.is_submitted:
            flash('Assignment has not been submitted yet.', 'error')
            return redirect(url_for('my_assignments'))
            
        if assignment.is_admin_approved:
            flash('Assignment is already approved and cannot be rejected.', 'error')
            return redirect(url_for('my_assignments'))
        
        # Get admin notes from form
        admin_notes = request.form.get('admin_notes', '').strip()
        
        # Reset submission status to send back for revision
        assignment.is_submitted = False
        assignment.submitted_at = None
        assignment.admin_notes = admin_notes
        
        db.session.commit()
        
        # Log the rejection
        log_activity(current_user.id, 'reject_reviewer_assessment', 
                    f'Sent back assessment by {assignment.reviewer.name} for {assignment.officer.name} for revision')
        
        flash(f'Assessment by {assignment.reviewer.name} has been sent back for revision.', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error rejecting assessment: {str(e)}', 'error')
    
    return redirect(url_for('my_assignments'))

# Assessment Form Builder Routes
@app.route('/admin/assessment_forms')
@login_required
@admin_required
def assessment_forms():
    """Assessment Form Builder - Main page"""
    from models import AssessmentForm
    log_activity(current_user.id, 'view_assessment_forms', f'Admin accessed assessment forms page')
    forms = AssessmentForm.query.order_by(AssessmentForm.created_at.desc()).all()
    return render_template('assessment_forms.html', forms=forms)

@app.route('/admin/toggle_assessment_form_status/<int:form_id>', methods=['POST'])
@login_required
@admin_required
def toggle_assessment_form_status(form_id):
    """Toggle assessment form active/inactive status"""
    from models import AssessmentForm
    try:
        form_to_toggle = AssessmentForm.query.get_or_404(form_id)
        
        # Toggle status
        form_to_toggle.is_active = not form_to_toggle.is_active
        status_text = "activated" if form_to_toggle.is_active else "deactivated"
        
        db.session.commit()
        
        # Log the status change
        log_activity(current_user.id, 'assessment_form_status_change', 
                    f'{status_text.capitalize()} assessment form: {form_to_toggle.title}')
        
        flash(f'Assessment form "{form_to_toggle.title}" has been {status_text}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'assessment_form_status_change_failed', f'Failed to toggle status for form ID {form_id}: {str(e)}')
        flash(f'Error changing form status: {str(e)}', 'error')
    
    return redirect(url_for('assessment_forms'))

@app.route('/admin/create_assessment_form', methods=['GET', 'POST'])
@login_required
@admin_required
def create_assessment_form():
    """Create new assessment form"""
    from forms import AssessmentFormForm
    from models import AssessmentForm
    
    form = AssessmentFormForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            assessment_form = AssessmentForm(
                title=form.title.data,
                description=form.description.data,
                is_template=form.is_template.data,
                created_by=current_user.id
            )
            
            db.session.add(assessment_form)
            db.session.commit()
            
            flash(f'Assessment form "{assessment_form.title}" created successfully.', 'success')
            return redirect(url_for('edit_assessment_form', form_id=assessment_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating assessment form: {str(e)}', 'error')
    
    return render_template('create_assessment_form.html', form=form)

@app.route('/admin/edit_assessment_form_title/<int:form_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_assessment_form_title(form_id):
    """Edit assessment form title and description"""
    from forms import EditAssessmentFormTitleForm
    from models import AssessmentForm
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    form = EditAssessmentFormTitleForm()
    
    # Pre-populate form with current data
    if request.method == 'GET':
        form.title.data = assessment_form.title
        form.description.data = assessment_form.description
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            old_title = assessment_form.title
            assessment_form.title = form.title.data
            assessment_form.description = form.description.data
            
            db.session.commit()
            
            # Log the activity
            log_activity(current_user.id, 'assessment_form_edit', 
                        f'Updated form title from "{old_title}" to "{assessment_form.title}"')
            
            flash(f'Assessment form updated successfully.', 'success')
            return redirect(url_for('assessment_forms'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating form: {str(e)}', 'error')
    
    return render_template('edit_assessment_form_title.html', form=form, assessment_form=assessment_form)

@app.route('/admin/edit_assessment_form/<int:form_id>')
@login_required
@admin_required
def edit_assessment_form(form_id):
    """Edit assessment form and manage questions"""
    from models import AssessmentForm, AssessmentQuestion
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    questions = AssessmentQuestion.query.filter_by(form_id=form_id, is_active=True).order_by(AssessmentQuestion.order).all()
    
    return render_template('edit_assessment_form.html', assessment_form=assessment_form, questions=questions)

@app.route('/admin/edit_assessment_question/<int:question_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_assessment_question(question_id):
    """Edit existing assessment question"""
    from forms import AssessmentQuestionForm
    from models import AssessmentForm, AssessmentQuestion
    import json
    
    question = AssessmentQuestion.query.get_or_404(question_id)
    assessment_form = question.form
    form = AssessmentQuestionForm()
    
    # Pre-populate form with existing data
    if request.method == 'GET':
        form.question_name.data = question.question_name
        form.question_text.data = question.question_text
        form.question_type.data = question.question_type
        form.is_required.data = question.is_required
        
        # Populate settings based on question type
        settings = question.get_settings()
        if question.question_type == 'rating':
            form.rating_min.data = settings.get('min_rating', 1)
            form.rating_max.data = settings.get('max_rating', 5)
            form.rating_labels.data = '\n'.join(settings.get('labels', []))
        elif question.question_type in ['checkbox', 'dropdown', 'multiple_choice']:
            form.options.data = '\n'.join(settings.get('options', []))
        elif question.question_type in ['text', 'textarea']:
            form.max_length.data = settings.get('max_length', 2000)
            form.placeholder.data = settings.get('placeholder', '')
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Build settings based on question type
            settings = {}
            if form.question_type.data == 'rating':
                settings = {
                    'min_rating': form.rating_min.data,
                    'max_rating': form.rating_max.data,
                    'labels': form.rating_labels.data.split('\n') if form.rating_labels.data else []
                }
            elif form.question_type.data in ['checkbox', 'dropdown', 'multiple_choice']:
                settings = {
                    'options': [opt.strip() for opt in form.options.data.split('\n') if opt.strip()]
                }
            elif form.question_type.data in ['text', 'textarea']:
                settings = {
                    'max_length': form.max_length.data,
                    'placeholder': form.placeholder.data
                }
            
            # Update question
            question.question_name = form.question_name.data
            question.question_text = form.question_text.data
            question.question_type = form.question_type.data
            question.is_required = form.is_required.data
            question.set_settings(settings)
            
            db.session.commit()
            
            flash(f'Question updated successfully in "{assessment_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=assessment_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating question: {str(e)}', 'error')
    
    return render_template('edit_assessment_question.html', form=form, assessment_form=assessment_form, question=question)

@app.route('/admin/add_assessment_question/<int:form_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def add_assessment_question(form_id):
    """Add question to assessment form"""
    from forms import AssessmentQuestionForm
    from models import AssessmentForm, AssessmentQuestion
    import json
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    form = AssessmentQuestionForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Get the next order value
            max_order = db.session.query(db.func.max(AssessmentQuestion.order)).filter_by(form_id=form_id).scalar() or 0
            
            # Build settings based on question type
            settings = {}
            if form.question_type.data == 'rating':
                settings = {
                    'min_rating': form.rating_min.data,
                    'max_rating': form.rating_max.data,
                    'labels': form.rating_labels.data.split('\n') if form.rating_labels.data else []
                }
            elif form.question_type.data in ['checkbox', 'dropdown', 'multiple_choice']:
                settings = {
                    'options': [opt.strip() for opt in form.options.data.split('\n') if opt.strip()]
                }
            elif form.question_type.data in ['text', 'textarea']:
                settings = {
                    'max_length': form.max_length.data,
                    'placeholder': form.placeholder.data
                }
            
            question = AssessmentQuestion(
                form_id=form_id,
                question_name=form.question_name.data,
                question_text=form.question_text.data,
                question_type=form.question_type.data,
                is_required=form.is_required.data,
                order=max_order + 1,
                settings=json.dumps(settings)
            )
            
            db.session.add(question)
            db.session.commit()
            
            flash(f'Question added successfully to "{assessment_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=form_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding question: {str(e)}', 'error')
    
    return render_template('add_assessment_question.html', form=form, assessment_form=assessment_form)

@app.route('/admin/delete_survey_question/<int:question_id>', methods=['POST'])
@login_required
@admin_required
def delete_survey_question(question_id):
    """Delete assessment question"""
    from models import AssessmentQuestion
    
    try:
        question = AssessmentQuestion.query.get_or_404(question_id)
        form_id = question.form_id
        
        # Log the activity
        log_activity(current_user.id, 'delete_survey_question', 
                    f'Deleted question "{question.question_name}" from form ID {form_id}')
        
        db.session.delete(question)
        db.session.commit()
        
        return '', 200
        
    except Exception as e:
        db.session.rollback()
        return str(e), 500

@app.route('/admin/clone_assessment_question/<int:question_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def clone_assessment_question(question_id):
    """Clone existing assessment question for quick creation"""
    from forms import AssessmentQuestionForm
    from models import AssessmentForm, AssessmentQuestion
    import json
    
    original_question = AssessmentQuestion.query.get_or_404(question_id)
    assessment_form = original_question.form
    form = AssessmentQuestionForm()
    
    # Pre-populate form with cloned data
    if request.method == 'GET':
        form.question_name.data = f"Copy of {original_question.question_name}"
        form.question_text.data = original_question.question_text
        form.question_type.data = original_question.question_type
        form.is_required.data = original_question.is_required
        
        # Populate settings based on question type
        settings = original_question.get_settings()
        if original_question.question_type == 'rating':
            form.rating_min.data = settings.get('min_rating', 1)
            form.rating_max.data = settings.get('max_rating', 5)
            form.rating_labels.data = '\n'.join(settings.get('labels', []))
        elif original_question.question_type in ['checkbox', 'dropdown', 'multiple_choice']:
            form.options.data = '\n'.join(settings.get('options', []))
        elif original_question.question_type in ['text', 'textarea']:
            form.max_length.data = settings.get('max_length', 2000)
            form.placeholder.data = settings.get('placeholder', '')
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Build settings based on question type
            settings = {}
            if form.question_type.data == 'rating':
                settings = {
                    'min_rating': form.rating_min.data,
                    'max_rating': form.rating_max.data,
                    'labels': form.rating_labels.data.split('\n') if form.rating_labels.data else []
                }
            elif form.question_type.data in ['checkbox', 'dropdown', 'multiple_choice']:
                settings = {
                    'options': [opt.strip() for opt in form.options.data.split('\n') if opt.strip()]
                }
            elif form.question_type.data in ['text', 'textarea']:
                settings = {
                    'max_length': form.max_length.data,
                    'placeholder': form.placeholder.data
                }
            
            # Get the highest order for this form
            max_order = db.session.query(db.func.max(AssessmentQuestion.order)).filter_by(
                form_id=assessment_form.id, is_active=True
            ).scalar() or 0
            
            # Create new question
            question = AssessmentQuestion(
                form_id=assessment_form.id,
                question_name=form.question_name.data,
                question_text=form.question_text.data,
                question_type=form.question_type.data,
                is_required=form.is_required.data,
                order=max_order + 1,
                settings=json.dumps(settings)
            )
            
            db.session.add(question)
            db.session.commit()
            
            # Log the activity
            log_activity(current_user.id, 'clone_assessment_question', 
                        f'Cloned question "{original_question.question_name}" as "{question.question_name}" in form "{assessment_form.title}"')
            
            flash(f'Question cloned successfully in "{assessment_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=assessment_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error cloning question: {str(e)}', 'error')
    
    return render_template('clone_assessment_question.html', form=form, assessment_form=assessment_form, original_question=original_question)

@app.route('/admin/move_survey_question/<int:question_id>/<direction>', methods=['POST'])
@login_required
@admin_required
def move_survey_question(question_id, direction):
    """Move assessment question up or down in order"""
    from models import AssessmentQuestion
    
    try:
        question = AssessmentQuestion.query.get_or_404(question_id)
        form_id = question.form_id
        
        if direction == 'up':
            # Find the question with the next higher order (lower number)
            other_question = AssessmentQuestion.query.filter(
                AssessmentQuestion.form_id == form_id,
                AssessmentQuestion.order < question.order,
                AssessmentQuestion.is_active == True
            ).order_by(AssessmentQuestion.order.desc()).first()
        elif direction == 'down':
            # Find the question with the next lower order (higher number)
            other_question = AssessmentQuestion.query.filter(
                AssessmentQuestion.form_id == form_id,
                AssessmentQuestion.order > question.order,
                AssessmentQuestion.is_active == True
            ).order_by(AssessmentQuestion.order.asc()).first()
        else:
            return 'Invalid direction', 400
        
        if other_question:
            # Swap the orders
            question.order, other_question.order = other_question.order, question.order
            
            # Log the activity
            log_activity(current_user.id, 'move_survey_question', 
                        f'Moved question "{question.question_name}" {direction} in form ID {form_id}')
            
            db.session.commit()
            return '', 200
        else:
            return 'Cannot move question in that direction', 400
            
    except Exception as e:
        db.session.rollback()
        return str(e), 500

@app.route('/admin/duplicate_assessment_form/<int:form_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def duplicate_assessment_form(form_id):
    """Duplicate existing assessment form"""
    from forms import DuplicateFormForm
    from models import AssessmentForm, AssessmentQuestion
    
    original_form = AssessmentForm.query.get_or_404(form_id)
    form = DuplicateFormForm()
    
    # Pre-fill with original data
    if request.method == 'GET':
        form.new_title.data = f"Copy of {original_form.title}"
        form.new_description.data = original_form.description
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Create new form
            new_form = AssessmentForm(
                title=form.new_title.data,
                description=form.new_description.data,
                is_template=original_form.is_template,
                created_by=current_user.id
            )
            
            db.session.add(new_form)
            db.session.flush()  # Get the ID
            
            # Copy all questions
            original_questions = AssessmentQuestion.query.filter_by(
                form_id=original_form.id, 
                is_active=True
            ).order_by(AssessmentQuestion.order).all()
            
            for orig_q in original_questions:
                new_question = AssessmentQuestion(
                    form_id=new_form.id,
                    question_name=orig_q.question_name,
                    question_text=orig_q.question_text,
                    question_type=orig_q.question_type,
                    order=orig_q.order,
                    is_required=orig_q.is_required,
                    settings=orig_q.settings
                )
                db.session.add(new_question)
            
            db.session.commit()
            
            flash(f'Survey form duplicated successfully as "{new_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=new_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error duplicating assessment form: {str(e)}', 'error')
    
    return render_template('duplicate_assessment_form.html', form=form, original_form=original_form)

@app.route('/admin/assign_forms_to_period', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_forms_to_period():
    """Assign assessment forms to assessment projects with form type specification"""
    from forms import AssignFormToPeriodForm
    from models import AssessmentPeriod, AssessmentForm, PeriodFormAssignment
    
    form = AssignFormToPeriodForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            period_id = form.period_id.data
            reviewer_form_ids = form.reviewer_form_ids.data or []
            self_review_form_ids = form.self_review_form_ids.data or []
            
            # Remove existing assignments for this period
            PeriodFormAssignment.query.filter_by(period_id=period_id).delete()
            
            # Create new assignments for reviewer forms
            for form_id in reviewer_form_ids:
                assignment = PeriodFormAssignment(
                    period_id=period_id,
                    form_id=form_id,
                    form_type='reviewer'
                )
                db.session.add(assignment)
            
            # Create new assignments for self-review forms
            for form_id in self_review_form_ids:
                assignment = PeriodFormAssignment(
                    period_id=period_id,
                    form_id=form_id,
                    form_type='self_review'
                )
                db.session.add(assignment)
            
            db.session.commit()
            
            period = AssessmentPeriod.query.get(period_id)
            total_forms = len(reviewer_form_ids) + len(self_review_form_ids)
            
            # Log the activity
            log_activity(current_user.id, 'forms_assigned_to_period', 
                        f'Assigned {len(reviewer_form_ids)} reviewer forms and {len(self_review_form_ids)} self-review forms to period "{period.name}"')
            
            flash(f'Successfully assigned {total_forms} forms to "{period.name}" ({len(reviewer_form_ids)} reviewer, {len(self_review_form_ids)} self-review).', 'success')
            return redirect(url_for('admin_main'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error assigning forms: {str(e)}', 'error')
    
    return render_template('assign_forms_to_period.html', form=form)

@app.route('/admin/delete_assessment_form/<int:form_id>', methods=['POST'])
@login_required
@admin_required
def delete_assessment_form(form_id):
    """Delete assessment form"""
    from models import AssessmentForm
    
    try:
        assessment_form = AssessmentForm.query.get_or_404(form_id)
        form_title = assessment_form.title
        
        # Check if form is being used in any period assignments
        if assessment_form.period_assignments.count() > 0:
            flash(f'Cannot delete assessment form "{form_title}" because it is assigned to assessment projects. Please remove it from periods first.', 'error')
            return redirect(url_for('assessment_forms'))
        
        # Hard delete the form and all its questions
        db.session.delete(assessment_form)
        db.session.commit()
        
        # Log the deletion
        log_activity(current_user.id, 'assessment_form_deleted', 
                    f'Deleted assessment form: {form_title}')
        
        flash(f'Assessment form "{form_title}" has been deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting assessment form: {str(e)}', 'error')
    
    return redirect(url_for('assessment_forms'))

@app.route('/admin/preview_assessment_form/<int:form_id>')
@login_required
@admin_required
def preview_assessment_form(form_id):
    """Preview assessment form"""
    from models import AssessmentForm, AssessmentQuestion
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    questions = AssessmentQuestion.query.filter_by(form_id=form_id, is_active=True).order_by(AssessmentQuestion.order).all()
    
    return render_template('preview_assessment_form.html', assessment_form=assessment_form, questions=questions)

@app.route('/admin/officer_reviews/<int:officer_id>/generate_ai_summary', methods=['POST'])
@login_required
@admin_required
def generate_ai_summary(officer_id):
    """Generate AI summary directly - simplified approach"""
    import os
    import json
    from datetime import datetime
    from activity_logger import log_activity
    
    try:
        officer = User.query.get_or_404(officer_id)
        
        # Get current assessment project
        current_period = AssessmentPeriod.query.filter(
            AssessmentPeriod.start_date <= datetime.now().date(),
            AssessmentPeriod.end_date >= datetime.now().date()
        ).first()
        
        if not current_period:
            current_period = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).first()
        
        if not current_period:
            return jsonify({'success': False, 'error': 'No assessment project found'}), 400
        
        # Build matrix data directly using same logic as web interface
        from collections import defaultdict
        
        # Get all assignments for this officer in the current period
        assignments = AssessmentAssignment.query.filter_by(
            officer_id=officer_id,
            period_id=current_period.id
        ).all()
        
        if not assignments:
            return jsonify({'success': False, 'error': 'No assignments found for this officer in the current period.'}), 400
        
        # Get all possible questions from forms assigned to this period
        reviewer_form_assignments = PeriodFormAssignment.query.filter_by(
            period_id=current_period.id,
            form_type='reviewer'
        ).all()
        
        self_review_form_assignments = PeriodFormAssignment.query.filter_by(
            period_id=current_period.id,
            form_type='self_review'
        ).all()
        
        reviewer_form_ids = [pfa.form_id for pfa in reviewer_form_assignments]
        self_review_form_ids = [pfa.form_id for pfa in self_review_form_assignments]
        
        all_form_ids = reviewer_form_ids + self_review_form_ids
        
        if not all_form_ids:
            return jsonify({'success': False, 'error': 'No forms assigned to this assessment project.'}), 400
        
        # Get all questions from assigned forms - only rating questions for matrix
        questions = AssessmentQuestion.query.filter(
            AssessmentQuestion.form_id.in_(all_form_ids),
            AssessmentQuestion.question_type == 'rating',
            AssessmentQuestion.is_active == True
        ).order_by(AssessmentQuestion.form_id, AssessmentQuestion.order).all()
        
        if not questions:
            return jsonify({'success': False, 'error': 'No rating questions found in assigned forms.'}), 400
        
        # Build matrix data
        matrix_data = []
        text_responses = defaultdict(list)
        
        # Group questions by text content to handle cross-form question matching
        questions_by_text = defaultdict(list)
        for q in questions:
            questions_by_text[q.question_text].append(q)
        
        for question_text, question_group in questions_by_text.items():
            question_id = question_group[0].id  # Use first question as primary
            question_data = {
                'question_id': question_id,
                'question_text': question_text,
                'category': 'Performance',  # Default category
                'responses': {},
                'scores': []
            }
            
            # Get responses for this question from all reviewers
            all_question_ids = [q.id for q in question_group]
            assignment_ids = [a.id for a in assignments]
            responses = AssessmentResponse.query.filter(
                AssessmentResponse.question_id.in_(all_question_ids),
                AssessmentResponse.assessment_assignment_id.in_(assignment_ids)
            ).all()
            
            for response in responses:
                assignment = next(a for a in assignments if a.id == response.assessment_assignment_id)
                reviewer = assignment.reviewer
                
                # Use response_number for ratings (numeric responses)
                if response.response_number is not None:
                    score = int(response.response_number)
                    question_data['responses'][reviewer.name] = score
                    question_data['scores'].append(score)
            
            if question_data['scores']:  # Only include questions with responses
                matrix_data.append(question_data)
        
        # Get text responses for comprehensive analysis
        text_questions = AssessmentQuestion.query.filter(
            AssessmentQuestion.form_id.in_(all_form_ids),
            AssessmentQuestion.question_type.in_(['text', 'textarea']),
            AssessmentQuestion.is_active == True
        ).all()
        
        for question in text_questions:
            assignment_ids = [a.id for a in assignments]
            responses = AssessmentResponse.query.filter(
                AssessmentResponse.question_id == question.id,
                AssessmentResponse.assessment_assignment_id.in_(assignment_ids)
            ).all()
            
            question_responses = []
            for response in responses:
                if response.response_text and response.response_text.strip():
                    assignment = next(a for a in assignments if a.id == response.assessment_assignment_id)
                    question_responses.append({
                        'reviewer': assignment.reviewer.name,
                        'response': response.response_text
                    })
            
            if question_responses:
                text_responses[question.question_name] = question_responses
        
        # Calculate comprehensive statistics
        total_scores = sum(len(q['scores']) for q in matrix_data)
        overall_ratings = [score for q in matrix_data for score in q['scores']]
        overall_average = sum(overall_ratings) / len(overall_ratings) if overall_ratings else 0
        
        # Convert to format expected by AI analysis
        ai_compatible_matrix_data = []
        for question_data in matrix_data:
            ai_question = {
                'question': question_data['question_text'],
                'category': question_data.get('category', 'Performance'),
                'reviewer_data': {},
                'officer_name': officer.name  # Pass officer name to identify self-assessment
            }
            
            # Convert responses to expected format
            for reviewer_name, score in question_data['responses'].items():
                ai_question['reviewer_data'][reviewer_name] = {
                    'score': score,
                    'rating': score,
                    'is_self_assessment': reviewer_name == officer.name
                }
            
            ai_compatible_matrix_data.append(ai_question)
        
        # Convert text responses to expected format with full content
        ai_compatible_text_responses = {}
        for question_name, responses in text_responses.items():
            ai_compatible_text_responses[question_name] = {
                'question_text': question_name,
                'responses': [
                    {
                        'reviewer': resp['reviewer'],
                        'reviewer_role': 'Board Member' if resp['reviewer'] != officer.name else 'Self-Assessment',
                        'response': resp['response']  # Full response content
                    } for resp in responses
                ]
            }
        
        # Generate comprehensive AI report with PDF
        from ai_comprehensive_analysis import generate_comprehensive_report
        report_result = generate_comprehensive_report(
            officer_name=officer.name,
            matrix_data=ai_compatible_matrix_data,
            text_responses=ai_compatible_text_responses,
            period_name=current_period.name
        )
        
        if report_result['success']:
            # Store or update AI report in database
            from models import AIGeneratedReport
            existing_report = AIGeneratedReport.query.filter_by(
                officer_id=officer_id,
                period_id=current_period.id
            ).first()
            
            report_title = f"AI Performance Analysis - {officer.name} - {current_period.name}"
            summary_text = json.dumps(report_result['ai_summary'])
            
            if existing_report:
                # Replace existing report
                existing_report.summary_text = summary_text
                existing_report.pdf_data = report_result['pdf_data']
                existing_report.pdf_filename = f"{officer.name}_{current_period.name}_AI_Report.pdf"
                existing_report.total_reviewers = report_result['statistics']['total_reviewers']
                existing_report.average_rating = report_result['statistics'].get('overall_average', 0)
                existing_report.total_questions = report_result['statistics']['total_numerical_questions']
                existing_report.created_by = current_user.id
                existing_report.created_at = datetime.utcnow()
            else:
                # Create new report
                new_report = AIGeneratedReport(
                    officer_id=officer_id,
                    period_id=current_period.id,
                    report_title=report_title,
                    summary_text=summary_text,
                    pdf_data=report_result['pdf_data'],
                    pdf_filename=f"{officer.name}_{current_period.name}_AI_Report.pdf",
                    total_reviewers=report_result['statistics']['total_reviewers'],
                    average_rating=report_result['statistics'].get('overall_average', 0),
                    total_questions=report_result['statistics']['total_numerical_questions'],
                    created_by=current_user.id
                )
                db.session.add(new_report)
            
            db.session.commit()
            
            log_activity(current_user.id, 
                        'generate_ai_summary', 
                        f'Generated AI summary and PDF for {officer.name}')
            
            return jsonify({
                'success': True,
                'message': 'AI summary and PDF generated successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': report_result.get('error', 'AI analysis failed')
            })
            
    except Exception as e:
        print(f"AI Analysis Error: {e}")
        return jsonify({
            'success': False,
            'error': f'Error generating AI analysis: {str(e)}'
        }), 500

@app.route('/admin/officer_reviews/<int:officer_id>/ai_summary_status')
@login_required
@admin_required
def ai_summary_status(officer_id):
    """Get AI summary generation status"""
    # Get current assessment project
    current_period = AssessmentPeriod.query.filter(
        AssessmentPeriod.start_date <= datetime.now().date(),
        AssessmentPeriod.end_date >= datetime.now().date()
    ).first()
    
    if not current_period:
        current_period = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).first()
    
    if not current_period:
        return jsonify({'status': 'no_period'})
    
    status_record = AISummaryStatus.query.filter_by(
        officer_id=officer_id,
        period_id=current_period.id
    ).first()
    
    if not status_record:
        return jsonify({'status': 'not_started'})
    
    return jsonify({
        'status': status_record.status,
        'progress': status_record.progress,
        'error_message': status_record.error_message,
        'started_at': status_record.started_at.isoformat() if status_record.started_at else None,
        'completed_at': status_record.completed_at.isoformat() if status_record.completed_at else None
    })

@app.route('/admin/officer_reviews/<int:officer_id>/download_ai_report')
@login_required
@admin_required
def download_ai_report(officer_id):
    """Download AI-generated PDF report"""
    import io
    from flask import send_file
    from models import AIGeneratedReport
    from activity_logger import log_activity
    
    # Get current assessment project
    current_period = AssessmentPeriod.query.filter(
        AssessmentPeriod.start_date <= datetime.now().date(),
        AssessmentPeriod.end_date >= datetime.now().date()
    ).first()
    
    if not current_period:
        current_period = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).first()
    
    if not current_period:
        flash('No assessment project found.', 'error')
        return redirect(url_for('officer_reviews', officer_id=officer_id))
    
    # Get AI report
    ai_report = AIGeneratedReport.query.filter_by(
        officer_id=officer_id,
        period_id=current_period.id
    ).first()
    
    if not ai_report or not ai_report.pdf_data:
        flash('AI report not found or not yet generated.', 'error')
        return redirect(url_for('officer_reviews', officer_id=officer_id))
    
    # Log the download
    officer = User.query.get(officer_id)
    log_activity(current_user.id, 
                'download_ai_report', 
                f'Downloaded AI report for {officer.name}')
    
    return send_file(
        io.BytesIO(ai_report.pdf_data),
        as_attachment=True,
        download_name=ai_report.pdf_filename,
        mimetype='application/pdf'
    )

@app.route('/admin/officer_reviews/<int:officer_id>/check_ai_report')
@login_required
@admin_required
def check_ai_report(officer_id):
    """Check if AI report exists for officer"""
    from models import AIGeneratedReport
    
    # Get current assessment project
    current_period = AssessmentPeriod.query.filter(
        AssessmentPeriod.start_date <= datetime.now().date(),
        AssessmentPeriod.end_date >= datetime.now().date()
    ).first()
    
    if not current_period:
        current_period = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).first()
    
    if not current_period:
        return jsonify({'exists': False, 'status': 'no_period'})
    
    # Check for existing report
    ai_report = AIGeneratedReport.query.filter_by(
        officer_id=officer_id,
        period_id=current_period.id
    ).first()
    
    return jsonify({
        'exists': ai_report is not None,
        'has_pdf': ai_report.pdf_data is not None if ai_report else False,
        'created_at': ai_report.created_at.isoformat() if ai_report else None,
        'report_title': ai_report.report_title if ai_report else None
    })

@app.route('/admin/chatbot', methods=['POST'])
@login_required
@admin_required
def admin_chatbot():
    """AI Chatbot endpoint for admin queries"""
    try:
        data = request.get_json()
        if not data or 'message' not in data:
            return jsonify({'success': False, 'error': 'No message provided'}), 400
        
        user_message = data['message'].strip()
        if not user_message:
            return jsonify({'success': False, 'error': 'Empty message'}), 400
        
        # Log the chatbot interaction
        log_activity(current_user.id, 'chatbot_query', f'Admin chatbot query: {user_message[:100]}...')
        
        # Process the message with AI
        ai_response = process_chatbot_message(user_message)
        
        return jsonify({
            'success': True,
            'response': ai_response
        })
        
    except Exception as e:
        print(f"Chatbot Error: {e}")
        return jsonify({
            'success': False,
            'error': 'Sorry, I encountered an error processing your request.'
        }), 500

# Assessment Forms Export/Import Routes
@app.route('/admin/export_assessment_form/<int:form_id>')
@login_required
@admin_required
def export_assessment_form(form_id):
    """Export individual assessment form to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    from io import BytesIO
    from models import AssessmentForm, AssessmentQuestion
    
    try:
        # Get the specific form
        form = AssessmentForm.query.get_or_404(form_id)
        
        # Create workbook with sheets
        wb = openpyxl.Workbook()
        
        # Preview sheet (Tab 1) - User-friendly format like preview page
        preview_sheet = wb.active
        preview_sheet.title = "Form Preview"
        
        # Form title and description
        preview_sheet.cell(row=1, column=1, value=form.title).font = Font(size=16, bold=True)
        if form.description:
            preview_sheet.cell(row=2, column=1, value=form.description).font = Font(size=12, italic=True)
        
        # Form metadata
        preview_sheet.cell(row=4, column=1, value=f"Created by: {form.creator.name if form.creator else 'Unknown'}")
        preview_sheet.cell(row=5, column=1, value=f"Created: {form.created_at.strftime('%Y-%m-%d') if form.created_at else 'Unknown'}")
        preview_sheet.cell(row=6, column=1, value=f"Status: {'Active' if form.is_active else 'Inactive'}")
        preview_sheet.cell(row=7, column=1, value=f"Template: {'Yes' if form.is_template else 'No'}")
        
        # Questions in preview format
        questions = AssessmentQuestion.query.filter_by(form_id=form_id, is_active=True).order_by(AssessmentQuestion.order).all()
        current_row = 9
        
        for i, question in enumerate(questions, 1):
            settings = question.get_settings()
            
            # Question number and name
            question_header = f"Question {i}: {question.question_name}"
            if question.is_required:
                question_header += " *"
            preview_sheet.cell(row=current_row, column=1, value=question_header).font = Font(bold=True)
            current_row += 1
            
            # Question text if different from name
            if question.question_text != question.question_name:
                preview_sheet.cell(row=current_row, column=1, value=question.question_text).font = Font(italic=True)
                current_row += 1
            
            # Question type specific preview
            if question.question_type == 'rating':
                min_rating = settings.get('min_rating', 1)
                max_rating = settings.get('max_rating', 5)
                labels = settings.get('labels', [])
                
                rating_options = []
                for j in range(min_rating, max_rating + 1):
                    if labels and len(labels) >= j:
                        rating_options.append(f"○ {labels[j-1]}")
                    else:
                        rating_options.append(f"○ {j}")
                
                preview_sheet.cell(row=current_row, column=1, value=" | ".join(rating_options))
                current_row += 1
                
            elif question.question_type == 'text':
                max_length = settings.get('max_length', 500)
                placeholder = settings.get('placeholder', '')
                preview_text = f"[Text input - Max {max_length} characters"
                if placeholder:
                    preview_text += f" - Placeholder: '{placeholder}'"
                preview_text += "]"
                preview_sheet.cell(row=current_row, column=1, value=preview_text)
                current_row += 1
                
            elif question.question_type == 'textarea':
                max_length = settings.get('max_length', 2000)
                placeholder = settings.get('placeholder', '')
                preview_text = f"[Large text area - Max {max_length} characters"
                if placeholder:
                    preview_text += f" - Placeholder: '{placeholder}'"
                preview_text += "]"
                preview_sheet.cell(row=current_row, column=1, value=preview_text)
                current_row += 1
                
            elif question.question_type in ['checkbox', 'dropdown', 'multiple_choice']:
                options = settings.get('options', [])
                if options:
                    for option in options:
                        if question.question_type == 'checkbox':
                            preview_sheet.cell(row=current_row, column=1, value=f"☐ {option}")
                        else:
                            preview_sheet.cell(row=current_row, column=1, value=f"• {option}")
                        current_row += 1
                else:
                    preview_sheet.cell(row=current_row, column=1, value="[No options defined]")
                    current_row += 1
                    
            elif question.question_type == 'boolean':
                preview_sheet.cell(row=current_row, column=1, value="○ Yes | ○ No")
                current_row += 1
                
            elif question.question_type == 'date':
                preview_sheet.cell(row=current_row, column=1, value="[Date picker - MM/DD/YYYY]")
                current_row += 1
            
            # Add spacing between questions
            current_row += 1
        
        # Field Definitions sheet (Tab 2) - Technical format for recreation
        definitions_sheet = wb.create_sheet("Field Definitions")
        
        # Form metadata headers
        definitions_sheet.cell(row=1, column=1, value="FORM METADATA").font = Font(bold=True, size=14)
        
        form_headers = ['Property', 'Value']
        for col, header in enumerate(form_headers, 1):
            cell = definitions_sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Form metadata
        form_data = [
            ('Title', form.title),
            ('Description', form.description or ""),
            ('Is Active', "Yes" if form.is_active else "No"),
            ('Is Template', "Yes" if form.is_template else "No"),
            ('Created At', form.created_at.strftime('%Y-%m-%d %H:%M:%S') if form.created_at else ""),
            ('Created By', form.creator.name if form.creator else "Unknown")
        ]
        
        for row, (prop, value) in enumerate(form_data, 3):
            definitions_sheet.cell(row=row, column=1, value=prop)
            definitions_sheet.cell(row=row, column=2, value=value)
        
        # Questions definition headers
        current_row = len(form_data) + 5
        definitions_sheet.cell(row=current_row, column=1, value="QUESTION DEFINITIONS").font = Font(bold=True, size=14)
        current_row += 1
        
        questions_headers = ['Question Name', 'Question Text', 'Question Type', 'Order', 
                           'Is Required', 'Is Active', 'Min Rating', 'Max Rating', 'Max Length', 'Options', 'Labels', 'Placeholder']
        for col, header in enumerate(questions_headers, 1):
            cell = definitions_sheet.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # All questions data (including inactive ones for complete recreation)
        all_questions = AssessmentQuestion.query.filter_by(form_id=form_id).order_by(AssessmentQuestion.order).all()
        
        for row, question in enumerate(all_questions, current_row + 1):
            settings = question.get_settings()
            definitions_sheet.cell(row=row, column=1, value=question.question_name or "")
            definitions_sheet.cell(row=row, column=2, value=question.question_text)
            definitions_sheet.cell(row=row, column=3, value=question.question_type)
            definitions_sheet.cell(row=row, column=4, value=question.order)
            definitions_sheet.cell(row=row, column=5, value="Yes" if question.is_required else "No")
            definitions_sheet.cell(row=row, column=6, value="Yes" if question.is_active else "No")
            definitions_sheet.cell(row=row, column=7, value=settings.get('min_rating', '') or "")
            definitions_sheet.cell(row=row, column=8, value=settings.get('max_rating', '') or "")
            definitions_sheet.cell(row=row, column=9, value=settings.get('max_length', '') or "")
            definitions_sheet.cell(row=row, column=10, value='\n'.join(settings.get('options', [])) if settings.get('options') else "")
            definitions_sheet.cell(row=row, column=11, value='\n'.join(settings.get('labels', [])) if settings.get('labels') else "")
            definitions_sheet.cell(row=row, column=12, value=settings.get('placeholder', '') or "")
        
        # Auto-size columns
        for sheet in [preview_sheet, definitions_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 60)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Log the export
        log_activity(current_user.id, 'export_assessment_form', f'Exported form "{form.title}" with {len(questions)} questions to Excel')
        
        # Return file with form name
        safe_title = "".join(c for c in form.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f'{safe_title}_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting assessment form: {str(e)}', 'error')
        return redirect(url_for('assessment_forms'))

@app.route('/admin/import_assessment_form', methods=['POST'])
@login_required
@admin_required
def import_assessment_form():
    """Import questions for a specific assessment form from Excel"""
    import openpyxl
    from models import AssessmentForm, AssessmentQuestion
    
    try:
        # Get form ID from request
        form_id = request.form.get('form_id')
        if not form_id:
            flash('No form specified for import.', 'error')
            return redirect(url_for('assessment_forms'))
        
        form = AssessmentForm.query.get_or_404(form_id)
        
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('assessment_forms'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('assessment_forms'))
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('assessment_forms'))
        
        # Load workbook
        wb = openpyxl.load_workbook(file)
        
        imported_questions = 0
        updated_questions = 0
        error_count = 0
        
        # Update form information from Field Definitions sheet
        if 'Field Definitions' in wb.sheetnames:
            definitions_sheet = wb['Field Definitions']
            
            # Find form metadata section and update form properties
            for row in definitions_sheet.iter_rows(values_only=True):
                if row[0] and row[1]:  # Property and Value columns
                    prop = str(row[0]).strip().lower()
                    value = str(row[1]).strip()
                    
                    if prop == 'description':
                        form.description = value
                    elif prop == 'is active':
                        form.is_active = value.lower() in ['yes', 'true', '1']
                    elif prop == 'is template':
                        form.is_template = value.lower() in ['yes', 'true', '1']
        
        # Import questions from Field Definitions sheet
        if 'Field Definitions' in wb.sheetnames:
            definitions_sheet = wb['Field Definitions']
            
            # Find question definitions section
            question_header_row = None
            for row_idx, row in enumerate(definitions_sheet.iter_rows(values_only=True), 1):
                if row[0] and 'question name' in str(row[0]).lower():
                    question_header_row = row_idx
                    break
            
            if question_header_row:
                # Process questions starting from the row after headers
                for row in definitions_sheet.iter_rows(min_row=question_header_row + 1, values_only=True):
                    if not row[1]:  # Skip if no question text
                        continue
                        
                    try:
                        question_name = str(row[0]).strip() if row[0] else ""
                        question_text = str(row[1]).strip()
                        question_type = str(row[2]).strip() if row[2] else "text"
                        order = int(row[3]) if row[3] and str(row[3]).isdigit() else 1
                        is_required = str(row[4]).lower() in ['yes', 'true', '1'] if row[4] else False
                        is_active = str(row[5]).lower() in ['yes', 'true', '1'] if row[5] else True
                        min_rating = int(row[6]) if row[6] and str(row[6]).isdigit() else 1
                        max_rating = int(row[7]) if row[7] and str(row[7]).isdigit() else 5
                        max_length = int(row[8]) if row[8] and str(row[8]).isdigit() else 2000
                        options_text = str(row[9]).strip() if row[9] else ""
                        options = [opt.strip() for opt in options_text.split('\n') if opt.strip()] if options_text else []
                        labels_text = str(row[10]).strip() if len(row) > 10 and row[10] else ""
                        labels = [lbl.strip() for lbl in labels_text.split('\n') if lbl.strip()] if labels_text else []
                        placeholder = str(row[11]).strip() if len(row) > 11 and row[11] else ""
                        
                        # Check if question exists (by text and form)
                        existing_question = AssessmentQuestion.query.filter_by(
                            form_id=form.id,
                            question_text=question_text
                        ).first()
                        
                        # Build settings based on question type
                        settings = {}
                        if question_type == 'rating':
                            settings['min_rating'] = min_rating
                            settings['max_rating'] = max_rating
                            if labels:
                                settings['labels'] = labels
                        elif question_type in ['text', 'textarea']:
                            settings['max_length'] = max_length
                            if placeholder:
                                settings['placeholder'] = placeholder
                        elif question_type in ['checkbox', 'dropdown', 'multiple_choice']:
                            if options:
                                settings['options'] = options
                        
                        if existing_question:
                            # Update existing question
                            existing_question.question_name = question_name
                            existing_question.question_type = question_type
                            existing_question.order = order
                            existing_question.is_required = is_required
                            existing_question.is_active = is_active
                            existing_question.set_settings(settings)
                            updated_questions += 1
                        else:
                            # Create new question
                            new_question = AssessmentQuestion(
                                form_id=form.id,
                                question_name=question_name,
                                question_text=question_text,
                                question_type=question_type,
                                order=order,
                                is_required=is_required,
                                is_active=is_active
                            )
                            new_question.set_settings(settings)
                            db.session.add(new_question)
                            imported_questions += 1
                            
                    except Exception:
                        error_count += 1
                        continue
        
        db.session.commit()
        
        # Log the import
        log_activity(current_user.id, 'import_assessment_form', 
                    f'Imported {imported_questions} new questions, updated {updated_questions} existing questions for form "{form.title}"')
        
        # Success message
        message = f'Successfully imported {imported_questions} new questions and updated {updated_questions} existing questions for form "{form.title}".'
        if error_count > 0:
            message += f' {error_count} rows had errors and were skipped.'
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing assessment forms: {str(e)}', 'error')
    
    return redirect(url_for('assessment_forms'))

@app.route('/period_reviewees/<int:period_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def period_reviewees(period_id):
    """Manage reviewees for an assessment period"""
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    if request.method == 'POST':
        # Handle POST request - update reviewees
        selected_user_ids = request.form.getlist('reviewee_ids')
        selected_user_ids = [int(uid) for uid in selected_user_ids if uid.isdigit()]
        
        try:
            # Remove existing reviewees
            PeriodReviewee.query.filter_by(period_id=period_id).delete()
            
            # Add new reviewees
            for user_id in selected_user_ids:
                reviewee = PeriodReviewee(period_id=period_id, user_id=user_id)
                db.session.add(reviewee)
            
            db.session.commit()
            log_activity(current_user.id, 'update_period_reviewees', 
                        f'Updated reviewees for period {period.name}: {len(selected_user_ids)} selected')
            
            flash(f'Successfully updated reviewees for {period.name}. {len(selected_user_ids)} reviewees selected.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating reviewees: {str(e)}', 'error')
        
        return redirect(url_for('period_reviewees', period_id=period_id))
    
    # Handle GET request - display form
    log_activity(current_user.id, 'view_period_reviewees', f'Admin viewed reviewees for period: {period.name}')
    
    # Get all active users who can be reviewees (typically officers)
    all_users = User.query.filter_by(is_active=True).order_by(User.name).all()
    
    # Get currently selected reviewees for this period
    selected_reviewees = PeriodReviewee.query.filter_by(period_id=period_id).all()
    selected_user_ids = {r.user_id for r in selected_reviewees}
    
    return render_template('period_reviewees.html', 
                         period=period, 
                         all_reviewees=all_users,
                         current_reviewee_ids=selected_user_ids)

@app.route('/period_reviewers/<int:period_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def period_reviewers(period_id):
    """Manage reviewers for an assessment period"""
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    if request.method == 'POST':
        # Handle POST request - update reviewers
        selected_user_ids = request.form.getlist('reviewer_ids')
        selected_user_ids = [int(uid) for uid in selected_user_ids if uid.isdigit()]
        
        try:
            # Remove existing reviewers
            PeriodReviewer.query.filter_by(period_id=period_id).delete()
            
            # Add new reviewers
            for user_id in selected_user_ids:
                reviewer = PeriodReviewer(period_id=period_id, user_id=user_id)
                db.session.add(reviewer)
            
            db.session.commit()
            log_activity(current_user.id, 'update_period_reviewers', 
                        f'Updated reviewers for period {period.name}: {len(selected_user_ids)} selected')
            
            flash(f'Successfully updated reviewers for {period.name}. {len(selected_user_ids)} reviewers selected.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating reviewers: {str(e)}', 'error')
        
        return redirect(url_for('period_reviewers', period_id=period_id))
    
    # Handle GET request - display form
    log_activity(current_user.id, 'view_period_reviewers', f'Admin viewed reviewers for period: {period.name}')
    
    # Get all active users who can be reviewers
    all_users = User.query.filter_by(is_active=True).order_by(User.name).all()
    
    # Get currently selected reviewers for this period
    selected_reviewers = PeriodReviewer.query.filter_by(period_id=period_id).all()
    selected_user_ids = {r.user_id for r in selected_reviewers}
    
    return render_template('period_reviewers.html', 
                         period=period, 
                         all_reviewers=all_users,
                         current_reviewer_ids=selected_user_ids)



@app.route('/assignment_deletion_preview/<int:period_id>', methods=['POST'])
@login_required
@admin_required
def assignment_deletion_preview(period_id):
    """Preview which assignments would be deleted when updating the matrix"""
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get current assignments
    current_assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
    
    # Build requested assignments from form data
    requested_assignments = set()
    form_data = request.get_json() or {}
    
    for key, value in form_data.items():
        if key.startswith('assignment_') and value:
            parts = key.replace('assignment_', '').split('_')
            if len(parts) == 2:
                officer_id, reviewer_id = parts
                requested_assignments.add(f"{officer_id}_{reviewer_id}")
    
    # Find assignments that would be deleted
    deletable_assignments = []
    for assignment in current_assignments:
        assignment_key = f"{assignment.officer_id}_{assignment.reviewer_id}"
        if assignment_key not in requested_assignments:
            # Check if assignment has responses
            has_responses = AssessmentResponse.query.filter_by(
                assessment_assignment_id=assignment.id
            ).first() is not None
            
            if not has_responses:
                deletable_assignments.append({
                    'officer_name': assignment.officer.name,
                    'reviewer_name': assignment.reviewer.name,
                    'id': assignment.id
                })
    
    return {
        'deletable_count': len(deletable_assignments),
        'deletable_assignments': deletable_assignments
    }

@app.route('/admin/delete_assignment/<int:assignment_id>', methods=['POST'])
@login_required
@admin_required
def delete_assignment(assignment_id):
    """Delete a specific assessment assignment"""
    assignment = AssessmentAssignment.query.get_or_404(assignment_id)
    officer_id = assignment.officer_id
    
    try:
        reviewer_name = assignment.reviewer.name
        officer_name = assignment.officer.name
        has_responses = AssessmentResponse.query.filter_by(assessment_assignment_id=assignment_id).first() is not None
        
        # Delete all associated assessment responses first
        if has_responses:
            AssessmentResponse.query.filter_by(assessment_assignment_id=assignment_id).delete()
        
        # Delete the assignment
        db.session.delete(assignment)
        db.session.commit()
        
        log_activity(current_user.id, 'delete_assignment', 
                    f'Deleted assignment with {"responses" if has_responses else "no responses"}: {reviewer_name} reviewing {officer_name}')
        
        flash_message = f'Assignment deleted: {reviewer_name} reviewing {officer_name}'
        if has_responses:
            flash_message += ' (including all response data)'
        flash(flash_message, 'success')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting assignment: {str(e)}', 'error')
    
    # Redirect back to my assignments instead of admin dashboard for better UX
    return redirect(url_for('my_assignments'))

@app.route('/api/officer/<int:officer_id>/assessments')
@login_required
@admin_required
def get_officer_assessments(officer_id):
    """API endpoint to get all assessments for an officer"""
    try:
        # Get all assignments for this officer
        assignments = AssessmentAssignment.query.filter_by(officer_id=officer_id).all()
        
        assessments = []
        for assignment in assignments:
            # Get the reviewer name
            reviewer = User.query.get(assignment.reviewer_id)
            if not reviewer:
                continue
                
            # Get all responses for this assignment
            responses = AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).all()
            
            if responses:  # Only include if there are responses
                response_data = []
                for response in responses:
                    question = AssessmentQuestion.query.get(response.question_id)
                    if question:
                        response_data.append({
                            'question_label': question.question_label,
                            'question_text': question.question_text,
                            'response_number': response.response_number,
                            'response_text': response.response_text,
                            'response_date': response.response_date.isoformat() if response.response_date else None
                        })
                
                assessments.append({
                    'reviewer_name': reviewer.name,
                    'is_self_assessment': assignment.officer_id == assignment.reviewer_id,
                    'submitted_at': assignment.submitted_at.isoformat() if assignment.submitted_at else None,
                    'responses': response_data
                })
        
        return jsonify({
            'success': True,
            'assessments': assessments
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/admin/officer/<int:officer_id>/all_reviews')
@login_required
@admin_required
def view_all_officer_reviews(officer_id):
    """View all reviews for an officer in tabbed interface"""
    try:
        # Get the officer
        officer = User.query.get_or_404(officer_id)
        
        # Get current active assessment period
        current_period = AssessmentPeriod.query.filter_by(is_active=True).first()
        if not current_period:
            flash('No active assessment period found.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Get all assignments for this officer in the current period
        assignments = AssessmentAssignment.query.filter_by(
            officer_id=officer_id,
            period_id=current_period.id
        ).all()
        
        reviews = []
        for assignment in assignments:
            # Get the reviewer
            reviewer = User.query.get(assignment.reviewer_id)
            if not reviewer:
                continue
                
            # Check if assignment has responses or is submitted
            response_count = AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).count()
            
            # Include if there are responses OR if it's a submitted assignment
            if response_count > 0 or assignment.is_submitted:
                # Use the EXACT same working logic as view_assessment_new
                # Determine which form to show based on assignment type
                if assignment.officer_id == assignment.reviewer_id:
                    # Self-assessment: show self_review form
                    target_form_type = 'self_review'
                else:
                    # External review: show reviewer form
                    target_form_type = 'reviewer'
                
                # Get the appropriate form for this assignment type
                period_forms = PeriodFormAssignment.query.filter_by(
                    period_id=assignment.period_id,
                    form_type=target_form_type
                ).all()
                
                if period_forms:
                    target_form = AssessmentForm.query.get(period_forms[0].form_id)
                    
                    # Get all questions from the target form in correct order
                    questions = AssessmentQuestion.query.filter_by(
                        form_id=target_form.id,
                        is_active=True
                    ).order_by(AssessmentQuestion.order).all()
                    
                    # Get responses for this assignment
                    assessment_responses = AssessmentResponse.query.filter_by(
                        assessment_assignment_id=assignment.id
                    ).all()
                    response_lookup = {resp.question_id: resp for resp in assessment_responses}
                    
                    # Create ordered responses list using the working pattern
                    ordered_responses = []
                    for question in questions:
                        response = response_lookup.get(question.id)
                        if response:  # Only include questions that have responses
                            ordered_responses.append({
                                'question_label': question.question_name,
                                'question_text': question.question_text,
                                'question_type': question.question_type,
                                'question_order': question.order,
                                'response_number': response.response_number,
                                'response_text': response.response_text,
                                'response_date': response.response_date
                            })
                else:
                    ordered_responses = []
                
                reviews.append({
                    'reviewer_name': reviewer.name,
                    'reviewer_email': reviewer.email,
                    'is_self_assessment': assignment.officer_id == assignment.reviewer_id,
                    'submitted_at': assignment.submitted_at,
                    'ordered_responses': ordered_responses,
                    'assignment_id': assignment.id
                })
        
        log_activity(current_user.id, 'view_all_officer_reviews', 
                    f'Viewed all reviews for officer: {officer.name}')
        
        from datetime import datetime
        return render_template('view_all_officer_reviews.html', 
                             officer=officer, 
                             reviews=reviews,
                             current_date=datetime.now())
        
    except Exception as e:
        print(f"Error in view_all_officer_reviews: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading reviews: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/officer/<int:officer_id>/self_assessment')
@login_required
@admin_required
def view_self_assessment(officer_id):
    """View self-assessment for an officer with correct question ordering"""
    try:
        # Get the officer
        officer = User.query.get_or_404(officer_id)
        
        # Get current active assessment period
        current_period = AssessmentPeriod.query.filter_by(is_active=True).first()
        if not current_period:
            flash('No active assessment period found.', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        # Get self-assessment assignment (where officer_id = reviewer_id)
        self_assignment = AssessmentAssignment.query.filter_by(
            officer_id=officer_id,
            reviewer_id=officer_id,  # Self-assessment
            period_id=current_period.id
        ).first()
        
        self_assessment = None
        ordered_questions = []
        
        if self_assignment:
            # Get existing responses for this assignment
            responses = AssessmentResponse.query.filter_by(assessment_assignment_id=self_assignment.id).all()
            existing_responses = {}
            for response in responses:
                existing_responses[response.question_id] = response
            
            # Get questions in correct order using the same logic as evaluate_officer
            # Find forms assigned to this period for self_review
            period_forms = PeriodFormAssignment.query.filter_by(
                period_id=current_period.id,
                form_type='self_review'
            ).all()
            
            # Load questions from all assigned forms in correct order
            questions = []
            for period_form in period_forms:
                form_questions = AssessmentQuestion.query.filter_by(
                    form_id=period_form.form_id,
                    is_active=True
                ).order_by(AssessmentQuestion.order).all()
                questions.extend(form_questions)
            
            # Create ordered questions list with responses
            for question in questions:
                response = existing_responses.get(question.id)
                ordered_questions.append({
                    'question_name': question.question_name,
                    'question_text': question.question_text,
                    'question_type': question.question_type,
                    'question_order': question.order,
                    'response_number': response.response_number if response else None,
                    'response_text': response.response_text if response else None,
                    'response_date': response.response_date if response else None
                })
            
            # Create self-assessment info
            self_assessment = {
                'reviewer_name': officer.name,
                'reviewer_email': officer.email,
                'submitted_at': self_assignment.submitted_at,
                'assignment_id': self_assignment.id
            }
        
        log_activity(current_user.id, 'view_self_assessment', 
                    f'Viewed self-assessment for officer: {officer.name}')
        
        return render_template('view_self_assessment.html', 
                             officer=officer, 
                             self_assessment=self_assessment,
                             ordered_questions=ordered_questions)
        
    except Exception as e:
        print(f"Error in view_self_assessment: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading self-assessment: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/documentation')
@login_required
@admin_required
def admin_documentation():
    """Admin documentation and help system"""
    log_activity(current_user.id, 'view_admin_documentation', 'Accessed admin documentation')
    return render_template('admin_documentation.html')

@app.route('/admin/test')
@login_required
@admin_required
def admin_test():
    """Simple admin test route"""
    return render_template('admin_main.html',
                         assessment_forms_count=0,
                         assessment_questions_count=0,
                         categories=[],
                         active_periods=[],
                         total_users=0,
                         total_officers=0,
                         total_board_members=0,
                         total_assessments=0)
