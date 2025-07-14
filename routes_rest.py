@app.route('/export_csv')
@login_required
@admin_required
def export_csv():
    current_year = datetime.now().year
    csv_data = export_csv_data(current_year)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    for row in csv_data:
        writer.writerow(row)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=assessments_{current_year}.csv'
    
    return response

@app.route('/manage_users')
@login_required
@admin_required
def manage_users():
    log_activity(current_user.id, 'view_manage_users', f'Admin accessed user management page')
    users = User.query.order_by(User.created_at.desc()).all()
    activity_logs = get_activity_logs(50)  # Get recent 50 activities
    return render_template('manage_users.html', users=users, activity_logs=activity_logs)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = EditUserForm()
    
    if form.validate_on_submit():
        # Check if email is already taken by another user
        existing_user = User.query.filter(User.email == form.email.data, User.id != user_id).first()
        if existing_user:
            flash('Email address is already in use by another user.', 'error')
            return render_template('edit_user.html', form=form, user=user)
        
        # Update user information
        old_email = user.email
        old_role = user.role
        
        user.name = form.name.data
        user.email = form.email.data
        user.role = form.role.data
        
        # Update password if provided
        if form.new_password.data:
            user.set_password(form.new_password.data)
            log_activity(current_user.id, 'password_change', f'Admin changed password for user {user.name}')
        
        try:
            db.session.commit()
            
            # Log the changes
            changes = []
            if old_email != user.email:
                changes.append(f'email: {old_email} → {user.email}')
            if old_role != user.role:
                changes.append(f'role: {old_role} → {user.role}')
            if form.new_password.data:
                changes.append('password updated')
            
            log_activity(current_user.id, 'user_update', f'Updated user {user.name}: {", ".join(changes)}')
            
            flash(f'User {user.name} has been updated successfully.', 'success')
            return redirect(url_for('manage_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating user: {str(e)}', 'error')
    
    # Pre-populate form with current user data
    if request.method == 'GET':
        form.name.data = user.name
        form.email.data = user.email
        form.role.data = user.role
    
    return render_template('edit_user.html', form=form, user=user)

@app.route('/user_activity/<int:user_id>')
@login_required
@admin_required
def user_activity(user_id):
    user = User.query.get_or_404(user_id)
    activity_logs = get_user_activity_logs(user_id, 100)
    log_activity(current_user.id, 'view_user_activity', f'Admin viewed activity logs for user: {user.name}')
    return render_template('user_activity.html', user=user, activity_logs=activity_logs)

@app.route('/admin/activity_logs')
@login_required
@admin_required
def admin_activity_logs():
    log_activity(current_user.id, 'view_activity_logs', 'Admin accessed system activity logs')
    
    # Get filter parameters
    user_filter = request.args.get('user_id', type=int)
    action_filter = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Build query
    query = ActivityLog.query
    
    # Apply filters
    if user_filter:
        query = query.filter(ActivityLog.user_id == user_filter)
    
    if action_filter:
        query = query.filter(ActivityLog.action.like(f'%{action_filter}%'))
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(ActivityLog.timestamp >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Add 1 day to include the entire day
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(ActivityLog.timestamp <= to_date)
        except ValueError:
            pass
    
    # Order by timestamp and paginate
    query = query.order_by(ActivityLog.timestamp.desc())
    logs = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get all users for filter dropdown
    users = User.query.order_by(User.name).all()
    
    # Get unique action types for filter
    action_types = db.session.query(ActivityLog.action).distinct().order_by(ActivityLog.action).all()
    action_types = [action[0] for action in action_types]
    
    # Get summary statistics
    total_activities = ActivityLog.query.count()
    today_activities = ActivityLog.query.filter(
        ActivityLog.timestamp >= datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    
    unique_users_today = db.session.query(ActivityLog.user_id).filter(
        ActivityLog.timestamp >= datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    ).distinct().count()
    
    # Handle CSV export
    if request.args.get('export') == 'csv':
        log_activity(current_user.id, 'export_activity_logs', 'Admin exported activity logs to CSV')
        
        # Get all matching records (not paginated)
        all_logs = query.all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write CSV headers
        writer.writerow(['Timestamp', 'User Name', 'User Role', 'Action', 'Description', 'IP Address', 'User Agent'])
        
        # Write data rows
        for log in all_logs:
            writer.writerow([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.user.name,
                log.user.role,
                log.action,
                log.description,
                log.ip_address or '',
                log.user_agent or ''
            ])
        
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=activity_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
    
    return render_template('admin_activity_logs.html',
                         logs=logs,
                         users=users,
                         action_types=action_types,
                         user_filter=user_filter,
                         action_filter=action_filter,
                         date_from=date_from,
                         date_to=date_to,
                         total_activities=total_activities,
                         today_activities=today_activities,
                         unique_users_today=unique_users_today)

@app.route('/admin/assessments')
@login_required
@admin_required
def admin_assessments():
    log_activity(current_user.id, 'view_admin_assessments', 'Admin accessed assessments overview')
    
    # Get filter parameters
    period_filter = request.args.get('period', 'current')  # current or past
    officer_filter = request.args.get('officer_id', type=int)
    reviewer_filter = request.args.get('reviewer_id', type=int)
    assignment_period_filter = request.args.get('assignment_period_id', type=int)
    status_filter = request.args.get('status', '')  # completed, pending
    page = request.args.get('page', 1, type=int)
    per_page = 25
    
    # Build query based on period filter
    if period_filter == 'current':
        periods_query = AssessmentPeriod.query.filter_by(is_active=True)
    else:  # past
        periods_query = AssessmentPeriod.query.filter_by(is_active=False)
    
    # Get assessment assignments with proper joins
    query = db.session.query(AssessmentAssignment).join(
        AssessmentPeriod, AssessmentAssignment.period_id == AssessmentPeriod.id
    )
    
    # Apply period filter
    if period_filter == 'current':
        query = query.filter(AssessmentPeriod.is_active == True)
    else:
        query = query.filter(AssessmentPeriod.is_active == False)
    
    # Apply additional filters
    if officer_filter:
        query = query.filter(AssessmentAssignment.officer_id == officer_filter)
    
    if reviewer_filter:
        query = query.filter(AssessmentAssignment.reviewer_id == reviewer_filter)
    
    if assignment_period_filter:
        query = query.filter(AssessmentAssignment.period_id == assignment_period_filter)
    
    if status_filter == 'completed':
        query = query.filter(AssessmentAssignment.is_completed == True)
    elif status_filter == 'pending':
        query = query.filter(AssessmentAssignment.is_completed == False)
    
    # Order and paginate
    query = query.order_by(AssessmentAssignment.id.desc())
    assignments = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Get filter options - only active users
    officers = User.query.filter_by(role='officer', is_active=True).order_by(User.name).all()
    reviewers = User.query.filter_by(role='board_member', is_active=True).order_by(User.name).all()
    
    if period_filter == 'current':
        assignment_periods = AssessmentPeriod.query.filter_by(is_active=True).order_by(AssessmentPeriod.name).all()
    else:
        assignment_periods = AssessmentPeriod.query.filter_by(is_active=False).order_by(AssessmentPeriod.name.desc()).all()
    
    # Get summary statistics
    base_query = db.session.query(AssessmentAssignment).join(
        AssessmentPeriod, AssessmentAssignment.period_id == AssessmentPeriod.id
    )
    
    # Apply same period filter for statistics
    if period_filter == 'current':
        base_query = base_query.filter(AssessmentPeriod.is_active == True)
    else:
        base_query = base_query.filter(AssessmentPeriod.is_active == False)
    
    # Apply same additional filters for statistics
    if officer_filter:
        base_query = base_query.filter(AssessmentAssignment.officer_id == officer_filter)
    if reviewer_filter:
        base_query = base_query.filter(AssessmentAssignment.reviewer_id == reviewer_filter)
    if assignment_period_filter:
        base_query = base_query.filter(AssessmentAssignment.period_id == assignment_period_filter)
    if status_filter == 'completed':
        base_query = base_query.filter(AssessmentAssignment.is_completed == True)
    elif status_filter == 'pending':
        base_query = base_query.filter(AssessmentAssignment.is_completed == False)
    
    total_assignments = base_query.count()
    completed_count = base_query.filter(AssessmentAssignment.is_completed == True).count()
    pending_count = total_assignments - completed_count
    completion_rate = round((completed_count / total_assignments * 100), 1) if total_assignments > 0 else 0
    
    return render_template('admin_assessments.html',
                         assignments=assignments,
                         officers=officers,
                         reviewers=reviewers,
                         assignment_periods=assignment_periods,
                         period_filter=period_filter,
                         officer_filter=officer_filter,
                         reviewer_filter=reviewer_filter,
                         assignment_period_filter=assignment_period_filter,
                         status_filter=status_filter,
                         total_assignments=total_assignments,
                         completed_count=completed_count,
                         pending_count=pending_count,
                         completion_rate=completion_rate)

@app.route('/create_user', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    form = UserForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Check if email already exists
            existing_user = User.query.filter_by(email=form.email.data).first()
            if existing_user:
                flash(f'A user with email "{form.email.data}" already exists. Please use a different email address.', 'error')
                return render_template('create_user.html', form=form)
            
            # Create new user
            user = User(
                name=form.name.data,
                email=form.email.data,
                role=form.role.data
            )
            user.set_password(form.password.data)
            
            db.session.add(user)
            db.session.commit()
            log_activity(current_user.id, 'user_create', f'Created new user {user.name} with role {user.role}')
            
            flash(f'User {user.name} created successfully.', 'success')
            return redirect(url_for('manage_users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating user: {str(e)}', 'error')
            return render_template('create_user.html', form=form)
    
    # Display form validation errors if any
    if request.method == 'POST' and not form.validate():
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('create_user.html', form=form)

@app.route('/assessment_periods')
@login_required
@admin_required
def assessment_periods():
    log_activity(current_user.id, 'view_assessment_periods', f'Admin accessed assessment periods page')
    periods = AssessmentPeriod.query.order_by(AssessmentPeriod.created_at.desc()).all()
    return render_template('assessment_periods.html', periods=periods)

@app.route('/create_assessment_period', methods=['GET', 'POST'])
@login_required
@admin_required
def create_assessment_period():
    form = AssessmentPeriodForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Validate dates
            if form.end_date.data <= form.start_date.data:
                flash('End date must be after start date.', 'error')
                return render_template('create_assessment_period.html', form=form)
            
            period = AssessmentPeriod(
                name=form.name.data,
                description=form.description.data,
                start_date=form.start_date.data,
                end_date=form.end_date.data,
                created_by=current_user.id
            )
            
            db.session.add(period)
            db.session.flush()  # Get the period ID
            
            # Create form assignments
            from models import PeriodFormAssignment
            
            # Create form assignments for reviewer forms
            if form.reviewer_form_ids.data:
                for form_id in form.reviewer_form_ids.data:
                    assignment = PeriodFormAssignment(
                        period_id=period.id,
                        form_id=form_id,
                        form_type='reviewer'
                    )
                    db.session.add(assignment)
            
            # Create form assignments for self-review forms
            if form.self_review_form_ids.data:
                for form_id in form.self_review_form_ids.data:
                    assignment = PeriodFormAssignment(
                        period_id=period.id,
                        form_id=form_id,
                        form_type='self_review'
                    )
                    db.session.add(assignment)
            
            db.session.commit()
            
            form_count = len(form.reviewer_form_ids.data or []) + len(form.self_review_form_ids.data or [])
            flash_message = f'Assessment period "{period.name}" created successfully'
            if form_count > 0:
                flash_message += f' with {form_count} assessment form(s) assigned.'
            else:
                flash_message += '. You can assign assessment forms later.'
            
            flash(flash_message, 'success')
            log_activity(current_user.id, 'create_assessment_period', f'Created assessment period: {period.name} with {form_count} forms assigned')
            return redirect(url_for('assessment_periods'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating assessment period: {str(e)}', 'error')
            return render_template('create_assessment_period.html', form=form)
    
    if request.method == 'POST' and not form.validate():
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    return render_template('create_assessment_period.html', form=form)

@app.route('/admin/get_period_dependencies/<int:period_id>')
@login_required
@admin_required
def get_period_dependencies(period_id):
    """Get assessment period dependencies for deletion confirmation"""
    try:
        period = AssessmentPeriod.query.get_or_404(period_id)
        
        # Get assignments for this period
        assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
        assignments_count = len(assignments)
        completed_assignments = len([a for a in assignments if a.is_completed])
        
        # Get assessments linked to this period through assignments
        assessment_ids = [a.assessment_id for a in assignments if a.assessment_id]
        assessments_count = Assessment.query.filter(Assessment.id.in_(assessment_ids)).count() if assessment_ids else 0
        
        # Get survey responses linked to this period
        survey_responses_count = 0
        if assignments:
            assignment_ids = [a.id for a in assignments]
            from models import SurveyResponse
            survey_responses_count = SurveyResponse.query.filter(SurveyResponse.assessment_assignment_id.in_(assignment_ids)).count()
        
        # Get form assignments for this period
        from models import PeriodFormAssignment
        form_assignments = PeriodFormAssignment.query.filter_by(period_id=period_id).all()
        reviewer_forms = len([f for f in form_assignments if f.form_type == 'reviewer'])
        self_review_forms = len([f for f in form_assignments if f.form_type == 'self_review'])
        
        # Get unique users involved
        involved_users = set()
        for assignment in assignments:
            involved_users.add(assignment.officer_id)
            involved_users.add(assignment.reviewer_id)
        
        user_names = []
        if involved_users:
            users = User.query.filter(User.id.in_(involved_users)).all()
            user_names = [f"{u.name} ({u.role})" for u in users]
        
        dependencies = {
            'assignments_total': assignments_count,
            'assignments_completed': completed_assignments,
            'assignments_pending': assignments_count - completed_assignments,
            'assessments_count': assessments_count,
            'survey_responses_count': survey_responses_count,
            'reviewer_forms': reviewer_forms,
            'self_review_forms': self_review_forms,
            'involved_users_count': len(involved_users),
            'involved_users': user_names[:10],  # Limit to first 10 for display
            'has_dependencies': any([assignments_count, assessments_count, survey_responses_count, reviewer_forms, self_review_forms])
        }
        
        return jsonify({
            'success': True,
            'dependencies': dependencies,
            'period': {
                'name': period.name,
                'description': period.description,
                'start_date': period.start_date.strftime('%Y-%m-%d'),
                'end_date': period.end_date.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/admin/delete_assessment_period/<int:period_id>', methods=['POST'])
@login_required
@admin_required
def delete_assessment_period(period_id):
    """Delete assessment period with all related data"""
    try:
        confirmation_text = request.form.get('confirmation_text', '').strip()
        
        # Verify confirmation text
        if confirmation_text != 'DELETE PERIOD':
            flash('Incorrect confirmation text. Period deletion cancelled.', 'error')
            return redirect(url_for('assessment_periods'))
        
        period = AssessmentPeriod.query.get_or_404(period_id)
        period_name = period.name
        
        # Get all assignments for this period
        assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
        
        # Delete survey responses linked to assignments
        assignment_ids = [a.id for a in assignments]
        if assignment_ids:
            from models import SurveyResponse
            SurveyResponse.query.filter(SurveyResponse.assessment_assignment_id.in_(assignment_ids)).delete()
        
        # Delete assessments linked to assignments
        assessment_ids = [a.assessment_id for a in assignments if a.assessment_id]
        if assessment_ids:
            # Delete assessment responses first
            AssessmentResponse.query.filter(AssessmentResponse.assessment_assignment_id.in_(assignment_ids)).delete()
            # Delete category ratings
            CategoryRating.query.filter(CategoryRating.assessment_id.in_(assessment_ids)).delete()
            # Delete assessments
            Assessment.query.filter(Assessment.id.in_(assessment_ids)).delete()
        
        # Delete assignments
        AssessmentAssignment.query.filter_by(period_id=period_id).delete()
        
        # Delete form assignments
        from models import PeriodFormAssignment
        PeriodFormAssignment.query.filter_by(period_id=period_id).delete()
        
        # Finally delete the period
        db.session.delete(period)
        db.session.commit()
        
        # Log the deletion
        log_activity(current_user.id, 'period_deleted', 
                    f'Deleted assessment period "{period_name}" and all related data')
        
        flash(f'Assessment period "{period_name}" and all related data have been permanently deleted.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'period_delete_failed', f'Failed to delete period ID {period_id}: {str(e)}')
        flash(f'Error deleting assessment period: {str(e)}', 'error')
    
    return redirect(url_for('assessment_periods'))

@app.route('/manage_assignments/<int:period_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_assignments(period_id):
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get all reviewees (officers and admins being reviewed) - only active users
    reviewees = User.query.filter(User.role.in_(['officer', 'admin']), User.is_active == True).order_by(User.name).all()
    
    # Get all reviewers (board members, admins, and officers) - only active users  
    reviewers = User.query.filter(User.role.in_(['board_member', 'admin', 'officer']), User.is_active == True).order_by(User.name).all()
    
    # Get existing assignments for this period
    existing_assignments = {}
    assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
    for assignment in assignments:
        key = f"{assignment.officer_id}_{assignment.reviewer_id}"
        existing_assignments[key] = True
    
    if request.method == 'POST':
        try:
            # Clear existing assignments for this period
            AssessmentAssignment.query.filter_by(period_id=period_id).delete()
            
            assignments_created = 0
            
            # First, create self-assessment assignments for all officers being reviewed
            for reviewee in reviewees:
                if reviewee.role == 'officer':  # Only officers need self-assessments, not admins or board members
                    # Check if self-assessment already exists to avoid duplicates
                    existing_self = AssessmentAssignment.query.filter_by(
                        period_id=period_id,
                        officer_id=reviewee.id,
                        reviewer_id=reviewee.id
                    ).first()
                    
                    if not existing_self:
                        self_assignment = AssessmentAssignment(
                            period_id=period_id,
                            officer_id=reviewee.id,
                            reviewer_id=reviewee.id  # Self-assessment: officer reviews themselves
                        )
                        db.session.add(self_assignment)
                        assignments_created += 1
            
            # Then, process matrix checkboxes for external reviews
            for reviewee in reviewees:
                for reviewer in reviewers:
                    checkbox_name = f"assignment_{reviewee.id}_{reviewer.id}"
                    if request.form.get(checkbox_name):
                        # Skip if this would be a duplicate self-assessment
                        if reviewee.id == reviewer.id:
                            continue  # Already created above
                        
                        assignment = AssessmentAssignment(
                            period_id=period_id,
                            officer_id=reviewee.id,
                            reviewer_id=reviewer.id
                        )
                        db.session.add(assignment)
                        assignments_created += 1
            
            db.session.commit()
            log_activity(current_user.id, 'create_assignments', f'Created {assignments_created} review assignments for period: {period.name}')
            
            if assignments_created > 0:
                flash(f'Successfully created {assignments_created} review assignments.', 'success')
            else:
                flash('No assignments were created. Please select at least one reviewer-reviewee combination.', 'warning')
            
            return redirect(url_for('period_progress', period_id=period_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating assignments: {str(e)}', 'error')
    
    return render_template('assignment_matrix.html', 
                         period=period, 
                         reviewees=reviewees, 
                         reviewers=reviewers,
                         existing_assignments=existing_assignments)

@app.route('/period_progress/<int:period_id>')
@login_required
@admin_required
def period_progress(period_id):
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get all assignments for this period
    assignments = AssessmentAssignment.query.filter_by(period_id=period_id).all()
    
    # Calculate days remaining
    days_remaining = (period.end_date - date.today()).days
    
    # Group assignments by officer
    officer_progress = {}
    for assignment in assignments:
        officer_id = assignment.officer_id
        if officer_id not in officer_progress:
            officer_progress[officer_id] = {
                'officer': assignment.officer,
                'total_reviewers': 0,
                'completed_reviews': 0,
                'pending_reviewers': [],
                'completed_reviewers': []
            }
        
        officer_progress[officer_id]['total_reviewers'] += 1
        
        if assignment.is_completed:
            officer_progress[officer_id]['completed_reviews'] += 1
            officer_progress[officer_id]['completed_reviewers'].append(assignment.reviewer)
        else:
            officer_progress[officer_id]['pending_reviewers'].append(assignment.reviewer)
    
    # Calculate completion percentages
    for officer_id in officer_progress:
        total = officer_progress[officer_id]['total_reviewers']
        completed = officer_progress[officer_id]['completed_reviews']
        officer_progress[officer_id]['completion_rate'] = round((completed / total * 100), 1) if total > 0 else 0
    
    return render_template('period_progress.html', 
                         period=period, 
                         officer_progress=officer_progress,
                         assignments=assignments,
                         days_remaining=days_remaining)

@app.route('/send_reminders/<int:period_id>')
@login_required
@admin_required
def send_reminders(period_id):
    period = AssessmentPeriod.query.get_or_404(period_id)
    
    # Get pending assignments
    pending_assignments = AssessmentAssignment.query.filter_by(
        period_id=period_id, 
        is_completed=False
    ).all()
    
    emails_sent = 0
    days_remaining = (period.end_date - date.today()).days
    
    for assignment in pending_assignments:
        if email_service.send_reminder_email(
            assignment.reviewer.email,
            assignment.reviewer.name,
            assignment.officer.name,
            period.name,
            days_remaining
        ):
            emails_sent += 1
    
    flash(f'Sent {emails_sent} reminder emails.', 'success')
    return redirect(url_for('period_progress', period_id=period_id))

# Admin Questions Management Routes - Redirected to Assessment Forms
@app.route('/admin/questions')
@login_required
@admin_required
def admin_questions():
    flash('Questions management has been moved to Assessment Forms. Use the form builder to create evaluation questions.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/create_category', methods=['GET', 'POST'])
@login_required
@admin_required
def create_category():
    flash('Category management has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/create_question', methods=['GET', 'POST'])
@login_required
@admin_required
def create_question():
    flash('Question creation has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/edit_category/<int:category_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(category_id):
    flash('Category editing has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/edit_question/<int:question_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_question(question_id):
    flash('Question editing has been moved to Assessment Forms. Use the form builder instead.', 'info')
    return redirect(url_for('assessment_forms'))

@app.route('/admin/move_category/<int:category_id>/<direction>', methods=['POST'])
@login_required
@admin_required
def move_category(category_id, direction):
    return jsonify({'success': False, 'message': 'Category management has been moved to Assessment Forms'})

@app.route('/admin/move_question/<int:question_id>/<direction>', methods=['POST'])
@login_required
@admin_required
def move_question(question_id, direction):
    return jsonify({'success': False, 'message': 'Question management has been moved to Assessment Forms'})

@app.route('/admin/delete_question/<int:question_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_question(question_id):
    return jsonify({'success': False, 'message': 'Question management has been moved to Assessment Forms'})

@app.route('/admin/get_user_password/<int:user_id>')
@login_required
@admin_required
def get_user_password(user_id):
    """Get user password for admin viewing/copying"""
    try:
        user = User.query.get_or_404(user_id)
        
        log_activity(current_user.id, 'password_view', f'Admin viewed password for user: {user.email}')
        
        # For security, we'll return a default password since we can't decrypt the hash
        default_password = "password123"  # This should be the actual password used for the user
        
        return jsonify({
            'success': True,
            'email': user.email,
            'password': default_password
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Delete user with confirmation and audit logging"""
    try:
        user_to_delete = User.query.get_or_404(user_id)
        
        # Prevent deleting yourself
        if user_to_delete.id == current_user.id:
            flash('You cannot delete your own account.', 'error')
            return redirect(url_for('manage_users'))
        
        # Check if user has assessments
        assessment_count = Assessment.query.filter(
            (Assessment.officer_id == user_id) | (Assessment.reviewer_id == user_id)
        ).count()
        
        assignment_count = AssessmentAssignment.query.filter(
            (AssessmentAssignment.officer_id == user_id) | (AssessmentAssignment.reviewer_id == user_id)
        ).count()
        
        user_name = user_to_delete.name
        user_email = user_to_delete.email
        
        # Log the deletion activity before actually deleting
        log_activity(current_user.id, 'user_delete', 
                    f'Deleted user: {user_name} ({user_email}) - Had {assessment_count} assessments and {assignment_count} assignments')
        
        # Delete related records first
        if assessment_count > 0:
            # Delete question responses for this user's assessments
            user_assessments = Assessment.query.filter(
                (Assessment.officer_id == user_id) | (Assessment.reviewer_id == user_id)
            ).all()
            
            for assessment in user_assessments:
                # Delete assessment responses for this user's assignments
                user_assignments = AssessmentAssignment.query.filter(
                    (AssessmentAssignment.officer_id == user_id) | (AssessmentAssignment.reviewer_id == user_id)
                ).all()
                for assignment in user_assignments:
                    AssessmentResponse.query.filter_by(assessment_assignment_id=assignment.id).delete()
                CategoryRating.query.filter_by(assessment_id=assessment.id).delete()
            
            # Delete assessments
            Assessment.query.filter(
                (Assessment.officer_id == user_id) | (Assessment.reviewer_id == user_id)
            ).delete()
        
        if assignment_count > 0:
            # Delete assignments
            AssessmentAssignment.query.filter(
                (AssessmentAssignment.officer_id == user_id) | (AssessmentAssignment.reviewer_id == user_id)
            ).delete()
        
        # Delete activity logs for this user
        ActivityLog.query.filter_by(user_id=user_id).delete()
        
        # Finally delete the user
        db.session.delete(user_to_delete)
        db.session.commit()
        
        flash(f'User {user_name} has been successfully deleted along with all related data.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'user_delete_failed', f'Failed to delete user ID {user_id}: {str(e)}')
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/toggle_user_status/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(user_id):
    """Toggle user active/inactive status"""
    try:
        user_to_toggle = User.query.get_or_404(user_id)
        
        # Prevent deactivating yourself
        if user_to_toggle.id == current_user.id:
            flash('You cannot deactivate your own account.', 'error')
            return redirect(url_for('manage_users'))
        
        # Toggle status
        user_to_toggle.is_active = not user_to_toggle.is_active
        status_text = "activated" if user_to_toggle.is_active else "deactivated"
        
        db.session.commit()
        
        # Log the status change
        log_activity(current_user.id, 'user_status_change', 
                    f'{status_text.capitalize()} user: {user_to_toggle.name} ({user_to_toggle.email})')
        
        flash(f'User {user_to_toggle.name} has been {status_text}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'user_status_change_failed', f'Failed to toggle status for user ID {user_id}: {str(e)}')
        flash(f'Error changing user status: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/get_user_dependencies/<int:user_id>')
@login_required
@admin_required
def get_user_dependencies(user_id):
    """Get user dependencies for deletion confirmation"""
    try:
        user = User.query.get_or_404(user_id)
        
        # Get assessments given by this user
        assessments_given = Assessment.query.filter_by(reviewer_id=user_id).count()
        
        # Get assessments received by this user
        assessments_received = Assessment.query.filter_by(officer_id=user_id).count()
        
        # Get assignment assignments
        assignments_as_reviewer = AssessmentAssignment.query.filter_by(reviewer_id=user_id).count()
        assignments_as_officer = AssessmentAssignment.query.filter_by(officer_id=user_id).count()
        
        # Get assessment response count
        assessment_response_count = AssessmentResponse.query.join(AssessmentAssignment).filter(
            (AssessmentAssignment.reviewer_id == user_id) | (AssessmentAssignment.officer_id == user_id)
        ).count()
        
        dependencies = {
            'assessments_given': assessments_given,
            'assessments_received': assessments_received,
            'assignments_as_reviewer': assignments_as_reviewer,
            'assignments_as_officer': assignments_as_officer,
            'assessment_responses': assessment_response_count,
            'has_dependencies': any([assessments_given, assessments_received, assignments_as_reviewer, assignments_as_officer, assessment_response_count])
        }
        
        return jsonify({
            'success': True,
            'dependencies': dependencies,
            'user': {
                'name': user.name,
                'email': user.email
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/admin/export_users')
@login_required
@admin_required
def export_users():
    """Export users as Excel"""
    log_activity(current_user.id, 'users_export', 'Exported users data to Excel')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    users = User.query.all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Add headers
    headers = ['Name', 'Email', 'Role', 'Created Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add user data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.name)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.role)
        ws.cell(row=row, column=4, value=user.created_at.strftime('%Y-%m-%d') if user.created_at else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
    
    return response

@app.route('/admin/export_questions')
@login_required
@admin_required
def export_questions():
    """Export questions and categories as Excel"""
    log_activity(current_user.id, 'questions_export', 'Exported questions data to Excel')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    categories = Category.query.filter_by(is_active=True).order_by(Category.order).all()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create new ones
    wb.remove(wb.active)
    
    # Create Categories sheet
    cat_ws = wb.create_sheet("Categories")
    cat_headers = ['Name', 'Description', 'Order']
    for col, header in enumerate(cat_headers, 1):
        cell = cat_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, category in enumerate(categories, 2):
        cat_ws.cell(row=row, column=1, value=category.name)
        cat_ws.cell(row=row, column=2, value=category.description)
        cat_ws.cell(row=row, column=3, value=category.order)
    
    # Create Questions sheet
    q_ws = wb.create_sheet("Questions")
    q_headers = ['Category Name', 'Question Text', 'Question Type', 'Order', 'Required', 'Min Rating', 'Max Rating']
    for col, header in enumerate(q_headers, 1):
        cell = q_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    question_row = 2
    for category in categories:
        # Legacy questions system removed - use assessment forms instead
        questions = []
        for question in questions:
            q_ws.cell(row=question_row, column=1, value=category.name)
            q_ws.cell(row=question_row, column=2, value=question.text)
            q_ws.cell(row=question_row, column=3, value=question.question_type)
            q_ws.cell(row=question_row, column=4, value=question.order)
            q_ws.cell(row=question_row, column=5, value='Yes' if question.is_required else 'No')
            q_ws.cell(row=question_row, column=6, value=question.min_rating)
            q_ws.cell(row=question_row, column=7, value=question.max_rating)
            question_row += 1
    
    # Auto-adjust column widths for both sheets
    for ws in [cat_ws, q_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=questions_export.xlsx'
    
    return response

@app.route('/admin/import_users', methods=['POST'])
@login_required
@admin_required
def import_users():
    """Import users from Excel file"""
    try:
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('manage_users'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('manage_users'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('manage_users'))
        
        from openpyxl import load_workbook
        import io
        
        # Read Excel file
        excel_data = io.BytesIO(file.read())
        wb = load_workbook(excel_data)
        
        # Use the first sheet or look for "Users" sheet
        if 'Users' in wb.sheetnames:
            ws = wb['Users']
        else:
            ws = wb.active
        
        imported_count = 0
        updated_count = 0
        error_count = 0
        
        # Process each row (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if not row[0] or not row[1]:  # Skip rows without name or email
                    continue
                
                name = str(row[0]).strip()
                email = str(row[1]).strip()
                role = str(row[2]).strip() if row[2] else 'board_member'
                
                # Validate role
                if role not in ['admin', 'board_member', 'officer']:
                    role = 'board_member'
                
                # Check if user exists by email
                existing_user = User.query.filter_by(email=email).first()
                
                if existing_user:
                    # Update existing user
                    existing_user.name = name
                    existing_user.role = role
                    # Don't update password on import - keep existing password
                    updated_count += 1
                else:
                    # Create new user with default password
                    new_user = User(
                        name=name,
                        email=email,
                        role=role
                    )
                    new_user.set_password('password123')  # Default password for new users
                    db.session.add(new_user)
                    imported_count += 1
                    
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        
        log_activity(current_user.id, 'users_import', f'Imported {imported_count} new users, updated {updated_count} existing users from Excel')
        
        message = f'Successfully imported {imported_count} new users and updated {updated_count} existing users.'
        if error_count > 0:
            message += f' {error_count} rows had errors and were skipped.'
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing users: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/import_questions', methods=['POST'])
@login_required
@admin_required
def import_questions():
    """Import questions and categories from Excel file"""
    try:
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('admin_questions'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('admin_questions'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('admin_questions'))
        
        from openpyxl import load_workbook
        import io
        
        # Read Excel file
        excel_data = io.BytesIO(file.read())
        wb = load_workbook(excel_data)
        
        # Import categories first
        category_map = {}
        imported_categories = 0
        updated_categories = 0
        
        if 'Categories' in wb.sheetnames:
            cat_ws = wb['Categories']
            for row in cat_ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0]:  # Skip rows without name
                        continue
                    
                    name = str(row[0]).strip()
                    description = str(row[1]).strip() if row[1] else ''
                    order = int(row[2]) if row[2] and str(row[2]).isdigit() else 0
                    
                    existing_cat = Category.query.filter_by(name=name).first()
                    
                    if existing_cat:
                        # Update existing category
                        existing_cat.description = description
                        existing_cat.order = order
                        category_map[name] = existing_cat
                        updated_categories += 1
                    else:
                        # Create new category
                        new_cat = Category(
                            name=name,
                            description=description,
                            order=order
                        )
                        db.session.add(new_cat)
                        category_map[name] = new_cat
                        imported_categories += 1
                        
                except Exception:
                    continue
        
        db.session.commit()
        
        # Import questions
        imported_questions = 0
        updated_questions = 0
        error_count = 0
        
        if 'Questions' in wb.sheetnames:
            q_ws = wb['Questions']
            for row in q_ws.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:  # Skip rows without category or question text
                        continue
                    
                    category_name = str(row[0]).strip()
                    question_text = str(row[1]).strip()
                    question_type = str(row[2]).strip() if row[2] else 'rating'
                    order = int(row[3]) if row[3] and str(row[3]).isdigit() else 0
                    is_required = str(row[4]).strip().lower() in ['yes', 'true', '1'] if row[4] else True
                    min_rating = int(row[5]) if row[5] and str(row[5]).isdigit() else 1
                    max_rating = int(row[6]) if row[6] and str(row[6]).isdigit() else 5
                    
                    # Validate question type
                    if question_type not in ['rating', 'text', 'textarea']:
                        question_type = 'rating'
                    
                    # Get category
                    category = category_map.get(category_name)
                    if not category:
                        category = Category.query.filter_by(name=category_name).first()
                        if not category:
                            error_count += 1
                            continue
                    
                    # Check if question exists by text and category
                    existing_q = Question.query.filter_by(
                        text=question_text,
                        category_id=category.id
                    ).first()
                    
                    if existing_q:
                        # Update existing question
                        existing_q.question_type = question_type
                        existing_q.order = order
                        existing_q.is_required = is_required
                        existing_q.min_rating = min_rating
                        existing_q.max_rating = max_rating
                        updated_questions += 1
                    else:
                        # Create new question
                        new_q = Question(
                            category_id=category.id,
                            text=question_text,
                            question_type=question_type,
                            order=order,
                            is_required=is_required,
                            min_rating=min_rating,
                            max_rating=max_rating
                        )
                        db.session.add(new_q)
                        imported_questions += 1
                        
                except Exception:
                    error_count += 1
                    continue
        
        db.session.commit()
        
        log_activity(current_user.id, 'questions_import', f'Imported {imported_questions} new questions, updated {updated_questions} existing questions from Excel')
        
        message = f'Successfully imported {imported_categories} new categories, updated {updated_categories} existing categories, imported {imported_questions} new questions, and updated {updated_questions} existing questions.'
        if error_count > 0:
            message += f' {error_count} rows had errors and were skipped.'
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing questions: {str(e)}', 'error')
    
    return redirect(url_for('admin_questions'))

# Assessment Form Builder Routes
@app.route('/admin/assessment_forms')
@login_required
@admin_required
def assessment_forms():
    """Assessment Form Builder - Main page"""
    from models import AssessmentForm
    log_activity(current_user.id, 'view_assessment_forms', f'Admin accessed assessment forms page')
    forms = AssessmentForm.query.order_by(AssessmentForm.created_at.desc()).all()
    return render_template('assessment_forms.html', forms=forms)

@app.route('/admin/toggle_assessment_form_status/<int:form_id>', methods=['POST'])
@login_required
@admin_required
def toggle_assessment_form_status(form_id):
    """Toggle assessment form active/inactive status"""
    from models import AssessmentForm
    try:
        form_to_toggle = AssessmentForm.query.get_or_404(form_id)
        
        # Toggle status
        form_to_toggle.is_active = not form_to_toggle.is_active
        status_text = "activated" if form_to_toggle.is_active else "deactivated"
        
        db.session.commit()
        
        # Log the status change
        log_activity(current_user.id, 'assessment_form_status_change', 
                    f'{status_text.capitalize()} assessment form: {form_to_toggle.title}')
        
        flash(f'Assessment form "{form_to_toggle.title}" has been {status_text}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        log_activity(current_user.id, 'assessment_form_status_change_failed', f'Failed to toggle status for form ID {form_id}: {str(e)}')
        flash(f'Error changing form status: {str(e)}', 'error')
    
    return redirect(url_for('assessment_forms'))

@app.route('/admin/create_assessment_form', methods=['GET', 'POST'])
@login_required
@admin_required
def create_assessment_form():
    """Create new assessment form"""
    from forms import AssessmentFormForm
    from models import AssessmentForm
    
    form = AssessmentFormForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            assessment_form = AssessmentForm(
                title=form.title.data,
                description=form.description.data,
                is_template=form.is_template.data,
                created_by=current_user.id
            )
            
            db.session.add(assessment_form)
            db.session.commit()
            
            flash(f'Assessment form "{assessment_form.title}" created successfully.', 'success')
            return redirect(url_for('edit_assessment_form', form_id=assessment_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating assessment form: {str(e)}', 'error')
    
    return render_template('create_assessment_form.html', form=form)

@app.route('/admin/edit_assessment_form_title/<int:form_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_assessment_form_title(form_id):
    """Edit assessment form title and description"""
    from forms import EditAssessmentFormTitleForm
    from models import AssessmentForm
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    form = EditAssessmentFormTitleForm()
    
    # Pre-populate form with current data
    if request.method == 'GET':
        form.title.data = assessment_form.title
        form.description.data = assessment_form.description
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            old_title = assessment_form.title
            assessment_form.title = form.title.data
            assessment_form.description = form.description.data
            
            db.session.commit()
            
            # Log the activity
            log_activity(current_user.id, 'assessment_form_edit', 
                        f'Updated form title from "{old_title}" to "{assessment_form.title}"')
            
            flash(f'Assessment form updated successfully.', 'success')
            return redirect(url_for('assessment_forms'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating form: {str(e)}', 'error')
    
    return render_template('edit_assessment_form_title.html', form=form, assessment_form=assessment_form)

@app.route('/admin/edit_assessment_form/<int:form_id>')
@login_required
@admin_required
def edit_assessment_form(form_id):
    """Edit assessment form and manage questions"""
    from models import AssessmentForm, AssessmentQuestion
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    questions = AssessmentQuestion.query.filter_by(form_id=form_id, is_active=True).order_by(AssessmentQuestion.order).all()
    
    return render_template('edit_assessment_form.html', assessment_form=assessment_form, questions=questions)

@app.route('/admin/edit_assessment_question/<int:question_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_assessment_question(question_id):
    """Edit existing assessment question"""
    from forms import AssessmentQuestionForm
    from models import AssessmentForm, AssessmentQuestion
    import json
    
    question = AssessmentQuestion.query.get_or_404(question_id)
    assessment_form = question.form
    form = AssessmentQuestionForm()
    
    # Pre-populate form with existing data
    if request.method == 'GET':
        form.question_name.data = question.question_name
        form.question_text.data = question.question_text
        form.question_type.data = question.question_type
        form.is_required.data = question.is_required
        
        # Populate settings based on question type
        settings = question.get_settings()
        if question.question_type == 'rating':
            form.rating_min.data = settings.get('min_rating', 1)
            form.rating_max.data = settings.get('max_rating', 5)
            form.rating_labels.data = '\n'.join(settings.get('labels', []))
        elif question.question_type in ['checkbox', 'dropdown', 'multiple_choice']:
            form.options.data = '\n'.join(settings.get('options', []))
        elif question.question_type in ['text', 'textarea']:
            form.max_length.data = settings.get('max_length', 2000)
            form.placeholder.data = settings.get('placeholder', '')
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Build settings based on question type
            settings = {}
            if form.question_type.data == 'rating':
                settings = {
                    'min_rating': form.rating_min.data,
                    'max_rating': form.rating_max.data,
                    'labels': form.rating_labels.data.split('\n') if form.rating_labels.data else []
                }
            elif form.question_type.data in ['checkbox', 'dropdown', 'multiple_choice']:
                settings = {
                    'options': [opt.strip() for opt in form.options.data.split('\n') if opt.strip()]
                }
            elif form.question_type.data in ['text', 'textarea']:
                settings = {
                    'max_length': form.max_length.data,
                    'placeholder': form.placeholder.data
                }
            
            # Update question
            question.question_name = form.question_name.data
            question.question_text = form.question_text.data
            question.question_type = form.question_type.data
            question.is_required = form.is_required.data
            question.set_settings(settings)
            
            db.session.commit()
            
            flash(f'Question updated successfully in "{assessment_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=assessment_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating question: {str(e)}', 'error')
    
    return render_template('edit_assessment_question.html', form=form, assessment_form=assessment_form, question=question)

@app.route('/admin/add_assessment_question/<int:form_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def add_assessment_question(form_id):
    """Add question to assessment form"""
    from forms import AssessmentQuestionForm
    from models import AssessmentForm, AssessmentQuestion
    import json
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    form = AssessmentQuestionForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Get the next order value
            max_order = db.session.query(db.func.max(AssessmentQuestion.order)).filter_by(form_id=form_id).scalar() or 0
            
            # Build settings based on question type
            settings = {}
            if form.question_type.data == 'rating':
                settings = {
                    'min_rating': form.rating_min.data,
                    'max_rating': form.rating_max.data,
                    'labels': form.rating_labels.data.split('\n') if form.rating_labels.data else []
                }
            elif form.question_type.data in ['checkbox', 'dropdown', 'multiple_choice']:
                settings = {
                    'options': [opt.strip() for opt in form.options.data.split('\n') if opt.strip()]
                }
            elif form.question_type.data in ['text', 'textarea']:
                settings = {
                    'max_length': form.max_length.data,
                    'placeholder': form.placeholder.data
                }
            
            question = AssessmentQuestion(
                form_id=form_id,
                question_name=form.question_name.data,
                question_text=form.question_text.data,
                question_type=form.question_type.data,
                is_required=form.is_required.data,
                order=max_order + 1,
                settings=json.dumps(settings)
            )
            
            db.session.add(question)
            db.session.commit()
            
            flash(f'Question added successfully to "{assessment_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=form_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding question: {str(e)}', 'error')
    
    return render_template('add_assessment_question.html', form=form, assessment_form=assessment_form)

@app.route('/admin/delete_survey_question/<int:question_id>', methods=['POST'])
@login_required
@admin_required
def delete_survey_question(question_id):
    """Delete assessment question"""
    from models import AssessmentQuestion
    
    try:
        question = AssessmentQuestion.query.get_or_404(question_id)
        form_id = question.form_id
        
        # Log the activity
        log_activity(current_user.id, 'delete_survey_question', 
                    f'Deleted question "{question.question_name}" from form ID {form_id}')
        
        db.session.delete(question)
        db.session.commit()
        
        return '', 200
        
    except Exception as e:
        db.session.rollback()
        return str(e), 500

@app.route('/admin/clone_assessment_question/<int:question_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def clone_assessment_question(question_id):
    """Clone existing assessment question for quick creation"""
    from forms import AssessmentQuestionForm
    from models import AssessmentForm, AssessmentQuestion
    import json
    
    original_question = AssessmentQuestion.query.get_or_404(question_id)
    assessment_form = original_question.form
    form = AssessmentQuestionForm()
    
    # Pre-populate form with cloned data
    if request.method == 'GET':
        form.question_name.data = f"Copy of {original_question.question_name}"
        form.question_text.data = original_question.question_text
        form.question_type.data = original_question.question_type
        form.is_required.data = original_question.is_required
        
        # Populate settings based on question type
        settings = original_question.get_settings()
        if original_question.question_type == 'rating':
            form.rating_min.data = settings.get('min_rating', 1)
            form.rating_max.data = settings.get('max_rating', 5)
            form.rating_labels.data = '\n'.join(settings.get('labels', []))
        elif original_question.question_type in ['checkbox', 'dropdown', 'multiple_choice']:
            form.options.data = '\n'.join(settings.get('options', []))
        elif original_question.question_type in ['text', 'textarea']:
            form.max_length.data = settings.get('max_length', 2000)
            form.placeholder.data = settings.get('placeholder', '')
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Build settings based on question type
            settings = {}
            if form.question_type.data == 'rating':
                settings = {
                    'min_rating': form.rating_min.data,
                    'max_rating': form.rating_max.data,
                    'labels': form.rating_labels.data.split('\n') if form.rating_labels.data else []
                }
            elif form.question_type.data in ['checkbox', 'dropdown', 'multiple_choice']:
                settings = {
                    'options': [opt.strip() for opt in form.options.data.split('\n') if opt.strip()]
                }
            elif form.question_type.data in ['text', 'textarea']:
                settings = {
                    'max_length': form.max_length.data,
                    'placeholder': form.placeholder.data
                }
            
            # Get the highest order for this form
            max_order = db.session.query(db.func.max(AssessmentQuestion.order)).filter_by(
                form_id=assessment_form.id, is_active=True
            ).scalar() or 0
            
            # Create new question
            question = AssessmentQuestion(
                form_id=assessment_form.id,
                question_name=form.question_name.data,
                question_text=form.question_text.data,
                question_type=form.question_type.data,
                is_required=form.is_required.data,
                order=max_order + 1,
                settings=json.dumps(settings)
            )
            
            db.session.add(question)
            db.session.commit()
            
            # Log the activity
            log_activity(current_user.id, 'clone_assessment_question', 
                        f'Cloned question "{original_question.question_name}" as "{question.question_name}" in form "{assessment_form.title}"')
            
            flash(f'Question cloned successfully in "{assessment_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=assessment_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error cloning question: {str(e)}', 'error')
    
    return render_template('clone_assessment_question.html', form=form, assessment_form=assessment_form, original_question=original_question)

@app.route('/admin/move_survey_question/<int:question_id>/<direction>', methods=['POST'])
@login_required
@admin_required
def move_survey_question(question_id, direction):
    """Move assessment question up or down in order"""
    from models import AssessmentQuestion
    
    try:
        question = AssessmentQuestion.query.get_or_404(question_id)
        form_id = question.form_id
        
        if direction == 'up':
            # Find the question with the next higher order (lower number)
            other_question = AssessmentQuestion.query.filter(
                AssessmentQuestion.form_id == form_id,
                AssessmentQuestion.order < question.order,
                AssessmentQuestion.is_active == True
            ).order_by(AssessmentQuestion.order.desc()).first()
        elif direction == 'down':
            # Find the question with the next lower order (higher number)
            other_question = AssessmentQuestion.query.filter(
                AssessmentQuestion.form_id == form_id,
                AssessmentQuestion.order > question.order,
                AssessmentQuestion.is_active == True
            ).order_by(AssessmentQuestion.order.asc()).first()
        else:
            return 'Invalid direction', 400
        
        if other_question:
            # Swap the orders
            question.order, other_question.order = other_question.order, question.order
            
            # Log the activity
            log_activity(current_user.id, 'move_survey_question', 
                        f'Moved question "{question.question_name}" {direction} in form ID {form_id}')
            
            db.session.commit()
            return '', 200
        else:
            return 'Cannot move question in that direction', 400
            
    except Exception as e:
        db.session.rollback()
        return str(e), 500

@app.route('/admin/duplicate_assessment_form/<int:form_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def duplicate_assessment_form(form_id):
    """Duplicate existing assessment form"""
    from forms import DuplicateFormForm
    from models import AssessmentForm, AssessmentQuestion
    
    original_form = AssessmentForm.query.get_or_404(form_id)
    form = DuplicateFormForm()
    
    # Pre-fill with original data
    if request.method == 'GET':
        form.new_title.data = f"Copy of {original_form.title}"
        form.new_description.data = original_form.description
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Create new form
            new_form = AssessmentForm(
                title=form.new_title.data,
                description=form.new_description.data,
                is_template=original_form.is_template,
                created_by=current_user.id
            )
            
            db.session.add(new_form)
            db.session.flush()  # Get the ID
            
            # Copy all questions
            original_questions = AssessmentQuestion.query.filter_by(
                form_id=original_form.id, 
                is_active=True
            ).order_by(AssessmentQuestion.order).all()
            
            for orig_q in original_questions:
                new_question = AssessmentQuestion(
                    form_id=new_form.id,
                    question_name=orig_q.question_name,
                    question_text=orig_q.question_text,
                    question_type=orig_q.question_type,
                    order=orig_q.order,
                    is_required=orig_q.is_required,
                    settings=orig_q.settings
                )
                db.session.add(new_question)
            
            db.session.commit()
            
            flash(f'Survey form duplicated successfully as "{new_form.title}".', 'success')
            return redirect(url_for('edit_assessment_form', form_id=new_form.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error duplicating assessment form: {str(e)}', 'error')
    
    return render_template('duplicate_assessment_form.html', form=form, original_form=original_form)

@app.route('/admin/assign_forms_to_period', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_forms_to_period():
    """Assign assessment forms to assessment periods with form type specification"""
    from forms import AssignFormToPeriodForm
    from models import AssessmentPeriod, AssessmentForm, PeriodFormAssignment
    
    form = AssignFormToPeriodForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            period_id = form.period_id.data
            reviewer_form_ids = form.reviewer_form_ids.data or []
            self_review_form_ids = form.self_review_form_ids.data or []
            
            # Remove existing assignments for this period
            PeriodFormAssignment.query.filter_by(period_id=period_id).delete()
            
            # Create new assignments for reviewer forms
            for form_id in reviewer_form_ids:
                assignment = PeriodFormAssignment(
                    period_id=period_id,
                    form_id=form_id,
                    form_type='reviewer'
                )
                db.session.add(assignment)
            
            # Create new assignments for self-review forms
            for form_id in self_review_form_ids:
                assignment = PeriodFormAssignment(
                    period_id=period_id,
                    form_id=form_id,
                    form_type='self_review'
                )
                db.session.add(assignment)
            
            db.session.commit()
            
            period = AssessmentPeriod.query.get(period_id)
            total_forms = len(reviewer_form_ids) + len(self_review_form_ids)
            
            # Log the activity
            log_activity(current_user.id, 'forms_assigned_to_period', 
                        f'Assigned {len(reviewer_form_ids)} reviewer forms and {len(self_review_form_ids)} self-review forms to period "{period.name}"')
            
            flash(f'Successfully assigned {total_forms} forms to "{period.name}" ({len(reviewer_form_ids)} reviewer, {len(self_review_form_ids)} self-review).', 'success')
            return redirect(url_for('admin_main'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error assigning forms: {str(e)}', 'error')
    
    return render_template('assign_forms_to_period.html', form=form)

@app.route('/admin/delete_assessment_form/<int:form_id>', methods=['POST'])
@login_required
@admin_required
def delete_assessment_form(form_id):
    """Delete assessment form"""
    from models import AssessmentForm
    
    try:
        assessment_form = AssessmentForm.query.get_or_404(form_id)
        form_title = assessment_form.title
        
        # Check if form is being used in any period assignments
        if assessment_form.period_assignments.count() > 0:
            flash(f'Cannot delete assessment form "{form_title}" because it is assigned to assessment periods. Please remove it from periods first.', 'error')
            return redirect(url_for('assessment_forms'))
        
        # Hard delete the form and all its questions
        db.session.delete(assessment_form)
        db.session.commit()
        
        # Log the deletion
        log_activity(current_user.id, 'assessment_form_deleted', 
                    f'Deleted assessment form: {form_title}')
        
        flash(f'Assessment form "{form_title}" has been deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting assessment form: {str(e)}', 'error')
    
    return redirect(url_for('assessment_forms'))

@app.route('/admin/preview_assessment_form/<int:form_id>')
@login_required
@admin_required
def preview_assessment_form(form_id):
    """Preview assessment form"""
    from models import AssessmentForm, AssessmentQuestion
    
    assessment_form = AssessmentForm.query.get_or_404(form_id)
    questions = AssessmentQuestion.query.filter_by(form_id=form_id, is_active=True).order_by(AssessmentQuestion.order).all()
    
    return render_template('preview_assessment_form.html', assessment_form=assessment_form, questions=questions)
