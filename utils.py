from functools import wraps
from flask import flash, redirect, url_for
from flask_login import current_user
from app import db
from models import User, Category, Assessment, CategoryRating
from werkzeug.security import generate_password_hash
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Access denied. Administrator privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def initialize_default_data():
    """Initialize default categories and admin user if they don't exist"""
    
    # Create default categories if they don't exist
    if Category.query.count() == 0:
        categories = [
            {
                'name': 'Financial',
                'description': 'Key financial responsibilities include presenting an acceptable annual budget, achieving consistent and diversified revenue growth, maintaining and improving profitability, ensuring positive returns on investments, managing cash flow, controlling costs, providing accurate and timely financial reports.',
                'order': 1
            },
            {
                'name': 'Strategy Development',
                'description': 'Leads creation of AAAWCNY\'s strategic direction and effectively implement long-term strategies that meet the needs of the organization, associates, members and other stakeholders.',
                'order': 2
            },
            {
                'name': 'Talent Development, Organizational Structure & Culture',
                'description': 'Develops and leads a strong executive team to drive short and long-term results to meet strategic goals and make critical, timely decisions about acquisitions, mergers, product line expansions and personnel. Develops future leaders together with succession plans for long term succession and sudden loss of leadership. Creates a positive work environment for associates and promotes a culture reflecting AAAWCNY\'s mission, values, and business strategies. Promotes organizational visibility and positive reputation in the communities served.',
                'order': 3
            },
            {
                'name': 'Operations',
                'description': 'Oversees the efficient and effective operations of all departments and business lines ensuring positive operational performance and member satisfaction. Ensures adequate internal controls are in place and effectively assesses, manages and mitigates risk. Anticipates, assesses and responds to market trends and disruptions, risks, and opportunities that impact the success of AAAWCNY\'s strategy. Creates successful marketing strategies that support the growth and diversification of revenues. Ensures that investments in technology (e.g. SalesForce, NetSuite) meet or exceed the ROI and related benefits set forth in the business cases presented to the Board for approval.',
                'order': 4
            },
            {
                'name': 'Board & External Relations',
                'description': 'Forges collaborative, trusting, and transparent working relationships with the Board of Directors. Engages with the Board of Directors in establishing short-term objectives and long-term goals with regular updates on the status of operations. Works with the board and committee chairs to ensure meetings are productive. Demonstrates community leadership and actively engages with the business community.',
                'order': 5
            }
        ]
        
        for cat_data in categories:
            category = Category(**cat_data)
            db.session.add(category)
        
        db.session.commit()
        print("Default categories created.")
    
    # Create default admin user if no admin exists
    if User.query.filter_by(role='admin').count() == 0:
        admin_user = User(
            name='System Administrator',
            email='admin@aaaperformance.com',
            role='admin'
        )
        admin_user.set_password('admin123')  # Change this in production
        db.session.add(admin_user)
        db.session.commit()
        print("Default admin user created: admin@aaaperformance.com / admin123")

def generate_pdf_report(officer, assessments):
    """Generate PDF report for an officer's assessments"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    story.append(Paragraph(f"Performance Assessment Report - {officer.name}", title_style))
    story.append(Paragraph(f"Assessment Year: {assessments[0].year}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary Statistics
    story.append(Paragraph("Assessment Summary", styles['Heading2']))
    
    # Calculate overall statistics
    total_assessments = len(assessments)
    overall_ratings = [a.overall_rating for a in assessments if a.overall_rating]
    avg_overall = round(sum(overall_ratings) / len(overall_ratings), 2) if overall_ratings else 0
    
    summary_data = [
        ['Total Assessments Received:', str(total_assessments)],
        ['Average Overall Rating:', f"{avg_overall}/5.0"],
        ['Assessment Period:', f"January - December {assessments[0].year}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Category Averages
    story.append(Paragraph("Category Performance", styles['Heading2']))
    
    categories = Category.query.order_by(Category.order).all()
    category_data = [['Category', 'Average Rating', 'Description']]
    
    for category in categories:
        ratings = []
        for assessment in assessments:
            rating = CategoryRating.query.filter_by(
                assessment_id=assessment.id,
                category_id=category.id
            ).first()
            if rating:
                ratings.append(rating.rating)
        
        avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else 0
        category_data.append([
            category.name,
            f"{avg_rating}/5.0",
            category.description[:100] + "..." if len(category.description) > 100 else category.description
        ])
    
    category_table = Table(category_data, colWidths=[1.5*inch, 1*inch, 3.5*inch])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(category_table)
    story.append(Spacer(1, 20))
    
    # Individual Assessment Details
    story.append(Paragraph("Individual Assessment Details", styles['Heading2']))
    
    for i, assessment in enumerate(assessments, 1):
        story.append(Paragraph(f"Assessment #{i} - {assessment.reviewer.name}", styles['Heading3']))
        story.append(Paragraph(f"Submitted: {assessment.submitted_at.strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Paragraph(f"Overall Rating: {assessment.overall_rating}/5.0", styles['Normal']))
        story.append(Spacer(1, 10))
        
        if assessment.accomplishments:
            story.append(Paragraph("Key Accomplishments:", styles['Heading4']))
            story.append(Paragraph(assessment.accomplishments, styles['Normal']))
            story.append(Spacer(1, 10))
        
        if assessment.improvement_opportunities:
            story.append(Paragraph("Areas for Improvement:", styles['Heading4']))
            story.append(Paragraph(assessment.improvement_opportunities, styles['Normal']))
            story.append(Spacer(1, 10))
        
        if assessment.focus_for_next_year:
            story.append(Paragraph("Focus for Next Year:", styles['Heading4']))
            story.append(Paragraph(assessment.focus_for_next_year, styles['Normal']))
            story.append(Spacer(1, 10))
        
        story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.read()

def export_csv_data(year):
    """Export assessment data as CSV"""
    categories = Category.query.order_by(Category.order).all()
    
    # Create header row
    header = ['Officer Name', 'Reviewer Name', 'Submission Date', 'Overall Rating']
    for category in categories:
        header.append(f'{category.name} Rating')
    header.extend(['Accomplishments', 'Improvement Opportunities', 'Focus for Next Year'])
    
    rows = [header]
    
    # Get all assessments for the year
    assessments = Assessment.query.filter_by(year=year).order_by(Assessment.officer_id, Assessment.submitted_at).all()
    
    for assessment in assessments:
        row = [
            assessment.officer.name,
            assessment.reviewer.name,
            assessment.submitted_at.strftime('%Y-%m-%d %H:%M'),
            str(assessment.overall_rating) if assessment.overall_rating else ''
        ]
        
        # Add category ratings
        for category in categories:
            rating = CategoryRating.query.filter_by(
                assessment_id=assessment.id,
                category_id=category.id
            ).first()
            row.append(str(rating.rating) if rating else '')
        
        # Add text feedback
        row.extend([
            assessment.accomplishments or '',
            assessment.improvement_opportunities or '',
            assessment.focus_for_next_year or ''
        ])
        
        rows.append(row)
    
    return rows

def generate_matrix_pdf_report(officer, matrix_data, reviewers, overall_matrix_average, overall_ai_summary=None):
    """Generate PDF report for matrix view - exactly matching web matrix format with AI summaries"""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io
    
    buffer = io.BytesIO()
    
    # Use landscape orientation for better table fit
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                          topMargin=0.4*inch, bottomMargin=0.4*inch,
                          leftMargin=0.4*inch, rightMargin=0.4*inch)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles for compact layout
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=4,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Header
    title = Paragraph(f"{officer.name} - Performance Matrix", title_style)
    subtitle = Paragraph(f"Overall Average: {overall_matrix_average}/5 | {len(reviewers)} Reviewers | AI-Powered Analysis", subtitle_style)
    
    story.append(title)
    story.append(subtitle)
    
    # Build table data - exactly matching web format
    table_data = []
    
    # Header row - matching web table exactly
    header_row = ['Question']
    reviewer_list = list(reviewers.values())
    for reviewer in reviewer_list:
        # Use shorter name format for PDF headers to prevent overlap
        name_parts = reviewer.name.split()
        if len(name_parts) > 1:
            short_name = f"{name_parts[0]}\n{name_parts[-1]}"
        else:
            short_name = reviewer.name
        header_row.append(short_name)
    header_row.append('Average')
    header_row.append('AI Summary')
    
    table_data.append(header_row)
    
    # Question rows - with full data including AI summaries
    for question_data in matrix_data:
        question = question_data['question']
        
        # Use simple question numbering since no categories in new structure
        cat_short = f'Q{len(table_data)}'
        
        # Truncate question for PDF but keep readable
        question_text = question.question_text
        if len(question_text) > 45:
            question_text = question_text[:42] + "..."
        
        question_cell = f"{question_text}\n({cat_short})"
        row = [question_cell]
        
        # Reviewer ratings - exact format
        for reviewer in reviewer_list:
            response = question_data['responses'].get(reviewer.name)
            if response and response.get('rating'):
                row.append(str(response['rating']))
            else:
                row.append("—")
        
        # Question average
        avg = question_data['average_rating']
        row.append(str(avg) if avg > 0 else "—")
        
        # AI Summary - full text with paragraph formatting
        ai_analysis = question_data.get('ai_analysis', {})
        ai_summary = ai_analysis.get('summary', 'Analysis pending...')
        # Keep full AI summary text for PDF - no truncation
        from reportlab.platypus import Paragraph
        ai_paragraph = Paragraph(ai_summary, styles['Normal'])
        row.append(ai_paragraph)
        
        table_data.append(row)
    
    # Total Average Row - exactly like web
    total_row = ['TOTAL AVERAGE']
    
    # Calculate individual reviewer averages
    for reviewer in reviewer_list:
        reviewer_ratings = []
        for question_data in matrix_data:
            response = question_data['responses'].get(reviewer.name)
            if response and response.get('rating'):
                reviewer_ratings.append(response['rating'])
        
        if reviewer_ratings:
            reviewer_avg = round(sum(reviewer_ratings) / len(reviewer_ratings), 1)
            total_row.append(str(reviewer_avg))
        else:
            total_row.append("—")
    
    # Overall average
    total_row.append(str(overall_matrix_average))
    # Simple text for total row - full summary will be separate section
    total_row.append("See Overall Analysis Below")
    
    table_data.append(total_row)
    
    # Calculate column widths for complete matrix format - optimized for reviewer names
    page_width = landscape(letter)[0] - 0.8*inch  # Available width
    question_col_width = page_width * 0.28  # 28% for question
    reviewer_count = len(reviewer_list)
    # Wider columns for reviewer names with smaller font
    reviewer_col_width = 1.2*inch if reviewer_count > 0 else 1.0*inch  # Fixed width per reviewer
    avg_col_width = page_width * 0.08  # 8% for average
    # Calculate remaining width for AI summary
    used_width = question_col_width + (reviewer_col_width * reviewer_count) + avg_col_width
    ai_col_width = max(page_width - used_width, page_width * 0.30)  # At least 30% for AI
    
    col_widths = [question_col_width] + [reviewer_col_width] * reviewer_count + [avg_col_width, ai_col_width]
    
    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Table styling - matching web design
    table_style = [
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        
        # Smaller font for reviewer name columns to prevent overlap
        ('FONTSIZE', (1, 0), (reviewer_count, 0), 8),
        ('FONTNAME', (1, 0), (reviewer_count, 0), 'Helvetica-Bold'),
        
        # Question column alignment
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # AI Summary column alignment
        ('ALIGN', (-1, 1), (-1, -1), 'LEFT'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        
        # Total row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 8),
    ]
    
    table.setStyle(TableStyle(table_style))
    story.append(table)
    
    # Overall AI Summary Section - separate from table like web page
    story.append(Spacer(1, 0.3*inch))
    
    summary_title_style = ParagraphStyle(
        'SummaryTitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#2c3e50')
    )
    
    summary_text_style = ParagraphStyle(
        'SummaryText',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        alignment=TA_LEFT,
        leftIndent=0.2*inch
    )
    
    story.append(Paragraph("Overall Performance Analysis", summary_title_style))
    story.append(Paragraph(overall_ai_summary or "Overall analysis pending...", summary_text_style))
    
    # Legend - matching web format
    story.append(Spacer(1, 0.2*inch))
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    legend = Paragraph("Rating Scale: 5=Outstanding • 4=Exceeds Expectations • 3=Meets Expectations • 2=Below Expectations • 1=Unsatisfactory", legend_style)
    story.append(legend)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_matrix_excel_report(officer, matrix_data, reviewers, overall_matrix_average, overall_ai_summary=None):
    """Generate Excel report for matrix view - exactly matching web matrix format with AI summaries"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{officer.name} - Performance Matrix"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='000000')
    title_font = Font(name='Arial', size=16, bold=True, color='2c3e50')
    content_font = Font(name='Arial', size=10, color='000000')
    small_font = Font(name='Arial', size=9, color='000000')
    
    header_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    total_fill = PatternFill(start_color='e9ecef', end_color='e9ecef', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title and subtitle
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{officer.name} - Performance Matrix"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Overall Average: {overall_matrix_average}/5 | {len(reviewers)} Reviewers | AI-Powered Analysis"
    ws['A2'].font = content_font
    ws['A2'].alignment = center_align
    
    # Header row (starting at row 4)
    row = 4
    headers = ['Question']
    reviewer_list = list(reviewers.values())
    
    for reviewer in reviewer_list:
        headers.append(reviewer.name)
    headers.extend(['Average', 'AI Summary'])
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    
    # Question rows
    for question_data in matrix_data:
        question = question_data['question']
        
        # Use simple question numbering since no categories in new structure
        cat_short = f'Q{row-4}'
        
        question_text = f"{question.question_text}\n({cat_short})"
        
        col = 1
        cell = ws.cell(row=row, column=col, value=question_text)
        cell.font = content_font
        cell.alignment = left_align
        cell.border = thin_border
        
        # Reviewer ratings
        col += 1
        for reviewer in reviewer_list:
            response = question_data['responses'].get(reviewer.name)
            rating = None
            
            # Safe handling of response data
            if response is not None:
                if isinstance(response, dict):
                    rating = response.get('rating')
                elif isinstance(response, (int, float)):
                    # Handle case where response is stored as direct rating value
                    rating = response
                else:
                    rating = None
            
            # Display rating or empty indicator
            display_value = rating if rating is not None else "—"
            cell = ws.cell(row=row, column=col, value=display_value)
            cell.font = content_font
            cell.alignment = center_align
            cell.border = thin_border
            
            # Color coding for ratings
            if rating is not None and isinstance(rating, (int, float)):
                if rating >= 4:
                    cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                elif rating <= 2:
                    cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
            
            col += 1
        
        # Average
        avg_score = question_data.get('average_score', 0)
        cell = ws.cell(row=row, column=col, value=round(avg_score, 1) if avg_score else "—")
        cell.font = content_font
        cell.alignment = center_align
        cell.border = thin_border
        col += 1
        
        # AI Summary (safe conversion)
        ai_analysis = question_data.get('ai_analysis', {})
        if isinstance(ai_analysis, dict):
            ai_summary = ai_analysis.get('summary', 'Analysis pending...')
            if not isinstance(ai_summary, str):
                ai_summary = str(ai_summary)
        else:
            ai_summary = str(ai_analysis) if ai_analysis else 'Analysis pending...'
        
        # Ensure text length is within Excel limits
        if len(ai_summary) > 32000:
            ai_summary = ai_summary[:32000] + "..."
            
        cell = ws.cell(row=row, column=col, value=ai_summary)
        cell.font = small_font
        cell.alignment = left_align
        cell.border = thin_border
        
        row += 1
    
    # Total Average Row
    col = 1
    cell = ws.cell(row=row, column=col, value="TOTAL AVERAGE")
    cell.font = header_font
    cell.fill = total_fill
    cell.alignment = center_align
    cell.border = thin_border
    
    # Individual reviewer averages
    col += 1
    for reviewer in reviewer_list:
        reviewer_ratings = []
        for question_data in matrix_data:
            response = question_data['responses'].get(reviewer.name)
            if response is not None:
                rating = None
                if isinstance(response, dict):
                    rating = response.get('rating')
                elif isinstance(response, (int, float)):
                    rating = response
                else:
                    rating = None
                
                if rating is not None and isinstance(rating, (int, float)):
                    reviewer_ratings.append(rating)
        
        if reviewer_ratings:
            reviewer_avg = round(sum(reviewer_ratings) / len(reviewer_ratings), 1)
            cell = ws.cell(row=row, column=col, value=reviewer_avg)
        else:
            cell = ws.cell(row=row, column=col, value="—")
        
        cell.font = header_font
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border
        col += 1
    
    # Overall average
    cell = ws.cell(row=row, column=col, value=overall_matrix_average)
    cell.font = header_font
    cell.fill = total_fill
    cell.alignment = center_align
    cell.border = thin_border
    col += 1
    
    # Total row AI summary indicator
    cell = ws.cell(row=row, column=col, value="See Overall Analysis Below")
    cell.font = header_font
    cell.fill = total_fill
    cell.alignment = center_align
    cell.border = thin_border
    
    row += 2
    
    # Overall Performance Analysis section
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="Overall Performance Analysis")
    cell.font = Font(name='Arial', size=14, bold=True, color='2c3e50')
    cell.alignment = left_align
    
    row += 1
    ws.merge_cells(f'A{row}:F{row+10}')  # Merge for multi-line summary
    
    # Format overall AI summary as text (safe handling)
    if overall_ai_summary and isinstance(overall_ai_summary, dict):
        exec_summary = overall_ai_summary.get('executive_summary', 'Analysis pending...')
        if not isinstance(exec_summary, str):
            exec_summary = str(exec_summary)
            
        themes = overall_ai_summary.get('major_themes', [])
        if not isinstance(themes, list):
            themes = [str(themes)]
        
        sentiment = overall_ai_summary.get('overall_sentiment', 'Under review')
        if not isinstance(sentiment, str):
            sentiment = str(sentiment)
            
        summary_text = f"""Executive Summary:
{exec_summary}

Major Themes: {', '.join(themes)}

Performance Level: {sentiment.title()}"""
    else:
        summary_text = "Overall analysis pending..."
    
    # Ensure summary text length is within Excel limits
    if len(summary_text) > 32000:
        summary_text = summary_text[:32000] + "..."
    
    cell = ws.cell(row=row, column=1, value=summary_text)
    cell.font = content_font
    cell.alignment = left_align
    
    # Column width adjustments
    ws.column_dimensions['A'].width = 35  # Question column
    for i, reviewer in enumerate(reviewer_list, 2):
        ws.column_dimensions[get_column_letter(i)].width = 15  # Reviewer columns
    ws.column_dimensions[get_column_letter(len(reviewer_list) + 2)].width = 10  # Average column
    ws.column_dimensions[get_column_letter(len(reviewer_list) + 3)].width = 50  # AI Summary column
    
    # Row height adjustments
    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 30
    
    # Save to BytesIO buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
